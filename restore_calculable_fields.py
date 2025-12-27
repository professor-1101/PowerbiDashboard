#!/usr/bin/env python3
"""
Restore Calculable Fields with MOCK Data
User requested to bring back fields that are calculable with MOCK data
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random
from datetime import datetime, timedelta

print("=" * 80)
print("بازگرداندن فیلدهای قابل محاسبه با MOCK Data")
print("=" * 80)

# Load workbook
wb = load_workbook('BugTracking_Complete_FINAL.xlsx')
ws_raw = wb['raw_data']

# Get current headers
headers = []
for col in range(1, ws_raw.max_column + 1):
    headers.append(ws_raw.cell(1, col).value)

print(f"\n📊 فیلدهای فعلی: {len(headers)}")

# Define calculable fields to add back
# NOTE: AnalysisEffortHrs is NOT included - user confirmed it should be deleted
fields_to_add = [
    'IsDuplicate',          # Boolean - if CloseReason is "Duplicate"
    'DuplicateOfBugID',     # Number - which bug is this a duplicate of
    'CloseReason',          # Text - Fixed, Duplicate, By Design, Won't Fix, etc.
    'LeadTimeHrs',          # Number - CreatedDate to ClosedDate
    'CycleTimeHrs',         # Number - First InProgress to Resolved
    'FixEffortHrs',         # Number - Effort from related tasks
    'ReopenEffortHrs',      # Number - Effort after reopen
]

print(f"✅ فیلدهای قابل محاسبه برای افزودن: {len(fields_to_add)}")

# Add new columns
start_col = ws_raw.max_column + 1
for idx, field_name in enumerate(fields_to_add):
    col = start_col + idx
    ws_raw.cell(1, col, field_name)
    print(f"   + {field_name}")

# Helper: Get column index by name
def get_col_index(name):
    for col in range(1, ws_raw.max_column + 1):
        if ws_raw.cell(1, col).value == name:
            return col
    return None

# Get indices for needed columns
created_date_col = get_col_index('CreatedDate')
closed_date_col = get_col_index('ClosedDate')
state_col = get_col_index('State')
reopen_count_col = get_col_index('ReopenCount')

# Get indices for new columns
is_duplicate_col = get_col_index('IsDuplicate')
duplicate_of_col = get_col_index('DuplicateOfBugID')
close_reason_col = get_col_index('CloseReason')
lead_time_col = get_col_index('LeadTimeHrs')
cycle_time_col = get_col_index('CycleTimeHrs')
fix_effort_col = get_col_index('FixEffortHrs')
reopen_effort_col = get_col_index('ReopenEffortHrs')

print("\n📝 تولید MOCK Data...")

# Close reasons distribution
close_reasons = ['Fixed', 'Duplicate', 'By Design', 'Won\'t Fix', 'Cannot Reproduce', 'External Dependency']
close_reason_weights = [0.70, 0.10, 0.08, 0.05, 0.05, 0.02]  # 70% Fixed, 10% Duplicate, etc.

# Track used bug IDs for duplicates
all_bug_ids = []
for row in range(2, ws_raw.max_row + 1):
    bug_id = ws_raw.cell(row, 1).value  # Assuming BugID is column 1
    if bug_id:
        all_bug_ids.append(bug_id)

# Generate MOCK data for each row
stats = {
    'duplicates': 0,
    'with_lead_time': 0,
    'with_reopen_effort': 0,
}

for row in range(2, ws_raw.max_row + 1):
    state = ws_raw.cell(row, state_col).value if state_col else None
    created_date = ws_raw.cell(row, created_date_col).value if created_date_col else None
    closed_date = ws_raw.cell(row, closed_date_col).value if closed_date_col else None
    reopen_count = ws_raw.cell(row, reopen_count_col).value if reopen_count_col else 0

    # CloseReason (only for closed/resolved bugs)
    close_reason = None
    if state in ['Closed', 'Resolved']:
        close_reason = random.choices(close_reasons, weights=close_reason_weights)[0]
        ws_raw.cell(row, close_reason_col, close_reason)
    else:
        ws_raw.cell(row, close_reason_col, 'N/A')

    # IsDuplicate and DuplicateOfBugID
    is_duplicate = (close_reason == 'Duplicate')
    ws_raw.cell(row, is_duplicate_col, 'Yes' if is_duplicate else 'No')

    if is_duplicate:
        # Pick a random bug as the original
        original_bug = random.choice([bid for bid in all_bug_ids if bid != ws_raw.cell(row, 1).value])
        ws_raw.cell(row, duplicate_of_col, original_bug)
        stats['duplicates'] += 1
    else:
        ws_raw.cell(row, duplicate_of_col, 'N/A')

    # LeadTimeHrs (Created to Closed)
    if created_date and closed_date and isinstance(created_date, datetime) and isinstance(closed_date, datetime):
        lead_time = (closed_date - created_date).total_seconds() / 3600
        ws_raw.cell(row, lead_time_col, round(lead_time, 2))
        stats['with_lead_time'] += 1
    else:
        ws_raw.cell(row, lead_time_col, 'N/A')

    # CycleTimeHrs (MOCK: typically 60-80% of LeadTime)
    if ws_raw.cell(row, lead_time_col).value != 'N/A':
        lead_val = ws_raw.cell(row, lead_time_col).value
        cycle_time = lead_val * random.uniform(0.6, 0.8)
        ws_raw.cell(row, cycle_time_col, round(cycle_time, 2))
    else:
        ws_raw.cell(row, cycle_time_col, 'N/A')

    # FixEffortHrs (MOCK: realistic effort in hours)
    if state in ['Closed', 'Resolved']:
        # Small bugs: 1-8 hrs, Medium: 8-24 hrs, Large: 24-80 hrs
        effort_range = random.choice([(1, 8), (8, 24), (24, 80)])
        fix_effort = random.uniform(effort_range[0], effort_range[1])
        ws_raw.cell(row, fix_effort_col, round(fix_effort, 2))
    else:
        ws_raw.cell(row, fix_effort_col, 'N/A')

    # ReopenEffortHrs (only for reopened bugs)
    if reopen_count and reopen_count > 0:
        # Reopen effort typically 2-6 hours per reopen
        reopen_effort = reopen_count * random.uniform(2, 6)
        ws_raw.cell(row, reopen_effort_col, round(reopen_effort, 2))
        stats['with_reopen_effort'] += 1
    else:
        ws_raw.cell(row, reopen_effort_col, 0)

print(f"   ✅ MOCK data generated:")
print(f"      - Duplicates: {stats['duplicates']} bugs")
print(f"      - With LeadTime: {stats['with_lead_time']} bugs")
print(f"      - With ReopenEffort: {stats['with_reopen_effort']} bugs")

# Update راهنمای_فیلدها
print("\n📖 به‌روزرسانی راهنمای_فیلدها...")

ws_guide = wb['راهنمای_فیلدها']

# Find where to insert (after existing fields)
insert_row = ws_guide.max_row + 1

# Yellow fill for calculable fields
yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

# Add field documentation
calculable_fields_docs = [
    ['IsDuplicate', 'آیا تکراری است', 'Boolean', 'Calculable',
     'محاسبه از CloseReason=="Duplicate"',
     'آیا این باگ تکراری باگ دیگری است (MOCK - قابل محاسبه)', '🟡'],

    ['DuplicateOfBugID', 'تکرار کدام باگ', 'Number', 'Calculable',
     'اگر IsDuplicate==Yes، شناسه باگ اصلی',
     'شناسه باگی که این باگ تکرار آن است (MOCK - قابل محاسبه)', '🟡'],

    ['CloseReason', 'دلیل بستن', 'Text', 'Calculable',
     'Fixed | Duplicate | By Design | Won\'t Fix | Cannot Reproduce | External Dependency',
     'دلیل بستن یا رد باگ (MOCK - قابل محاسبه)', '🟡'],

    ['LeadTimeHrs', 'زمان کل (ساعت)', 'Number', 'Calculable',
     'محاسبه از ClosedDate - CreatedDate',
     'زمان کل از ایجاد تا بستن باگ به ساعت (MOCK - قابل محاسبه)', '🟡'],

    ['CycleTimeHrs', 'زمان چرخه (ساعت)', 'Number', 'Calculable',
     'محاسبه از WorkItemRevisions (اولین InProgress تا Resolved)',
     'زمان فعال کاری از شروع تا حل باگ (MOCK - قابل محاسبه)', '🟡'],

    ['FixEffortHrs', 'ساعات رفع باگ', 'Number', 'Calculable',
     'محاسبه از Related Tasks با Work Item Type==Task',
     'مجموع ساعات کاری صرف‌شده برای رفع باگ از Related Tasks (MOCK - قابل محاسبه)', '🟡'],

    ['ReopenEffortHrs', 'ساعات پس از بازگشایی', 'Number', 'Calculable',
     'محاسبه تلاش بعد از Reopen از WorkItemRevisions',
     'مجموع ساعات کاری پس از بازگشایی باگ (MOCK - قابل محاسبه)', '🟡'],
]

for field_data in calculable_fields_docs:
    ws_guide.cell(insert_row, 1, field_data[0])  # Field Name
    ws_guide.cell(insert_row, 2, field_data[1])  # Persian Name
    ws_guide.cell(insert_row, 3, field_data[2])  # Data Type
    ws_guide.cell(insert_row, 4, field_data[3])  # Source
    ws_guide.cell(insert_row, 5, field_data[4])  # Formula/Query
    ws_guide.cell(insert_row, 6, field_data[5])  # Description

    # Apply yellow fill
    for col in range(1, 7):
        ws_guide.cell(insert_row, col).fill = yellow_fill

    insert_row += 1
    print(f"   + {field_data[0]}")

# Note: Summary cells may be merged, so we don't update them automatically
print(f"   ℹ️  Final field count: {len(headers) + len(fields_to_add)}")

# Save
wb.save('BugTracking_Complete_FINAL.xlsx')

print("\n" + "=" * 80)
print("✅ فیلدهای قابل محاسبه با موفقیت اضافه شدند!")
print("=" * 80)
print(f"""
📊 آمار نهایی:
   - فیلدها: {len(headers)} → {len(headers) + len(fields_to_add)}
   - فیلدهای اضافه‌شده: {len(fields_to_add)}
   - باگ‌ها: {ws_raw.max_row - 1}
   - فیلدهای تکراری: {stats['duplicates']} باگ

✅ همه فیلدها با MOCK data واقع‌گرایانه پر شدند
✅ راهنمای_فیلدها به‌روز شد
✅ همه فیلدها به‌عنوان "قابل محاسبه" مستند شدند
""")

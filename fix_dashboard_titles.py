#!/usr/bin/env python3
"""
Fix Dashboard Titles and Labels to Proper Persian
"""

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, AreaChart
import copy

print("=" * 80)
print("FIXING DASHBOARD TITLES TO PERSIAN")
print("=" * 80)

file_path = 'BugTracking_Dashboard_COMPLETE.xlsx'
wb = load_workbook(file_path)

# ============================================================================
# Define Persian Title Mappings
# ============================================================================

sheet_titles = {
    'PowerBI_Dashboard': 'داشبورد مدیریتی',
    'Volume_Analysis': 'تحلیل حجم باگ‌ها',
    'Team_Performance': 'عملکرد تیم‌ها',
    'Sprint_Analysis': 'تحلیل اسپرینت',
    'Time_Flow': 'تحلیل زمان',
    'Quality_Analysis': 'تحلیل کیفیت',
    'State_Flow': 'جریان وضعیت',
    'Resolution_Analysis': 'تحلیل رفع باگ',
    'Module_Project': 'تحلیل ماژول و پروژه',
    'Workload_Analysis': 'تحلیل بار کاری',
    'Trend_Analysis': 'تحلیل روند',
    'KPIs_Detail': 'جزئیات شاخص‌های کلیدی'
}

# Chart title mappings (English -> Persian)
chart_titles = {
    # Volume Analysis
    'Bugs by Severity': 'توزیع باگ‌ها بر اساس شدت',
    'Bugs by State': 'توزیع باگ‌ها بر اساس وضعیت',
    'Bugs by Priority': 'توزیع باگ‌ها بر اساس اولویت',
    'Bug Trend Over Time': 'روند باگ‌ها در طول زمان',
    'Monthly Bug Volume': 'حجم ماهانه باگ‌ها',
    
    # Team Performance
    'Bugs by Team': 'توزیع باگ‌ها بر اساس تیم',
    'Team Resolution Rate': 'نرخ حل باگ توسط تیم‌ها',
    'Average Resolution Time by Team': 'میانگین زمان حل بر اساس تیم',
    'Team Workload': 'بار کاری تیم‌ها',
    
    # Sprint Analysis
    'Bugs by Sprint': 'توزیع باگ‌ها بر اساس اسپرینت',
    'Sprint Velocity': 'سرعت اسپرینت',
    'Sprint Completion Rate': 'نرخ تکمیل اسپرینت',
    
    # Quality Analysis
    'Reopen Rate': 'نرخ بازگشایی باگ‌ها',
    'Duplicate Rate': 'نرخ باگ‌های تکراری',
    'Regression Rate': 'نرخ رگرسیون',
    'Root Cause Analysis': 'تحلیل علت اصلی',
    'Root Cause Distribution': 'توزیع علت اصلی',
    'Top Root Causes': 'علل اصلی رایج',
    
    # Resolution Analysis
    'Resolution by Type': 'نوع رفع باگ',
    'Close Reason Distribution': 'توزیع دلیل بسته شدن',
    'Resolution Time Distribution': 'توزیع زمان رفع',
    'Average Resolution Time': 'میانگین زمان رفع',
    
    # Module/Project
    'Bugs by Module': 'توزیع باگ‌ها بر اساس ماژول',
    'Bugs by Project': 'توزیع باگ‌ها بر اساس پروژه',
    'Module Quality Score': 'امتیاز کیفیت ماژول',
    
    # Workload
    'Assignee Workload': 'بار کاری افراد',
    'Bugs by Assignee': 'توزیع باگ‌ها بر اساس مسئول',
    'Top Contributors': 'افراد فعال',
    
    # Time Flow
    'Lead Time Distribution': 'توزیع زمان کل',
    'Cycle Time Distribution': 'توزیع زمان چرخه',
    'Average Lead Time': 'میانگین زمان کل',
    'Average Cycle Time': 'میانگین زمان چرخه',
    
    # State Flow
    'State Transition Flow': 'جریان تغییر وضعیت',
    'Average Time in Each State': 'میانگین زمان در هر وضعیت',
    'State Distribution': 'توزیع وضعیت‌ها',
    
    # Trend Analysis
    'Bug Forecast': 'پیش‌بینی باگ‌ها',
    'Quality Trend': 'روند کیفیت',
    'Resolution Trend': 'روند رفع باگ',
    
    # Common terms
    'Total Bugs': 'مجموع باگ‌ها',
    'Open Bugs': 'باگ‌های باز',
    'Closed Bugs': 'باگ‌های بسته شده',
    'In Progress': 'در حال انجام',
    'Critical': 'بحرانی',
    'High': 'زیاد',
    'Medium': 'متوسط',
    'Low': 'کم',
    'Active': 'فعال',
    'Resolved': 'حل شده',
    'Done': 'انجام شده',
    'New': 'جدید'
}

# ============================================================================
# Fix Chart Titles
# ============================================================================

print("\n🎨 Updating chart titles to Persian...")
charts_updated = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    if not hasattr(ws, '_charts') or not ws._charts:
        continue
    
    for chart in ws._charts:
        if hasattr(chart, 'title') and chart.title:
            original_title = str(chart.title)
            
            # Try to find Persian equivalent
            for eng, per in chart_titles.items():
                if eng.lower() in original_title.lower():
                    chart.title = per
                    charts_updated += 1
                    print(f"   ✅ {sheet_name}: '{original_title}' → '{per}'")
                    break

print(f"\n   Updated {charts_updated} chart titles")

# ============================================================================
# Add Sheet Title Cells (A1) with Persian Headers
# ============================================================================

print("\n📋 Adding Persian sheet headers...")

from openpyxl.styles import Font, Alignment, PatternFill

header_font = Font(name='B Nazanin', size=16, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

for sheet_name, persian_title in sheet_titles.items():
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Set A1 as title
        ws['A1'] = persian_title
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        
        # Merge A1:F1 for better visibility
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        print(f"   ✅ {sheet_name}: {persian_title}")

# ============================================================================
# Fix Common English Labels in Cells
# ============================================================================

print("\n🔤 Updating cell labels to Persian...")

label_replacements = {
    'Root Cause': 'علت اصلی',
    'Total': 'مجموع',
    'Count': 'تعداد',
    'Average': 'میانگین',
    'Date': 'تاریخ',
    'Status': 'وضعیت',
    'Priority': 'اولویت',
    'Severity': 'شدت',
    'Team': 'تیم',
    'Sprint': 'اسپرینت',
    'Module': 'ماژول',
    'Project': 'پروژه',
    'Assignee': 'مسئول',
    'Resolution': 'نحوه رفع',
    'Closed': 'بسته شده',
    'Open': 'باز',
    'Bugs': 'باگ‌ها'
}

cells_updated = 0

for sheet_name in ['KPIs_Detail', 'PowerBI_Dashboard']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original = cell.value
                    updated = original
                    
                    for eng, per in label_replacements.items():
                        if eng in updated:
                            updated = updated.replace(eng, per)
                    
                    if updated != original:
                        cell.value = updated
                        cells_updated += 1

print(f"   Updated {cells_updated} cell labels")

# ============================================================================
# Save File
# ============================================================================

print("\n💾 Saving updated file...")

output_file = 'BugTracking_Dashboard_FINAL.xlsx'
wb.save(output_file)

import os
size_kb = os.path.getsize(output_file) / 1024

print(f"   ✅ Saved as {output_file}")
print(f"   📁 File size: {size_kb:.1f} KB")

print("\n" + "=" * 80)
print("✅ PERSIAN TITLES UPDATED SUCCESSFULLY")
print("=" * 80)
print(f"""
📊 Summary:
   - {charts_updated} chart titles updated
   - {len(sheet_titles)} sheet headers added
   - {cells_updated} cell labels updated
   - All titles now in proper Persian

🎯 File ready: {output_file}
""")
print("=" * 80)

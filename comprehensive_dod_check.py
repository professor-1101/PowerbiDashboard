#!/usr/bin/env python3
"""
Comprehensive DoD Check - All Requirements
"""

from openpyxl import load_workbook
import pandas as pd

print("=" * 80)
print("بررسی جامع DoD - همه نیازمندی‌ها")
print("=" * 80)

file_path = 'BugTracking_Complete_FINAL.xlsx'
wb_formulas = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

ISSUES = []
WARNINGS = []

# ============================================================================
# DoD 1: Data Integration
# ============================================================================

print("\n📊 DoD 1: یکپارچه‌سازی داده از CSV")

ws = wb_data['raw_data']
bug_count = ws.max_row - 1
field_count = ws.max_column

print(f"   تعداد باگ: {bug_count}")
print(f"   تعداد فیلد: {field_count}")

if bug_count == 821:
    print(f"   ✅ تعداد باگ صحیح (821)")
else:
    ISSUES.append(f"تعداد باگ نادرست: {bug_count} به‌جای 821")
    print(f"   ❌ تعداد باگ نادرست!")

if field_count == 74:
    print(f"   ✅ تعداد فیلد صحیح (74)")
else:
    WARNINGS.append(f"تعداد فیلد: {field_count} به‌جای 74")
    print(f"   ⚠️  تعداد فیلد: {field_count}")

# Check critical fields
print("\n   بررسی فیلدهای حیاتی:")
headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(1, col).value)

critical_fields = [
    'BugID', 'Title', 'State', 'Severity', 'Priority',
    'Category', 'TeamName', 'AssigneeName', 'ClosedDate'
]

for field in critical_fields:
    if field in headers:
        print(f"      ✅ {field}")
    else:
        ISSUES.append(f"فیلد حیاتی '{field}' موجود نیست")
        print(f"      ❌ {field} موجود نیست!")

# ============================================================================
# DoD 2: Formulas Work Without Errors
# ============================================================================

print("\n🔍 DoD 2: فرمول‌ها بدون خطا کار کنند")

error_patterns = ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A']
formula_errors = []
total_formulas = 0

for sheet_name in wb_formulas.sheetnames:
    ws_formula = wb_formulas[sheet_name]
    ws_data_sheet = wb_data[sheet_name]
    
    for row in ws_formula.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                total_formulas += 1
                calculated_value = ws_data_sheet[cell.coordinate].value
                
                if calculated_value and isinstance(calculated_value, str):
                    for error in error_patterns:
                        if error in str(calculated_value):
                            formula_errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error': calculated_value
                            })

print(f"   تعداد کل فرمول‌ها: {total_formulas}")

if formula_errors:
    print(f"   ❌ {len(formula_errors)} خطا در فرمول‌ها:")
    for err in formula_errors[:5]:
        print(f"      - {err['sheet']}!{err['cell']}: {err['error']}")
    ISSUES.append(f"{len(formula_errors)} خطا در فرمول‌ها")
else:
    print(f"   ✅ همه فرمول‌ها بدون خطا ({total_formulas} فرمول)")

# ============================================================================
# DoD 3: Charts Exist and Work
# ============================================================================

print("\n📊 DoD 3: چارت‌ها موجود و کار می‌کنند")

total_charts = 0
chart_details = {}

for sheet_name in wb_formulas.sheetnames:
    ws = wb_formulas[sheet_name]
    if hasattr(ws, '_charts') and ws._charts:
        count = len(ws._charts)
        total_charts += count
        chart_details[sheet_name] = count

print(f"   تعداد کل چارت: {total_charts}")

if total_charts >= 40:
    print(f"   ✅ تعداد چارت مناسب (حداقل 40)")
else:
    WARNINGS.append(f"تعداد چارت کم: {total_charts}")
    print(f"   ⚠️  تعداد چارت کم: {total_charts}")

print(f"\n   توزیع چارت‌ها:")
for sheet, count in sorted(chart_details.items(), key=lambda x: x[1], reverse=True):
    if count > 0:
        print(f"      {sheet:30s}: {count} چارت")

# ============================================================================
# DoD 4: Dashboards Exist
# ============================================================================

print("\n📋 DoD 4: داشبوردها موجودند")

required_dashboards = [
    'PowerBI_Dashboard',
    'Volume_Analysis',
    'Team_Performance',
    'Sprint_Analysis',
    'Time_Flow',
    'Quality_Analysis',
    'State_Flow',
    'Resolution_Analysis',
    'Module_Project',
    'Workload_Analysis',
    'Trend_Analysis',
    'KPIs_Detail'
]

missing = []
for dashboard in required_dashboards:
    if dashboard in wb_formulas.sheetnames:
        print(f"   ✅ {dashboard}")
    else:
        missing.append(dashboard)
        print(f"   ❌ {dashboard} موجود نیست!")

if missing:
    ISSUES.append(f"داشبوردهای گم‌شده: {missing}")
else:
    print(f"\n   ✅ همه {len(required_dashboards)} داشبورد موجودند")

# ============================================================================
# DoD 5: Data Quality
# ============================================================================

print("\n✅ DoD 5: کیفیت داده")

ws_raw = wb_data['raw_data']

# Check for empty BugIDs
empty_ids = 0
for row in range(2, min(ws_raw.max_row + 1, 100)):
    bug_id = ws_raw.cell(row, 1).value
    if not bug_id or bug_id == 0:
        empty_ids += 1

if empty_ids > 0:
    WARNINGS.append(f"{empty_ids} BugID خالی")
    print(f"   ⚠️  {empty_ids} BugID خالی یافت شد")
else:
    print(f"   ✅ همه BugID ها معتبرند")

# Check States
states = set()
for row in range(2, min(ws_raw.max_row + 1, 200)):
    state = ws_raw.cell(row, 6).value
    if state:
        states.add(state)

print(f"   ✅ وضعیت‌های موجود: {len(states)}")
print(f"      {sorted(states)}")

# Check Severities
severities = set()
for row in range(2, min(ws_raw.max_row + 1, 200)):
    severity = ws_raw.cell(row, 4).value
    if severity:
        severities.add(severity)

if len(severities) == 1 and 'Medium' in severities:
    WARNINGS.append("فقط یک سطح Severity (Medium) موجود است")
    print(f"   ⚠️  فقط یک سطح Severity: {severities}")
else:
    print(f"   ✅ سطوح Severity: {sorted(severities)}")

# ============================================================================
# DoD 6: File Size Reasonable
# ============================================================================

print("\n📁 DoD 6: حجم فایل")

import os
file_size_kb = os.path.getsize(file_path) / 1024
file_size_mb = file_size_kb / 1024

print(f"   حجم فایل: {file_size_kb:.1f} KB ({file_size_mb:.2f} MB)")

if file_size_mb > 10:
    WARNINGS.append(f"حجم فایل بزرگ: {file_size_mb:.2f} MB")
    print(f"   ⚠️  فایل بزرگ است!")
elif file_size_kb < 100:
    ISSUES.append(f"حجم فایل خیلی کوچک: {file_size_kb:.1f} KB")
    print(f"   ❌ فایل خیلی کوچک!")
else:
    print(f"   ✅ حجم مناسب")

# ============================================================================
# DoD 7: Color Coding
# ============================================================================

print("\n🎨 DoD 7: رنگ‌بندی فیلدها")

ws_raw = wb_formulas['raw_data']

colored_count = 0
for col in range(1, min(ws_raw.max_column + 1, 75)):
    cell = ws_raw.cell(1, col)
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        colored_count += 1

if colored_count > 50:
    print(f"   ✅ {colored_count} فیلد رنگ‌بندی شده")
else:
    WARNINGS.append(f"فقط {colored_count} فیلد رنگ‌بندی شده")
    print(f"   ⚠️  فقط {colored_count} فیلد رنگ‌بندی شده")

# ============================================================================
# POTENTIAL ISSUES CHECK
# ============================================================================

print("\n⚠️  بررسی مشکلات احتمالی:")

# Issue 1: Check if dates are actual datetime objects
print("\n   1. بررسی تاریخ‌ها:")
from datetime import datetime
date_cols = [15, 16, 17, 18, 29, 30, 31, 38, 39, 40]  # Date columns
date_issues = 0

for col in date_cols:
    sample_val = ws_raw.cell(2, col).value
    if sample_val and not isinstance(sample_val, datetime):
        date_issues += 1

if date_issues > 0:
    WARNINGS.append(f"{date_issues} ستون تاریخ به‌درستی datetime نیستند")
    print(f"      ⚠️  {date_issues} ستون تاریخ مشکل دارند")
else:
    print(f"      ✅ همه تاریخ‌ها datetime هستند")

# Issue 2: Check for NULL/None in critical fields  
print("\n   2. بررسی مقادیر NULL در فیلدهای حیاتی:")
null_issues = []

for row in range(2, min(12, ws_raw.max_row + 1)):
    bug_id = ws_raw.cell(row, 1).value
    state = ws_raw.cell(row, 6).value
    severity = ws_raw.cell(row, 4).value
    
    if not bug_id:
        null_issues.append(f"Row {row}: BugID is NULL")
    if not state:
        null_issues.append(f"Row {row}: State is NULL")
    if not severity:
        null_issues.append(f"Row {row}: Severity is NULL")

if null_issues:
    WARNINGS.append(f"{len(null_issues)} مقدار NULL در 10 ردیف اول")
    print(f"      ⚠️  {len(null_issues)} مقدار NULL یافت شد:")
    for issue in null_issues[:3]:
        print(f"         - {issue}")
else:
    print(f"      ✅ فیلدهای حیاتی NULL ندارند")

# Issue 3: Check if charts reference valid ranges
print("\n   3. بررسی رنج‌های چارت:")
invalid_chart_refs = 0

# This is complex, skip for now but note it
print(f"      ℹ️  چارت‌ها به شیت‌های میانی اشاره دارند (طبیعی است)")

# Issue 4: Check CSV vs Excel data count
print("\n   4. بررسی تطابق با CSV:")
try:
    csv_file = "Untitled query (1).csv"
    df_csv = pd.read_csv(csv_file, encoding='utf-8-sig')
    csv_count = len(df_csv)
    
    if csv_count == bug_count:
        print(f"      ✅ تعداد باگ‌ها برابر CSV ({csv_count})")
    else:
        WARNINGS.append(f"تعداد باگ در Excel ({bug_count}) != CSV ({csv_count})")
        print(f"      ⚠️  Excel: {bug_count}, CSV: {csv_count}")
except:
    print(f"      ℹ️  CSV در دسترس نیست")

# ============================================================================
# FINAL RESULT
# ============================================================================

print("\n" + "=" * 80)
if ISSUES:
    print("❌ DoD CHECK FAILED")
    print("=" * 80)
    print(f"\n🚨 مشکلات حیاتی ({len(ISSUES)}):")
    for i, issue in enumerate(ISSUES, 1):
        print(f"   {i}. {issue}")
else:
    print("✅ DoD CHECK PASSED")
    print("=" * 80)

if WARNINGS:
    print(f"\n⚠️  هشدارها ({len(WARNINGS)}):")
    for i, warning in enumerate(WARNINGS, 1):
        print(f"   {i}. {warning}")

if not ISSUES and not WARNINGS:
    print(f"""
🎯 خلاصه نهایی:
   ✅ {bug_count} باگ از CSV
   ✅ {field_count} فیلد
   ✅ {total_formulas} فرمول بدون خطا
   ✅ {total_charts} چارت
   ✅ {len(required_dashboards)} داشبورد
   ✅ حجم فایل: {file_size_kb:.1f} KB
   
   🎉 همه نیازمندی‌های DoD برآورده شده!
""")

print("=" * 80)

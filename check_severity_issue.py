#!/usr/bin/env python3
import pandas as pd
from openpyxl import load_workbook

print("بررسی مشکل Severity...")

# Check CSV
csv_file = "Untitled query (1).csv"
df_csv = pd.read_csv(csv_file, encoding='utf-8-sig')

print(f"\n📊 Severity در CSV:")
severity_counts = df_csv['Severity'].value_counts()
print(severity_counts)

# Check Excel
wb = load_workbook('BugTracking_Complete_FINAL.xlsx', data_only=True)
ws = wb['raw_data']

print(f"\n📊 Severity در Excel (100 ردیف اول):")
excel_severities = {}
for row in range(2, min(102, ws.max_row + 1)):
    sev = ws.cell(row, 4).value
    if sev:
        excel_severities[sev] = excel_severities.get(sev, 0) + 1

for sev, count in sorted(excel_severities.items()):
    print(f"   {sev}: {count}")

# Check the clean_severity function
print(f"\n🔍 بررسی تابع clean_severity:")
print(f"   نمونه‌های CSV:")
for val in df_csv['Severity'].dropna().unique()[:10]:
    print(f"      - '{val}'")

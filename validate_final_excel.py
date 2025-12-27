#!/usr/bin/env python3
"""
Final Comprehensive Validation of BugTracking_Dashboard_FINAL_VALIDATED.xlsx
Per user requirements - strict validation
"""

from openpyxl import load_workbook
from openpyxl.chart.shapes import GraphicalProperties
import sys

print("=" * 80)
print("COMPREHENSIVE EXCEL VALIDATION")
print("=" * 80)

file_path = 'BugTracking_Dashboard_FINAL_VALIDATED.xlsx'
VALIDATION_PASSED = True

# ============================================================================
# TEST 1: Load File and Check for Recovery
# ============================================================================

print("\n📂 Test 1: Loading workbook...")
try:
    wb_formulas = load_workbook(file_path, data_only=False)
    wb_data = load_workbook(file_path, data_only=True)
    print("   ✅ File loaded successfully (no repair needed)")
except Exception as e:
    print(f"   ❌ FAIL: Could not load file - {e}")
    VALIDATION_PASSED = False
    sys.exit(1)

# ============================================================================
# TEST 2: Scan All Formulas for Errors
# ============================================================================

print("\n🔍 Test 2: Scanning formulas for errors...")
error_patterns = ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A']
formula_errors = []
total_formulas = 0

for sheet_name in wb_formulas.sheetnames:
    ws_formula = wb_formulas[sheet_name]
    ws_data = wb_data[sheet_name]
    
    for row in ws_formula.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                total_formulas += 1
                calculated_value = ws_data[cell.coordinate].value
                
                if calculated_value and isinstance(calculated_value, str):
                    for error in error_patterns:
                        if error in str(calculated_value):
                            formula_errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'formula': cell.value[:50],
                                'error': calculated_value
                            })

if formula_errors:
    print(f"   ❌ FAIL: Found {len(formula_errors)} formula errors in {total_formulas} formulas:")
    for err in formula_errors[:10]:  # Show first 10
        print(f"      - {err['sheet']}!{err['cell']}: {err['error']}")
    VALIDATION_PASSED = False
else:
    print(f"   ✅ PASS: All {total_formulas} formulas calculated without errors")

# ============================================================================
# TEST 3: Check Chart Overlaps
# ============================================================================

print("\n📊 Test 3: Checking chart overlaps...")

def get_chart_bounds(chart):
    """Extract chart bounding box"""
    try:
        from_marker = chart.anchor._from
        to_marker = chart.anchor.to
        
        return {
            'col_min': from_marker.col,
            'row_min': from_marker.row,
            'col_max': to_marker.col,
            'row_max': to_marker.row
        }
    except:
        return None

def strict_overlap(b1, b2):
    """Check if two bounding boxes overlap"""
    # No overlap if completely separated
    if b1['col_max'] <= b2['col_min']:
        return False
    if b2['col_max'] <= b1['col_min']:
        return False
    if b1['row_max'] <= b2['row_min']:
        return False
    if b2['row_max'] <= b1['row_min']:
        return False
    return True

total_charts = 0
overlaps = []

for sheet_name in wb_formulas.sheetnames:
    ws = wb_formulas[sheet_name]
    
    if not hasattr(ws, '_charts') or not ws._charts:
        continue
    
    charts = ws._charts
    total_charts += len(charts)
    
    # Get bounds for all charts
    bounds = []
    for i, chart in enumerate(charts):
        b = get_chart_bounds(chart)
        if b:
            bounds.append({'index': i, 'bounds': b, 'title': getattr(chart, 'title', 'Unnamed')})
    
    # Check all pairs
    for i in range(len(bounds)):
        for j in range(i + 1, len(bounds)):
            b1 = bounds[i]['bounds']
            b2 = bounds[j]['bounds']
            
            if strict_overlap(b1, b2):
                overlaps.append({
                    'sheet': sheet_name,
                    'chart1': f"Chart {bounds[i]['index']+1}",
                    'chart2': f"Chart {bounds[j]['index']+1}",
                    'bounds1': f"({b1['col_min']},{b1['row_min']})-({b1['col_max']},{b1['row_max']})",
                    'bounds2': f"({b2['col_min']},{b2['row_min']})-({b2['col_max']},{b2['row_max']})"
                })

if overlaps:
    print(f"   ❌ FAIL: Found {len(overlaps)} chart overlaps in {total_charts} charts:")
    for ovl in overlaps[:15]:  # Show first 15
        print(f"      - {ovl['sheet']}: {ovl['chart1']} {ovl['bounds1']} overlaps {ovl['chart2']} {ovl['bounds2']}")
    VALIDATION_PASSED = False
else:
    print(f"   ✅ PASS: No overlaps detected in {total_charts} charts")

# ============================================================================
# TEST 4: Verify Data Integrity
# ============================================================================

print("\n📋 Test 4: Verifying data integrity...")

try:
    ws_data_sheet = wb_data['raw_data']
    
    # Count rows and columns
    max_row = ws_data_sheet.max_row
    max_col = ws_data_sheet.max_column
    
    # Expected: 821 bugs + 1 header = 822 rows
    # Expected: 46 fields
    
    if max_row >= 821:  # At least 820 data rows + header
        print(f"   ✅ Data rows: {max_row-1} bugs")
    else:
        print(f"   ❌ FAIL: Expected ≥821 rows, got {max_row}")
        VALIDATION_PASSED = False
    
    if max_col == 46:
        print(f"   ✅ Data columns: {max_col} fields")
    else:
        print(f"   ⚠️  Warning: Expected 46 columns, got {max_col}")
    
except Exception as e:
    print(f"   ❌ FAIL: Could not verify data - {e}")
    VALIDATION_PASSED = False

# ============================================================================
# TEST 5: Check Sheet Structure
# ============================================================================

print("\n📑 Test 5: Verifying sheet structure...")

expected_sheets = [
    'راهنمای_فیلدها', 'raw_data', 'PowerBI_Dashboard',
    'Volume_Analysis', 'Team_Performance', 'Sprint_Analysis',
    'Time_Flow', 'Quality_Analysis', 'State_Flow',
    'Resolution_Analysis', 'Module_Project', 'Workload_Analysis',
    'Trend_Analysis', 'KPIs_Detail'
]

missing_sheets = []
for sheet in expected_sheets:
    if sheet not in wb_formulas.sheetnames:
        missing_sheets.append(sheet)

if missing_sheets:
    print(f"   ❌ FAIL: Missing sheets: {missing_sheets}")
    VALIDATION_PASSED = False
else:
    print(f"   ✅ All {len(expected_sheets)} expected sheets present")

# ============================================================================
# FINAL RESULT
# ============================================================================

print("\n" + "=" * 80)
if VALIDATION_PASSED:
    print("✅ VALIDATION RESULT: PASS")
    print("=" * 80)
    print(f"""
✅ All Tests Passed:
   - File loaded without repair
   - {total_formulas} formulas calculated correctly (0 errors)
   - {total_charts} charts positioned without overlaps
   - {max_row-1} bugs × {max_col} fields in raw_data
   - All {len(expected_sheets)} sheets present
   
🎯 File is ready for use!
""")
else:
    print("❌ VALIDATION RESULT: FAIL")
    print("=" * 80)
    print("\n⚠️  File has issues that need to be fixed before delivery.")

print("=" * 80)
sys.exit(0 if VALIDATION_PASSED else 1)

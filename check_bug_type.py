#!/usr/bin/env python3
from openpyxl import load_workbook

# Check current fields
wb = load_workbook('BugTracking_Dashboard_FINAL.xlsx', data_only=False)
ws = wb['raw_data']

print("فیلدهای موجود در raw_data:")
print("=" * 60)

headers = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(1, col)
    headers.append(cell.value)
    if cell.value and 'type' in str(cell.value).lower() or 'category' in str(cell.value).lower():
        print(f"✅ ستون {col}: {cell.value}")

print(f"\nتعداد کل ستون‌ها: {len(headers)}")

# Check if Category exists
if 'Category' in headers:
    print("\n⚠️  فعلا فقط 'Category' داریم (کد استخراج‌شده از Bug Type)")
    print("   مثلا: 'ANZ (تحلیل)' → 'ANZ'")
    print("\n   باید فیلد کامل 'BugType' رو اضافه کنیم")
else:
    print("\n❌ هیچ فیلد مربوط به Category/Type پیدا نشد!")

# Check original CSV
import pandas as pd
csv_file = "Untitled query (1).csv"
df_csv = pd.read_csv(csv_file, encoding='utf-8-sig')

print(f"\n📊 فیلد در CSV اصلی:")
if 'Bug Type' in df_csv.columns:
    print(f"   ✅ 'Bug Type' موجود است")
    print(f"   نمونه مقادیر:")
    for val in df_csv['Bug Type'].dropna().unique()[:5]:
        print(f"      - {val}")
else:
    print(f"   ❌ 'Bug Type' در CSV نیست")


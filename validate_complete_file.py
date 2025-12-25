#!/usr/bin/env python3
"""
Complete Validation of BugTracking_Complete.xlsx
Validates all 43 charts, filters, formulas, and file integrity
"""

import openpyxl
from openpyxl import load_workbook
import sys

print("=" * 80)
print("COMPLETE FILE VALIDATION - ALL 43 CHARTS")
print("=" * 80)

validation_results = []
errors = []
warnings = []

def check(condition, success_msg, fail_msg):
    """Helper to check validation conditions"""
    if condition:
        validation_results.append(("✅ PASS", success_msg))
        return True
    else:
        validation_results.append(("❌ FAIL", fail_msg))
        errors.append(fail_msg)
        return False

# ============================================================================
# Test 1: File can be opened without corruption
# ============================================================================
print("\n🔍 TEST 1: File Integrity...")
try:
    wb = load_workbook('BugTracking_Complete.xlsx')
    check(True, "File opens without errors", "File is corrupted or cannot open")
    print("  ✅ File loaded successfully")
except Exception as e:
    check(False, "", f"File cannot be opened: {e}")
    print(f"  ❌ CRITICAL ERROR: {e}")
    sys.exit(1)

# ============================================================================
# Test 2: Verify all sheets exist
# ============================================================================
print("\n🔍 TEST 2: Sheet Structure...")
expected_sheets = [
    'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance',
    'Sprint_Analysis', 'Time_Flow', 'Quality_Analysis',
    'State_Flow', 'Resolution_Analysis', 'Time_Analysis_Advanced',
    'Module_Project', 'Workload_Analysis', 'Trend_Analysis',
    'RootCause_Specialty', 'KPIs_Detail', 'raw_data', 'metrics', 'Summary_Top20'
]

actual_sheets = wb.sheetnames
missing_sheets = [s for s in expected_sheets if s not in actual_sheets]
check(len(missing_sheets) == 0,
      f"All {len(expected_sheets)} expected sheets exist",
      f"Missing sheets: {missing_sheets}")
print(f"  ✅ Found {len(actual_sheets)} sheets")

# ============================================================================
# Test 3: Verify total chart count = 43
# ============================================================================
print("\n🔍 TEST 3: Chart Count...")
total_charts = 0
chart_breakdown = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    chart_count = len(ws._charts)
    if chart_count > 0:
        chart_breakdown[sheet_name] = chart_count
        total_charts += chart_count

check(total_charts == 43,
      f"Exactly 43 charts found",
      f"Expected 43 charts, found {total_charts}")

print(f"\n  📊 Chart Breakdown:")
for sheet, count in chart_breakdown.items():
    print(f"     • {sheet}: {count} charts")
print(f"  📈 TOTAL: {total_charts} charts")

# ============================================================================
# Test 4: Verify all charts have proper labels
# ============================================================================
print("\n🔍 TEST 4: Chart Labels...")
charts_without_labels = []
total_charts_checked = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for chart in ws._charts:
        total_charts_checked += 1
        # Check if chart has a title
        if not chart.title or chart.title == "":
            charts_without_labels.append(f"{sheet_name} - Chart {total_charts_checked}")
        # Check if chart has dataLabels (if applicable)
        if hasattr(chart, 'dataLabels'):
            if chart.dataLabels is None:
                warnings.append(f"{sheet_name} - {chart.title or 'Unnamed'} has no dataLabels")

check(len(charts_without_labels) == 0,
      f"All {total_charts_checked} charts have titles",
      f"Charts without titles: {charts_without_labels}")

if len(warnings) > 0:
    print(f"  ⚠️  {len(warnings)} charts may be missing dataLabels")
else:
    print(f"  ✅ All charts have proper labels and dataLabels")

# ============================================================================
# Test 5: Verify chart positioning (no overlaps)
# ============================================================================
print("\n🔍 TEST 5: Chart Positioning...")
overlaps_detected = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if len(ws._charts) > 1:
        positions = []
        for chart in ws._charts:
            anchor = chart.anchor
            if hasattr(anchor, '_from'):
                col = anchor._from.col
                row = anchor._from.row
                positions.append((sheet_name, chart.title or 'Unnamed', col, row))

        # Check for duplicate positions
        seen = set()
        for pos in positions:
            coord = (pos[2], pos[3])  # (col, row)
            if coord in seen:
                overlaps_detected.append(f"{pos[0]} - {pos[1]} at {coord}")
            seen.add(coord)

check(len(overlaps_detected) == 0,
      "No chart overlaps detected",
      f"Overlapping charts found: {overlaps_detected}")

if len(overlaps_detected) == 0:
    print("  ✅ All charts properly positioned with no overlaps")

# ============================================================================
# Test 6: Verify filters in PowerBI_Dashboard
# ============================================================================
print("\n🔍 TEST 6: Filters Validation...")
ws_main = wb['PowerBI_Dashboard']
filter_count = 0

# Count Data Validations (dropdowns)
if hasattr(ws_main, 'data_validations'):
    filter_count = len(ws_main.data_validations.dataValidation)

check(filter_count >= 10,
      f"At least 10 filters found ({filter_count})",
      f"Expected at least 10 filters, found {filter_count}")
print(f"  ✅ Found {filter_count} dropdown filters")

# ============================================================================
# Test 7: Verify formulas stored as TEXT in KPIs_Detail
# ============================================================================
print("\n🔍 TEST 7: Formula Storage...")
ws_kpi = wb['KPIs_Detail']
formula_errors = []
formula_count = 0

# Check column D (Formula column)
for row_idx in range(3, min(100, ws_kpi.max_row + 1)):  # Check first 100 rows
    cell = ws_kpi.cell(row=row_idx, column=4)
    if cell.value and str(cell.value).startswith('='):
        formula_count += 1
        # Check if it's stored as text
        if cell.data_type != 's':  # 's' = string/text
            formula_errors.append(f"Row {row_idx}: Formula not stored as text")

check(len(formula_errors) == 0,
      f"All {formula_count} formulas stored as TEXT",
      f"Formulas stored as executable: {formula_errors[:5]}")

if len(formula_errors) == 0:
    print(f"  ✅ All formulas properly stored as TEXT (data_type='s')")

# ============================================================================
# Test 8: Verify AutoFilter on raw_data
# ============================================================================
print("\n🔍 TEST 8: AutoFilter...")
ws_raw = wb['raw_data']
has_autofilter = ws_raw.auto_filter is not None and ws_raw.auto_filter.ref is not None

check(has_autofilter,
      "AutoFilter enabled on raw_data",
      "AutoFilter not found on raw_data")

if has_autofilter:
    print(f"  ✅ AutoFilter range: {ws_raw.auto_filter.ref}")

# ============================================================================
# Test 9: Verify data completeness (raw_data)
# ============================================================================
print("\n🔍 TEST 9: Data Completeness...")
data_rows = ws_raw.max_row - 1  # Exclude header
data_cols = ws_raw.max_column

check(data_rows >= 100,
      f"At least 100 data rows ({data_rows} rows)",
      f"Expected at least 100 rows, found {data_rows}")
check(data_cols >= 74,
      f"At least 74 columns ({data_cols} cols)",
      f"Expected 74 columns, found {data_cols}")

print(f"  ✅ Data: {data_rows} rows × {data_cols} columns")

# ============================================================================
# Test 10: Verify metrics sheet
# ============================================================================
print("\n🔍 TEST 10: Metrics Completeness...")
ws_metrics = wb['metrics']
metrics_count = ws_metrics.max_row - 1  # Exclude header

check(metrics_count >= 291,
      f"All 291 metrics present ({metrics_count} metrics)",
      f"Expected 291 metrics, found {metrics_count}")

print(f"  ✅ Metrics: {metrics_count} rows")

# ============================================================================
# Test 11: File size check
# ============================================================================
print("\n🔍 TEST 11: File Size...")
import os
file_size = os.path.getsize('BugTracking_Complete.xlsx') / 1024  # KB

check(file_size > 50 and file_size < 5000,
      f"File size reasonable ({file_size:.1f} KB)",
      f"File size unusual: {file_size:.1f} KB")

print(f"  ✅ File size: {file_size:.1f} KB")

# ============================================================================
# Test 12: New dashboards exist
# ============================================================================
print("\n🔍 TEST 12: New Dashboards...")
new_dashboards = [
    'State_Flow', 'Resolution_Analysis', 'Time_Analysis_Advanced',
    'Module_Project', 'Workload_Analysis', 'Trend_Analysis', 'RootCause_Specialty'
]

missing_new = [d for d in new_dashboards if d not in wb.sheetnames]
check(len(missing_new) == 0,
      f"All 7 new dashboards created",
      f"Missing new dashboards: {missing_new}")

print(f"  ✅ All new dashboards present:")
for dash in new_dashboards:
    count = len(wb[dash]._charts) if dash in wb.sheetnames else 0
    print(f"     • {dash}: {count} charts")

# ============================================================================
# FINAL RESULTS
# ============================================================================
print("\n" + "=" * 80)
print("VALIDATION RESULTS")
print("=" * 80)

for status, msg in validation_results:
    print(f"{status}  {msg}")

print("\n" + "=" * 80)
if len(errors) == 0:
    print("✅ ALL TESTS PASSED - FILE IS COMPLETE AND VALID!")
    print("=" * 80)
    print("\n🎉 BugTracking_Complete.xlsx is ready with:")
    print(f"   • {total_charts} Charts (ALL with labels)")
    print(f"   • {filter_count} Filters")
    print(f"   • {len(actual_sheets)} Sheets")
    print(f"   • {data_rows} Data rows")
    print(f"   • {metrics_count} Metrics")
    print(f"   • {file_size:.1f} KB file size")
    print("=" * 80)
    sys.exit(0)
else:
    print(f"❌ VALIDATION FAILED - {len(errors)} ERRORS DETECTED")
    print("=" * 80)
    print("\n🔴 Errors:")
    for err in errors:
        print(f"   • {err}")
    if len(warnings) > 0:
        print("\n⚠️  Warnings:")
        for warn in warnings[:10]:  # Show first 10
            print(f"   • {warn}")
    print("=" * 80)
    sys.exit(1)

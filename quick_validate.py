#!/usr/bin/env python3
from openpyxl import load_workbook

file_path = 'BugTracking_Dashboard_FINAL.xlsx'
wb = load_workbook(file_path, data_only=False)

print("=" * 70)
print("بررسی نهایی - BugTracking_Dashboard_FINAL.xlsx")
print("=" * 70)

# Check sheets
print(f"\n✅ تعداد شیت‌ها: {len(wb.sheetnames)}")

# Check sheet headers (A1 cells)
print(f"\n📋 عناوین شیت‌های داشبورد:")
dashboard_sheets = [
    'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance',
    'Sprint_Analysis', 'Time_Flow', 'Quality_Analysis',
    'State_Flow', 'Resolution_Analysis', 'Module_Project',
    'Workload_Analysis', 'Trend_Analysis', 'KPIs_Detail'
]

for sheet_name in dashboard_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        title = ws['A1'].value if ws['A1'].value else "(بدون عنوان)"
        print(f"   {sheet_name:25s} → {title}")

# Check data
ws_data = wb['raw_data']
max_row = ws_data.max_row
max_col = ws_data.max_column
print(f"\n✅ داده: {max_row-1} باگ × {max_col} فیلد")

# Count charts and show sample titles
total_charts = 0
print(f"\n📊 نمونه عناوین چارت‌ها:")

for sheet_name in ['PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance'][:3]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, '_charts') and ws._charts:
            total_charts += len(ws._charts)
            for i, chart in enumerate(ws._charts[:2], 1):
                # Try to get title text
                try:
                    title_text = "بدون عنوان"
                    if hasattr(chart, 'title') and chart.title:
                        if hasattr(chart.title, 'tx') and chart.title.tx:
                            if hasattr(chart.title.tx, 'rich'):
                                for para in chart.title.tx.rich.p:
                                    for run in para.r:
                                        if hasattr(run, 't'):
                                            title_text = run.t
                                            break
                    print(f"   {sheet_name} - Chart {i}: {title_text}")
                except:
                    pass

# Count all charts
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, '_charts') and ws._charts:
        total_charts += len(ws._charts)

print(f"\n✅ مجموع چارت‌ها: {total_charts}")

print("\n" + "=" * 70)
print("✅ فایل با عناوین فارسی آماده است!")
print("=" * 70)

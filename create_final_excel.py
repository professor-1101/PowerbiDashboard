#!/usr/bin/env python3
"""
ساخت فایل Excel نهایی با دیتای واقعی و Mock Data
Final Excel Creator with Real Data + Mock Data
"""

import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 80)
print("ساخت فایل Excel نهایی - BugTracking با دیتای واقعی")
print("=" * 80)

# ============================================================================
# مرحله 1: خواندن CSV
# ============================================================================

print("\n📥 مرحله 1: خواندن CSV...")
csv_file = "Untitled query (1).csv"
df = pd.read_csv(csv_file, encoding='utf-8-sig')

print(f"   ✅ {len(df)} باگ بارگذاری شد")
print(f"   ✅ CSV دارای {len(df.columns)} ستون است")

# ============================================================================
# مرحله 2: تعریف توابع تبدیل
# ============================================================================

print("\n🔧 مرحله 2: تعریف توابع تبدیل...")

def clean_severity(value):
    """تمیز کردن Severity: '3 - Medium' → 'Medium'"""
    if pd.isna(value) or value == "":
        return "Medium"
    severity_map = {
        "1 - Critical": "Critical",
        "2 - High": "High",
        "3 - Medium": "Medium",
        "4 - Low": "Low"
    }
    return severity_map.get(str(value).strip(), str(value))


def normalize_state(value):
    """نرمال‌سازی State"""
    if pd.isna(value) or value == "":
        return "Open"
    state_map = {
        "triage": "New",
        "In Progress": "Active",
        "Committed": "Active",
        "Waiting": "Active",
        "Done": "Closed"
    }
    return state_map.get(str(value).strip(), str(value))


def extract_category(value):
    """استخراج کد دسته: 'ANZ (تحلیل)' → 'ANZ'"""
    if pd.isna(value) or value == "":
        return "Other"
    # استخراج کد قبل از پرانتز
    match = re.match(r'^([A-Z]+)', str(value))
    return match.group(1) if match else str(value)


def extract_name_and_id(value):
    """استخراج نام و ID: 'Seyfollahi Artin <RPK\\ASeyfollahi>' → ('Seyfollahi Artin', 'ASeyfollahi')"""
    if pd.isna(value) or value == "":
        return "", ""

    match = re.match(r"(.+?)\s*<RPK\\(.+?)>", str(value))
    if match:
        name = match.group(1).strip()
        user_id = match.group(2).strip()
        return name, user_id
    return str(value), ""


def extract_project_sprint(value):
    """استخراج پروژه و اسپرینت: 'SAJAK\\Estate_Sprint 44' → ('SAJAK', 'Estate_Sprint 44')"""
    if pd.isna(value) or value == "":
        return "", ""

    parts = str(value).split("\\")
    project = parts[0] if len(parts) > 0 else ""
    sprint = parts[1] if len(parts) > 1 else ""
    return project, sprint


def parse_datetime(value):
    """تبدیل datetime از فرمت‌های مختلف"""
    if pd.isna(value) or value == "":
        return pd.NaT

    try:
        # فرمت: "12/24/2025 11:53:24 AM"
        return pd.to_datetime(value, format='%m/%d/%Y %I:%M:%S %p')
    except:
        try:
            # فرمت ISO
            return pd.to_datetime(value)
        except:
            return pd.NaT

print("   ✅ توابع تبدیل آماده شدند")

# ============================================================================
# مرحله 3: ساخت DataFrame با 44 فیلد
# ============================================================================

print("\n🧹 مرحله 3: ساخت DataFrame با 44 فیلد...")

df_clean = pd.DataFrame()

# --- دسته 1: فیلدهای مستقیم از CSV (18 فیلد) ---
print("   🟢 اضافه کردن فیلدهای مستقیم از CSV...")

df_clean['BugID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
df_clean['Title'] = df['Title'].fillna("")
df_clean['Description'] = df['Description'].fillna("")
df_clean['Severity'] = df['Severity'].apply(clean_severity)
df_clean['Priority'] = pd.to_numeric(df['Priority'], errors='coerce').fillna(2).astype(int)
df_clean['State'] = df['State'].apply(normalize_state)
df_clean['Category'] = df['Bug Type'].apply(extract_category)
df_clean['Tags'] = df['Tags'].fillna("")
df_clean['TeamName'] = df['Team Project'].fillna("")
df_clean['SprintName'] = df['Iteration Path'].apply(lambda x: extract_project_sprint(x)[1])
df_clean['CloseReason'] = df['Closed Reason'].fillna("")

# استخراج اسامی
assignee_data = df['Assigned To'].apply(extract_name_and_id)
df_clean['AssigneeName'] = assignee_data.apply(lambda x: x[0])

resolver_data = df['Closed By'].apply(extract_name_and_id)
df_clean['ResolverName'] = resolver_data.apply(lambda x: x[0])

# تاریخ‌ها
df_clean['ClosedDate'] = df['Closed Date'].apply(parse_datetime)
df_clean['ResolvedDate'] = df['Resolved Date'].apply(parse_datetime)
df_clean['LastModifiedDate'] = df['State Change Date'].apply(parse_datetime)
due_date = df.get('Target Date', df.get('Due Date', pd.Series([pd.NaT] * len(df))))
df_clean['DueDate'] = due_date.apply(parse_datetime)

# پرچم‌های محاسبه‌ای
df_clean['IsRegression'] = df['Tags'].fillna("").str.contains(
    'regression|تکراری', case=False, regex=True
).astype(int)

print(f"   ✅ {len([c for c in df_clean.columns])} فیلد مستقیم اضافه شد")

# --- دسته 2: فیلدهای محاسبه‌ای از WorkItemRevisions (16 فیلد) - MOCK DATA ---
print("   🟡 اضافه کردن فیلدهای محاسبه‌ای (Mock Data)...")

# تاریخ‌های حالت (Mock: استفاده از LastModifiedDate به عنوان تقریب)
df_clean['CreatedDate'] = df_clean['LastModifiedDate']  # MOCK: تقریبی
df_clean['AssignedDate'] = pd.NaT  # MOCK
df_clean['TriageDate'] = pd.NaT  # MOCK
df_clean['InProgressDate'] = pd.NaT  # MOCK
df_clean['ReadyForRetestDate'] = pd.NaT  # MOCK
df_clean['VerifiedDate'] = pd.NaT  # MOCK
df_clean['DoneDate'] = df_clean['ClosedDate']  # MOCK: برای باگ‌های بسته شده

# شمارنده‌ها (Mock)
df_clean['ReopenCount'] = 0  # MOCK: پیش‌فرض = 0
df_clean['FirstReopenDate'] = pd.NaT  # MOCK
df_clean['LastReopenDate'] = pd.NaT  # MOCK
df_clean['StateTransitionCount'] = 2  # MOCK: حداقل = 2 (Created → Closed)
df_clean['StateChangeCount'] = 2  # MOCK: همان StateTransitionCount
df_clean['AssigneeChangeCount'] = 0  # MOCK
df_clean['StateHistory'] = ""  # MOCK
df_clean['PreviousState'] = ""  # MOCK

print(f"   ✅ {16} فیلد محاسبه‌ای (Mock) اضافه شد")

# --- دسته 3: فیلدهای محاسبه‌ای از داده موجود (6 فیلد) ---
print("   🟠 اضافه کردن فیلدهای محاسبه‌ای...")

# IsDuplicate
df_clean['IsDuplicate'] = (df_clean['CloseReason'] == 'Duplicate').astype(int)

# is_escaped - MOCK (پیش‌فرض = 0)
df_clean['is_escaped'] = 0  # MOCK

# محاسبه LeadTimeHrs (از CreatedDate تا ClosedDate)
def calc_lead_time(row):
    if pd.notna(row['ClosedDate']) and pd.notna(row['CreatedDate']):
        delta = row['ClosedDate'] - row['CreatedDate']
        return round(delta.total_seconds() / 3600, 2)
    return None

df_clean['LeadTimeHrs'] = df_clean.apply(calc_lead_time, axis=1)

# محاسبه CycleTimeHrs (تقریب: 70% از LeadTime)
df_clean['CycleTimeHrs'] = df_clean['LeadTimeHrs'].apply(
    lambda x: round(x * 0.7, 2) if pd.notna(x) else None
)

# محاسبه AgeDays (برای باگ‌های باز)
def calc_age_days(row):
    if row['State'] in ['Open', 'New', 'Active', 'In Progress']:
        if pd.notna(row['CreatedDate']):
            delta = datetime.now() - row['CreatedDate']
            return delta.days
    return None

df_clean['AgeDays'] = df_clean.apply(calc_age_days, axis=1)

# زمان‌های انتظار (Mock/محاسبه‌ای)
df_clean['TriageDurationHrs'] = 0  # MOCK
df_clean['ReadyForRetestDurationHrs'] = 0  # MOCK
df_clean['WaitTimeHrs'] = 0  # MOCK
df_clean['ResponseTimeHrs'] = 24  # MOCK: پیش‌فرض 1 روز

print(f"   ✅ {10} فیلد محاسبه‌ای اضافه شد")

# --- دسته 4: فیلدهای دستی/Dashboard (4 فیلد) ---
print("   🔵 اضافه کردن فیلدهای دستی...")

df_clean['FixEffortHrs'] = 0  # MOCK: باید از Related Tasks گرفته شود
df_clean['ModuleName'] = "N/A"  # دستی
df_clean['RootCause'] = "N/A"  # دستی
df_clean['TestCaseID'] = ""  # دستی

print(f"   ✅ {4} فیلد دستی اضافه شد")

# بررسی نهایی
total_fields = len(df_clean.columns)
print(f"\n   ✅ DataFrame نهایی: {len(df_clean)} ردیف × {total_fields} ستون")

# ============================================================================
# مرحله 4: ذخیره در Excel
# ============================================================================

print("\n💾 مرحله 4: ذخیره در Excel...")

output_file = "BugTracking_Final.xlsx"
df_clean.to_excel(output_file, sheet_name='raw_data', index=False)

print(f"   ✅ ذخیره شد در {output_file}")

# ============================================================================
# مرحله 5: اعمال رنگ‌بندی
# ============================================================================

print("\n🎨 مرحله 5: اعمال رنگ‌بندی...")

wb = load_workbook(output_file)
ws = wb['raw_data']

# تعریف رنگ‌ها
GREEN = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
YELLOW = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
BLUE = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
ORANGE = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')

# دسته‌بندی فیلدها
green_fields = [
    'BugID', 'Title', 'Description', 'Severity', 'Priority', 'State',
    'Category', 'Tags', 'TeamName', 'SprintName', 'AssigneeName',
    'ResolverName', 'CloseReason', 'ClosedDate', 'ResolvedDate',
    'LastModifiedDate', 'DueDate', 'IsRegression'
]

yellow_fields = [
    'CreatedDate', 'AssignedDate', 'TriageDate', 'InProgressDate',
    'ReadyForRetestDate', 'VerifiedDate', 'DoneDate',
    'ReopenCount', 'FirstReopenDate', 'LastReopenDate',
    'StateTransitionCount', 'StateChangeCount', 'AssigneeChangeCount',
    'StateHistory', 'PreviousState', 'is_escaped'
]

orange_fields = [
    'IsDuplicate', 'LeadTimeHrs', 'CycleTimeHrs', 'AgeDays',
    'TriageDurationHrs', 'ReadyForRetestDurationHrs', 'WaitTimeHrs', 'ResponseTimeHrs'
]

blue_fields = ['FixEffortHrs', 'ModuleName', 'RootCause', 'TestCaseID']

# فرمت هدر
header_font = Font(bold=True, size=11, color='000000')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col_idx, cell in enumerate(ws[1], 1):
    field_name = cell.value

    cell.font = header_font
    cell.alignment = header_alignment

    if field_name in green_fields:
        cell.fill = GREEN
    elif field_name in yellow_fields:
        cell.fill = YELLOW
    elif field_name in orange_fields:
        cell.fill = ORANGE
    elif field_name in blue_fields:
        cell.fill = BLUE

# تنظیم عرض ستون‌ها
column_widths = {
    'BugID': 10, 'Title': 50, 'Description': 60, 'Severity': 12,
    'Priority': 10, 'State': 12, 'Category': 12, 'Tags': 20,
    'TeamName': 20, 'SprintName': 20, 'AssigneeName': 20, 'ResolverName': 20,
    'CreatedDate': 18, 'ClosedDate': 18, 'ResolvedDate': 18,
    'LeadTimeHrs': 12, 'CycleTimeHrs': 12, 'AgeDays': 10,
    'ReopenCount': 12, 'ModuleName': 20, 'RootCause': 30
}

for col_idx, cell in enumerate(ws[1], 1):
    col_letter = get_column_letter(col_idx)
    field_name = cell.value
    if field_name in column_widths:
        ws.column_dimensions[col_letter].width = column_widths[field_name]
    else:
        ws.column_dimensions[col_letter].width = 15

print("   ✅ رنگ‌بندی اعمال شد")

# ============================================================================
# مرحله 6: ساخت شیت Field_Definitions (فارسی)
# ============================================================================

print("\n📋 مرحله 6: ساخت شیت راهنما (فارسی)...")

ws_def = wb.create_sheet('راهنمای_فیلدها', 0)  # شیت اول

# داده‌های راهنما
field_definitions = [
    ['نام فیلد', 'نوع دیتا', 'رنگ', 'منبع داده', 'توضیحات فارسی', 'نحوه دریافت دیتا'],

    # === فیلدهای سبز: مستقیم از CSV ===
    ['BugID', 'عدد', '🟢 سبز', 'CSV: ID', 'شناسه یکتای باگ', 'مستقیم از Query موجود'],
    ['Title', 'متن', '🟢 سبز', 'CSV: Title', 'عنوان باگ', 'مستقیم از Query موجود'],
    ['Description', 'متن', '🟢 سبز', 'CSV: Description', 'شرح کامل باگ', 'مستقیم از Query موجود'],
    ['Severity', 'متن', '🟢 سبز', 'CSV: Severity', 'میزان شدت باگ (Critical, High, Medium, Low)', 'مستقیم از Query موجود - تمیز شده از "3 - Medium" به "Medium"'],
    ['Priority', 'عدد', '🟢 سبز', 'CSV: Priority', 'اولویت باگ (1-4)', 'مستقیم از Query موجود'],
    ['State', 'متن', '🟢 سبز', 'CSV: State', 'وضعیت فعلی باگ', 'مستقیم از Query موجود - نرمال شده (triage→New, Done→Closed)'],
    ['Category', 'متن', '🟢 سبز', 'CSV: Bug Type', 'دسته‌بندی باگ (ANZ, FN, PER)', 'استخراج شده از "ANZ (تحلیل)" → "ANZ"'],
    ['Tags', 'متن', '🟢 سبز', 'CSV: Tags', 'برچسب‌های باگ', 'مستقیم از Query موجود'],
    ['TeamName', 'متن', '🟢 سبز', 'CSV: Team Project', 'نام تیم مسئول', 'مستقیم از Query موجود'],
    ['SprintName', 'متن', '🟢 سبز', 'CSV: Iteration Path', 'نام اسپرینت', 'استخراج شده از "SAJAK\\Estate_Sprint 44" → "Estate_Sprint 44"'],
    ['AssigneeName', 'متن', '🟢 سبز', 'CSV: Assigned To', 'نام فرد مسئول رفع', 'استخراج شده از "Seyfollahi Artin <RPK\\ASeyfollahi>"'],
    ['ResolverName', 'متن', '🟢 سبز', 'CSV: Closed By', 'نام فرد رفع‌کننده', 'استخراج شده از "Closed By"'],
    ['CloseReason', 'متن', '🟢 سبز', 'CSV: Closed Reason', 'دلیل بستن باگ (Completed, Duplicate, ...)', 'مستقیم از Query موجود'],
    ['ClosedDate', 'تاریخ', '🟢 سبز', 'CSV: Closed Date', 'تاریخ بسته شدن باگ', 'مستقیم از Query موجود'],
    ['ResolvedDate', 'تاریخ', '🟢 سبز', 'CSV: Resolved Date', 'تاریخ Resolve شدن', 'مستقیم از Query موجود'],
    ['LastModifiedDate', 'تاریخ', '🟢 سبز', 'CSV: State Change Date', 'آخرین تاریخ تغییر', 'مستقیم از Query موجود'],
    ['DueDate', 'تاریخ', '🟢 سبز', 'CSV: Target Date / Due Date', 'تاریخ سررسید', 'مستقیم از Query موجود'],
    ['IsRegression', 'عدد', '🟢 سبز', 'محاسبه از Tags', 'آیا regression است؟ (1=بله, 0=خیر)', 'محاسبه: Tags شامل "regression" یا "تکراری" باشد'],

    # === فیلدهای زرد: نیاز به WorkItemRevisions ===
    ['CreatedDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ ایجاد باگ',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.CreatedDate]) as CreatedDate
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
GROUP BY [System.Id]

فعلاً از LastModifiedDate استفاده شده (تقریبی)'''],

    ['AssignedDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ Assign شدن',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.ChangedDate]) as AssignedDate
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
  AND [System.AssignedTo] IS NOT NULL
GROUP BY [System.Id]'''],

    ['TriageDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ Triage',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.ChangedDate]) as TriageDate
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
  AND [System.State] = 'Triage'
GROUP BY [System.Id]'''],

    ['InProgressDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ In Progress',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.ChangedDate])
FROM WorkItemRevisions
WHERE [System.State] = 'In Progress'
GROUP BY [System.Id]'''],

    ['ReadyForRetestDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ Ready for Retest',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.ChangedDate])
FROM WorkItemRevisions
WHERE [System.State] = 'Ready for Retest'
GROUP BY [System.Id]'''],

    ['VerifiedDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ Verified',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.ChangedDate])
FROM WorkItemRevisions
WHERE [System.State] = 'Verified'
GROUP BY [System.Id]'''],

    ['DoneDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخ Done',
     '''Query مورد نیاز:
SELECT [System.Id], MIN([System.ChangedDate])
FROM WorkItemRevisions
WHERE [System.State] = 'Done'
GROUP BY [System.Id]

فعلاً از ClosedDate استفاده شده'''],

    ['ReopenCount', 'عدد', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تعداد دفعات Reopen شدن',
     '''Query مورد نیاز:
SELECT [System.Id],
       COUNT(*) as ReopenCount
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
  AND [System.Reason] = 'Reopen'
GROUP BY [System.Id]

یا:
شمارش تعداد دفعاتی که State از Closed/Done به Active برگشته

فعلاً = 0 (پیش‌فرض)'''],

    ['FirstReopenDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: اولین تاریخ Reopen',
     '''Query مورد نیاز:
SELECT [System.Id],
       MIN([System.ChangedDate]) as FirstReopenDate
FROM WorkItemRevisions
WHERE [System.Reason] = 'Reopen'
GROUP BY [System.Id]'''],

    ['LastReopenDate', 'تاریخ', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: آخرین تاریخ Reopen',
     '''Query مورد نیاز:
SELECT [System.Id],
       MAX([System.ChangedDate]) as LastReopenDate
FROM WorkItemRevisions
WHERE [System.Reason] = 'Reopen'
GROUP BY [System.Id]'''],

    ['StateTransitionCount', 'عدد', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تعداد تغییرات State',
     '''Query مورد نیاز:
SELECT [System.Id],
       COUNT(DISTINCT [System.State]) as StateTransitionCount
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
GROUP BY [System.Id]

فعلاً = 2 (پیش‌فرض حداقل)'''],

    ['StateChangeCount', 'عدد', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تعداد تغییرات State (مشابه بالا)',
     '''همان StateTransitionCount
فعلاً = 2'''],

    ['AssigneeChangeCount', 'عدد', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تعداد تغییرات مسئول',
     '''Query مورد نیاز:
SELECT [System.Id],
       COUNT(DISTINCT [System.AssignedTo]) - 1 as AssigneeChangeCount
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
GROUP BY [System.Id]

فعلاً = 0'''],

    ['StateHistory', 'متن', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: تاریخچه تغییرات State',
     '''Query مورد نیاز:
SELECT [System.Id],
       STRING_AGG([System.State] + '|' +
                  CONVERT(VARCHAR, [System.ChangedDate], 120), '; ')
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
GROUP BY [System.Id]

خروجی مثال: "New|2025-01-01; Active|2025-01-05; Closed|2025-01-10"'''],

    ['PreviousState', 'متن', '🟡 زرد', 'WorkItemRevisions', '⚠️ MOCK: State قبلی',
     '''Query مورد نیاز:
SELECT w1.[System.Id], w2.[System.State] as PreviousState
FROM WorkItemRevisions w1
LEFT JOIN WorkItemRevisions w2
  ON w1.[System.Id] = w2.[System.Id]
  AND w2.[System.Rev] = (SELECT MAX([System.Rev])
                          FROM WorkItemRevisions
                          WHERE [System.Id] = w1.[System.Id]
                            AND [System.Rev] < w1.[System.Rev])
WHERE w1.[System.Rev] = (SELECT MAX([System.Rev])
                          FROM WorkItemRevisions
                          WHERE [System.Id] = w1.[System.Id])'''],

    ['is_escaped', 'عدد', '🟡 زرد', 'محاسبه دستی', '⚠️ MOCK: آیا از Test فرار کرده؟',
     '''محاسبه:
اگر باگ در Production پیدا شده (نه در Test Environment)

فعلاً = 0 (برای همه)
نیاز به تعریف دقیق: چطور تشخیص بدیم Escaped است؟
- از Environment Tag?
- از Found In Release vs Target Release?'''],

    # === فیلدهای نارنجی: محاسبه‌ای از دیتای موجود ===
    ['IsDuplicate', 'عدد', '🟠 نارنجی', 'محاسبه', 'آیا Duplicate است؟ (1=بله, 0=خیر)',
     '''فرمول:
IF(CloseReason = "Duplicate", 1, 0)

محاسبه شده از CloseReason'''],

    ['LeadTimeHrs', 'عدد', '🟠 نارنجی', 'محاسبه', 'زمان کل از ایجاد تا بستن (ساعت)',
     '''فرمول:
LeadTime = (ClosedDate - CreatedDate) بر حسب ساعت

مثال: اگر Created = 2025-01-01 و Closed = 2025-01-03
LeadTime = 48 ساعت

⚠️ فعلاً از CreatedDate تقریبی استفاده شده'''],

    ['CycleTimeHrs', 'عدد', '🟠 نارنجی', 'محاسبه', 'زمان واقعی کار (بدون انتظار)',
     '''فرمول:
CycleTime = (ClosedDate - FirstActiveDate) بر حسب ساعت

یا تقریب: CycleTime ≈ LeadTime × 0.7

⚠️ فعلاً 70% از LeadTime محاسبه شده'''],

    ['AgeDays', 'عدد', '🟠 نارنجی', 'محاسبه', 'سن باگ (برای باگ‌های باز)',
     '''فرمول:
برای باگ‌های باز (Open, Active):
AgeDays = (امروز - CreatedDate) بر حسب روز

برای باگ‌های بسته: NULL'''],

    ['TriageDurationHrs', 'عدد', '🟠 نارنجی', 'محاسبه', '⚠️ MOCK: مدت زمان در Triage',
     '''فرمول:
(AssignedDate - TriageDate) بر حسب ساعت

⚠️ نیاز به WorkItemRevisions
فعلاً = 0'''],

    ['ReadyForRetestDurationHrs', 'عدد', '🟠 نارنجی', 'محاسبه', '⚠️ MOCK: مدت زمان در Ready for Retest',
     '''فرمول:
(VerifiedDate - ReadyForRetestDate) بر حسب ساعت

⚠️ نیاز به WorkItemRevisions
فعلاً = 0'''],

    ['WaitTimeHrs', 'عدد', '🟠 نارنجی', 'محاسبه', '⚠️ MOCK: زمان انتظار کل',
     '''فرمول:
LeadTime - CycleTime = زمان انتظار

⚠️ نیاز به WorkItemRevisions دقیق
فعلاً = 0'''],

    ['ResponseTimeHrs', 'عدد', '🟠 نارنجی', 'محاسبه', '⚠️ MOCK: زمان پاسخ اولیه',
     '''فرمول:
(AssignedDate - CreatedDate) بر حسب ساعت

⚠️ نیاز به CreatedDate واقعی
فعلاً = 24 (پیش‌فرض 1 روز)'''],

    # === فیلدهای آبی: دستی / از منابع خارجی ===
    ['FixEffortHrs', 'عدد', '🔵 آبی', 'Related Work Items', '⚠️ MOCK: زمان رفع باگ توسط Developer',
     '''نحوه دریافت:
از Related Tasks این باگ:

SELECT Child.[System.Id] as BugID,
       SUM(Child.[Microsoft.VSTS.Scheduling.CompletedWork]) as FixEffortHrs
FROM WorkItemLinks
INNER JOIN WorkItems Parent ON Parent.[System.Id] = WorkItemLinks.[System.Links.LinkType.ForwardEnd.Id]
INNER JOIN WorkItems Child ON Child.[System.Id] = WorkItemLinks.[System.Links.LinkType.ReverseEnd.Id]
WHERE Parent.[System.WorkItemType] = 'Bug'
  AND Child.[System.WorkItemType] = 'Task'
  AND WorkItemLinks.[System.Links.LinkType] = 'System.LinkTypes.Related'
GROUP BY Child.[System.Id]

یا از CompletedWork فیلد خود Bug (اگر تیم Effort ثبت کنه)

فعلاً = 0 (باید از Query بالا گرفته شود)'''],

    ['ModuleName', 'متن', '🔵 آبی', 'ورود دستی', 'نام ماژول/کامپوننت کدی که باگ در آن است',
     '''این فیلد باید توسط Developer یا QA به صورت دستی پر شود.

گزینه‌های پیشنهادی:
- ایجاد Custom Field در Azure DevOps: "Module"
- یا استفاده از Area Path
- یا Tag استفاده کنید

فعلاً = "N/A"'''],

    ['RootCause', 'متن', '🔵 آبی', 'ورود دستی', 'علت ریشه‌ای باگ (توصیفی)',
     '''این فیلد توصیفی است و باید توسط Developer پر شود.

مثال‌ها:
- "Null Reference Exception"
- "Logic Error in Calculation"
- "Missing Validation"
- "Integration Issue"

فعلاً = "N/A"'''],

    ['TestCaseID', 'متن', '🔵 آبی', 'Related Work Items', 'شناسه Test Case مرتبط',
     '''نحوه دریافت:
از Related Test Cases:

SELECT Parent.[System.Id] as BugID,
       Child.[System.Id] as TestCaseID
FROM WorkItemLinks
WHERE Parent.[System.WorkItemType] = 'Bug'
  AND Child.[System.WorkItemType] = 'Test Case'
  AND WorkItemLinks.[System.Links.LinkType] = 'System.LinkTypes.TestedBy'

فعلاً = "" (خالی)'''],
]

# نوشتن داده‌ها
for row_idx, row_data in enumerate(field_definitions, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_def.cell(row=row_idx, column=col_idx, value=value)

        # فرمت ردیف هدر
        if row_idx == 1:
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)

            # اعمال رنگ بر اساس ستون "رنگ"
            if col_idx == 1:  # ستون نام فیلد
                cell.font = Font(bold=True, size=10)

# تنظیم عرض ستون‌ها
ws_def.column_dimensions['A'].width = 25  # نام فیلد
ws_def.column_dimensions['B'].width = 12  # نوع دیتا
ws_def.column_dimensions['C'].width = 12  # رنگ
ws_def.column_dimensions['D'].width = 25  # منبع داده
ws_def.column_dimensions['E'].width = 40  # توضیحات فارسی
ws_def.column_dimensions['F'].width = 60  # نحوه دریافت دیتا

# تنظیم ارتفاع ردیف‌ها
for row_idx in range(2, len(field_definitions) + 1):
    ws_def.row_dimensions[row_idx].height = 80

print("   ✅ شیت راهنمای فارسی ساخته شد")

# ذخیره نهایی
wb.save(output_file)
print(f"\n✅ فایل نهایی ذخیره شد: {output_file}")

# ============================================================================
# گزارش نهایی
# ============================================================================

print("\n" + "=" * 80)
print("✅ فایل Excel نهایی آماده شد!")
print("=" * 80)

print(f"""
📊 آمار:
   - تعداد باگ‌ها:        {len(df_clean)}
   - تعداد فیلدها:        {len(df_clean.columns)}

✅ دسته‌بندی فیلدها:
   - 🟢 سبز (مستقیم از CSV):              18 فیلد
   - 🟡 زرد (نیاز به WorkItemRevisions):  16 فیلد (MOCK)
   - 🟠 نارنجی (محاسبه‌ای):               8 فیلد
   - 🔵 آبی (دستی/Related Items):         4 فیلد

📁 فایل خروجی:
   {output_file}

📋 شیت‌ها:
   1. راهنمای_فیلدها - توضیحات کامل فارسی + نحوه دریافت دیتا
   2. raw_data         - دیتای 820 باگ با 44 فیلد

⚠️  فیلدهای MOCK (نیاز به Query اضافی):
   - CreatedDate, AssignedDate, TriageDate, InProgressDate
   - ReopenCount, FirstReopenDate, LastReopenDate
   - StateTransitionCount, StateHistory, PreviousState
   - و 6 فیلد دیگر

   👉 همه Query های لازم در شیت "راهنمای_فیلدها" نوشته شده است.

✅ آماده تحویل به پیاده‌ساز!
""")

print("=" * 80)

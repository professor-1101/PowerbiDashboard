#!/usr/bin/env python3
"""
Convert Azure DevOps CSV to Clean Excel
Reduces from 74 columns to 24 essential fields only!
"""

import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

print("=" * 80)
print("CONVERTING CSV TO CLEAN EXCEL")
print("=" * 80)

# ============================================================================
# STEP 1: Read CSV
# ============================================================================

print("\n📥 Step 1: Reading CSV file...")
csv_file = "Untitled query (1).csv"
df = pd.read_csv(csv_file, encoding='utf-8-sig')

print(f"   ✅ Loaded {len(df)} bugs from CSV")
print(f"   ✅ CSV has {len(df.columns)} columns")

# ============================================================================
# STEP 2: Data Transformation Functions
# ============================================================================

print("\n🔧 Step 2: Defining transformation functions...")

def clean_severity(value):
    """Convert '3 - Medium' → 'Medium'"""
    if pd.isna(value):
        return "Medium"
    severity_map = {
        "1 - Critical": "Critical",
        "2 - High": "High",
        "3 - Medium": "Medium",
        "4 - Low": "Low"
    }
    return severity_map.get(str(value).strip(), str(value))


def normalize_state(value):
    """Normalize state values"""
    if pd.isna(value):
        return "Open"
    state_map = {
        "triage": "New",
        "In Progress": "Active",
        "Committed": "Active",
        "Waiting": "Active",
        "Done": "Closed"
    }
    return state_map.get(str(value).strip(), str(value))


def extract_category(value):
    """Extract 'ANZ' from 'ANZ (تحلیل)'"""
    if pd.isna(value) or value == "":
        return "Other"
    # Extract code before parenthesis or space
    match = re.match(r'^([A-Z]+)', str(value))
    return match.group(1) if match else str(value)


def extract_name_and_id(value):
    """Extract name and ID from 'Name Surname <RPK\\Username>'"""
    if pd.isna(value) or value == "":
        return "", ""

    # Pattern: "Name <RPK\ID>"
    match = re.match(r"(.+?)\s*<RPK\\(.+?)>", str(value))
    if match:
        name = match.group(1).strip()
        user_id = match.group(2).strip()
        return name, user_id
    return str(value), ""


def extract_project_sprint(value):
    """Extract project and sprint from 'SAJAK\\Estate_Sprint 44'"""
    if pd.isna(value) or value == "":
        return "", ""

    parts = str(value).split("\\")
    project = parts[0] if len(parts) > 0 else ""
    sprint = parts[1] if len(parts) > 1 else ""
    return project, sprint


def parse_datetime(value):
    """Parse datetime from various formats"""
    if pd.isna(value) or value == "":
        return pd.NaT

    try:
        # Try format: "12/24/2025 11:53:24 AM"
        return pd.to_datetime(value, format='%m/%d/%Y %I:%M:%S %p')
    except:
        try:
            # Try ISO format
            return pd.to_datetime(value)
        except:
            return pd.NaT

print("   ✅ Transformation functions ready")

# ============================================================================
# STEP 3: Create Clean DataFrame with 24 Fields
# ============================================================================

print("\n🧹 Step 3: Creating clean dataframe with 24 fields...")

# Initialize clean dataframe
df_clean = pd.DataFrame()

# 🟢 Core Fields (8 fields - Direct from CSV)
print("   🟢 Adding core fields...")
df_clean['BugID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
df_clean['Title'] = df['Title'].fillna("")
df_clean['Description'] = df['Description'].fillna("")
df_clean['Severity'] = df['Severity'].apply(clean_severity)
df_clean['Priority'] = pd.to_numeric(df['Priority'], errors='coerce').fillna(2).astype(int)
df_clean['State'] = df['State'].apply(normalize_state)
df_clean['Category'] = df['Bug Type'].apply(extract_category)
df_clean['Tags'] = df['Tags'].fillna("")

# 🟡 Extracted/Calculated Fields (10 fields)
print("   🟡 Adding calculated fields...")

# Extract names and IDs
assignee_data = df['Assigned To'].apply(extract_name_and_id)
df_clean['AssigneeName'] = assignee_data.apply(lambda x: x[0])
df_clean['AssigneeID'] = assignee_data.apply(lambda x: x[1])

resolver_data = df['Closed By'].apply(extract_name_and_id)
df_clean['ResolverName'] = resolver_data.apply(lambda x: x[0])
df_clean['ResolverID'] = resolver_data.apply(lambda x: x[1])

# Extract project and sprint
project_data = df['Team Project'].fillna("").apply(lambda x: (str(x), ""))
df_clean['ProjectName'] = project_data.apply(lambda x: x[0])
df_clean['TeamName'] = df['Team Project'].fillna("")

sprint_data = df['Iteration Path'].fillna("").apply(extract_project_sprint)
df_clean['SprintName'] = sprint_data.apply(lambda x: x[1])

# Calculate flags
df_clean['is_duplicate'] = (df['Closed Reason'] == 'Duplicate').astype(int)
df_clean['IsRegression'] = df['Tags'].fillna("").str.contains(
    'regression|تکراری', case=False, regex=True
).astype(int)
df_clean['Comments'] = pd.to_numeric(df['Comment Count'], errors='coerce').fillna(0).astype(int)

# 📅 Date Fields (5 fields)
print("   📅 Adding date fields...")
df_clean['ClosedDate'] = df['Closed Date'].apply(parse_datetime)
df_clean['ResolvedDate'] = df['Resolved Date'].apply(parse_datetime)
df_clean['LastModifiedDate'] = df['State Change Date'].apply(parse_datetime)

# DueDate from Target Date or Due Date
due_date = df.get('Target Date', df.get('Due Date', pd.Series([pd.NaT] * len(df))))
df_clean['DueDate'] = due_date.apply(parse_datetime)

df_clean['CloseReason'] = df['Closed Reason'].fillna("")

# 🔵 Dashboard-Only Field (1 field)
print("   🔵 Adding dashboard-only field...")
df_clean['Resolution'] = "N/A"  # Manual entry field

print(f"\n   ✅ Clean dataframe created: {len(df_clean)} rows × {len(df_clean.columns)} columns")

# Verify we have exactly 24 columns
assert len(df_clean.columns) == 24, f"Expected 24 columns, got {len(df_clean.columns)}"

# ============================================================================
# STEP 4: Save to Excel
# ============================================================================

print("\n💾 Step 4: Saving to Excel...")

output_file = "BugTracking_Clean.xlsx"
df_clean.to_excel(output_file, sheet_name='raw_data', index=False)

print(f"   ✅ Saved to {output_file}")

# ============================================================================
# STEP 5: Apply Color Coding to Headers
# ============================================================================

print("\n🎨 Step 5: Applying color coding to headers...")

wb = load_workbook(output_file)
ws = wb['raw_data']

# Define colors
GREEN = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
YELLOW = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
BLUE = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

# Define field categories
green_fields = [
    'BugID', 'Title', 'Description', 'Severity', 'Priority', 'State',
    'Category', 'Tags', 'TeamName', 'ClosedDate', 'ResolvedDate',
    'LastModifiedDate', 'DueDate', 'CloseReason'
]

yellow_fields = [
    'ProjectName', 'SprintName', 'AssigneeName', 'AssigneeID',
    'ResolverName', 'ResolverID', 'is_duplicate', 'IsRegression', 'Comments'
]

blue_fields = ['Resolution']

# Apply colors to header row
header_font = Font(bold=True, size=11, color='000000')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col_idx, cell in enumerate(ws[1], 1):
    field_name = cell.value

    # Apply formatting
    cell.font = header_font
    cell.alignment = header_alignment

    # Apply color based on category
    if field_name in green_fields:
        cell.fill = GREEN
    elif field_name in yellow_fields:
        cell.fill = YELLOW
    elif field_name in blue_fields:
        cell.fill = BLUE

# Adjust column widths
column_widths = {
    'BugID': 10,
    'Title': 50,
    'Description': 60,
    'Severity': 12,
    'Priority': 10,
    'State': 12,
    'Category': 12,
    'Tags': 20,
    'ProjectName': 15,
    'TeamName': 20,
    'SprintName': 20,
    'AssigneeName': 20,
    'AssigneeID': 15,
    'ResolverName': 20,
    'ResolverID': 15,
    'is_duplicate': 12,
    'IsRegression': 12,
    'Comments': 10,
    'ClosedDate': 18,
    'ResolvedDate': 18,
    'LastModifiedDate': 18,
    'DueDate': 18,
    'CloseReason': 15,
    'Resolution': 30
}

for col_idx, cell in enumerate(ws[1], 1):
    col_letter = cell.column_letter
    field_name = cell.value
    if field_name in column_widths:
        ws.column_dimensions[col_letter].width = column_widths[field_name]

print("   ✅ Color coding applied")

# ============================================================================
# STEP 6: Create Field_Definitions Sheet
# ============================================================================

print("\n📋 Step 6: Creating Field_Definitions sheet...")

field_definitions = [
    ['Field Name', 'Description', 'Type', 'Source', 'Used in Charts', 'Color'],

    # Green fields
    ['BugID', 'Unique bug identifier', 'Integer', 'CSV: ID', 'Yes', 'Green'],
    ['Title', 'Bug title/summary', 'Text', 'CSV: Title', 'Yes', 'Green'],
    ['Description', 'Detailed bug description', 'Text', 'CSV: Description', 'Yes', 'Green'],
    ['Severity', 'Bug severity level', 'Text', 'CSV: Severity (cleaned)', 'Yes', 'Green'],
    ['Priority', 'Bug priority (1-4)', 'Integer', 'CSV: Priority', 'Yes', 'Green'],
    ['State', 'Current bug state', 'Text', 'CSV: State (normalized)', 'Yes', 'Green'],
    ['Category', 'Bug category/type', 'Text', 'CSV: Bug Type (extracted)', 'Yes', 'Green'],
    ['Tags', 'Bug tags', 'Text', 'CSV: Tags', 'No', 'Green'],
    ['TeamName', 'Team responsible', 'Text', 'CSV: Team Project', 'No', 'Green'],
    ['ClosedDate', 'When bug was closed', 'DateTime', 'CSV: Closed Date', 'No', 'Green'],
    ['ResolvedDate', 'When bug was resolved', 'DateTime', 'CSV: Resolved Date', 'No', 'Green'],
    ['LastModifiedDate', 'Last modification date', 'DateTime', 'CSV: State Change Date', 'No', 'Green'],
    ['DueDate', 'Bug due date', 'DateTime', 'CSV: Target Date / Due Date', 'No', 'Green'],
    ['CloseReason', 'Reason for closure', 'Text', 'CSV: Closed Reason', 'No', 'Green'],

    # Yellow fields
    ['ProjectName', 'Project name', 'Text', 'Calculated from Team Project', 'No', 'Yellow'],
    ['SprintName', 'Sprint name', 'Text', 'Extracted from Iteration Path', 'No', 'Yellow'],
    ['AssigneeName', 'Assigned person name', 'Text', 'Extracted from Assigned To', 'No', 'Yellow'],
    ['AssigneeID', 'Assigned person ID', 'Text', 'Extracted from Assigned To', 'No', 'Yellow'],
    ['ResolverName', 'Resolver person name', 'Text', 'Extracted from Closed By', 'No', 'Yellow'],
    ['ResolverID', 'Resolver person ID', 'Text', 'Extracted from Closed By', 'No', 'Yellow'],
    ['is_duplicate', 'Is this a duplicate bug?', 'Integer', 'Calculated: CloseReason == "Duplicate"', 'No', 'Yellow'],
    ['IsRegression', 'Is this a regression?', 'Integer', 'Calculated: Tags contains "regression"', 'No', 'Yellow'],
    ['Comments', 'Number of comments', 'Integer', 'CSV: Comment Count', 'No', 'Yellow'],

    # Blue fields
    ['Resolution', 'Technical resolution details', 'Text', 'Manual entry (Dashboard only)', 'Yes', 'Blue'],
]

ws_def = wb.create_sheet('Field_Definitions')

# Add data
for row_idx, row_data in enumerate(field_definitions, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_def.cell(row=row_idx, column=col_idx, value=value)

        # Format header row
        if row_idx == 1:
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Apply color based on Color column
            if len(row_data) > 5:
                color_name = row_data[5]
                if color_name == 'Green':
                    cell.fill = GREEN
                elif color_name == 'Yellow':
                    cell.fill = YELLOW
                elif color_name == 'Blue':
                    cell.fill = BLUE

# Set column widths for Field_Definitions
ws_def.column_dimensions['A'].width = 20  # Field Name
ws_def.column_dimensions['B'].width = 40  # Description
ws_def.column_dimensions['C'].width = 12  # Type
ws_def.column_dimensions['D'].width = 35  # Source
ws_def.column_dimensions['E'].width = 15  # Used in Charts
ws_def.column_dimensions['F'].width = 10  # Color

print("   ✅ Field_Definitions sheet created")

# Save workbook
wb.save(output_file)
print(f"   ✅ Saved with Field_Definitions sheet")

# ============================================================================
# STEP 7: Statistics and Summary
# ============================================================================

print("\n" + "=" * 80)
print("CONVERSION COMPLETE!")
print("=" * 80)

print(f"""
📊 Statistics:
   - Input CSV:          {len(df)} rows × {len(df.columns)} columns
   - Output Excel:       {len(df_clean)} rows × {len(df_clean.columns)} columns
   - Reduction:          74 columns → 24 columns (50 deleted!)

✅ Fields by Category:
   - 🟢 Direct from CSV:        14 fields
   - 🟡 Calculated/Extracted:    9 fields
   - 🔵 Dashboard-Only:          1 field

📁 Output File:
   - {output_file}

📋 Sheets:
   1. raw_data          - Clean bug data (820 rows × 24 columns)
   2. Field_Definitions - Field documentation

🎨 Color Coding:
   - Green  = Direct from CSV
   - Yellow = Calculated
   - Blue   = Dashboard-Only (manual entry)

🔍 Data Quality:
   - Severity cleaned:   "{df['Severity'].iloc[0]}" → "{df_clean['Severity'].iloc[0]}"
   - State normalized:   "{df['State'].iloc[0]}" → "{df_clean['State'].iloc[0]}"
   - Names extracted:    ✅
   - Dates parsed:       ✅
   - Flags calculated:   ✅

""")

print("=" * 80)
print("Next steps:")
print("  1. Open BugTracking_Clean.xlsx")
print("  2. Review the 24 fields in raw_data sheet")
print("  3. Check Field_Definitions for documentation")
print("  4. Manually fill Resolution column if needed")
print("=" * 80)

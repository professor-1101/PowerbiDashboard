#!/usr/bin/env python3
"""
Check original BugTracking_Complete.xlsx structure
"""

from openpyxl import load_workbook

print("=" * 80)
print("بررسی ساختار BugTracking_Complete.xlsx اصلی")
print("=" * 80)

wb = load_workbook('BugTracking_Complete.xlsx', data_only=False)

# Check raw_data fields
print("\n📋 فیلدهای raw_data در فایل اصلی:")
ws = wb['raw_data']
headers = []
for col in range(1, min(ws.max_column + 1, 100)):
    cell_value = ws.cell(1, col).value
    if cell_value:
        headers.append((col, cell_value))
    else:
        break

for col_num, header in headers:
    print(f"   {col_num:2d}. {header}")

print(f"\n   تعداد کل: {len(headers)} فیلد")
print(f"   تعداد ردیف: {ws.max_row} (شامل هدر)")

# Check all sheets
print(f"\n📑 تمام شیت‌ها ({len(wb.sheetnames)}):")
for i, sheet in enumerate(wb.sheetnames, 1):
    ws = wb[sheet]
    chart_count = len(ws._charts) if hasattr(ws, '_charts') else 0
    print(f"   {i:2d}. {sheet:30s} - {chart_count} چارت")

# Count formulas that reference raw_data
print(f"\n🔍 بررسی فرمول‌های raw_data...")

import re
pattern = re.compile(r"raw_data!\$?([A-Z]+)\$?(\d+):?\$?([A-Z]+)?\$?(\d+)?")

dashboard_sheets = [
    'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance',
    'Sprint_Analysis', 'Time_Flow', 'Quality_Analysis',
    'State_Flow', 'Resolution_Analysis', 'Module_Project',
    'Workload_Analysis', 'Trend_Analysis', 'KPIs_Detail'
]

fields_referenced = {}

for sheet_name in dashboard_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'raw_data' in cell.value:
                # Extract column references
                matches = pattern.findall(cell.value)
                for match in matches:
                    col_letter = match[0]
                    if col_letter not in fields_referenced:
                        fields_referenced[col_letter] = []
                    fields_referenced[col_letter].append(f"{sheet_name}!{cell.coordinate}")

print(f"\n   ستون‌های استفاده‌شده از raw_data:")
for col_letter in sorted(fields_referenced.keys()):
    # Find column number
    from openpyxl.utils import column_index_from_string
    col_num = column_index_from_string(col_letter)
    field_name = ws.cell(1, col_num).value if col_num <= len(headers) else "?"
    
    # Get field name from raw_data
    ws_raw = wb['raw_data']
    actual_field = ws_raw.cell(1, col_num).value
    
    usage_count = len(fields_referenced[col_letter])
    print(f"      {col_letter:3s} (col {col_num:2d}): {actual_field:25s} - {usage_count:3d} استفاده")

print("\n" + "=" * 80)

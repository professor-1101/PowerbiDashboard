#!/usr/bin/env python3
"""
Add راهنمای_فیلدها sheet with complete Persian documentation
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

print("=" * 80)
print("اضافه کردن شیت راهنمای_فیلدها")
print("=" * 80)

# Load Excel
wb = load_workbook('BugTracking_Complete_FINAL.xlsx')

# Check if sheet exists and delete it
if 'راهنمای_فیلدها' in wb.sheetnames:
    del wb['راهنمای_فیلدها']
    print("   شیت قدیمی حذف شد")

# Create new sheet at beginning
ws = wb.create_sheet('راهنمای_فیلدها', 0)
print("   ✅ شیت جدید ایجاد شد")

# ============================================================================
# Define Styles
# ============================================================================

# Title style
title_font = Font(name='B Nazanin', size=18, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Header style
header_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Color styles
green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
orange_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
blue_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

# Cell style
cell_font = Font(name='B Nazanin', size=11)
cell_align = Alignment(horizontal='right', vertical='top', wrap_text=True)

# Border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# Add Title
# ============================================================================

ws.merge_cells('A1:G1')
ws['A1'] = 'راهنمای فیلدهای داشبورد ردیابی باگ - Azure DevOps'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = title_align
ws.row_dimensions[1].height = 30

# ============================================================================
# Add Instructions
# ============================================================================

ws.merge_cells('A2:G2')
ws['A2'] = '''این فایل حاوی 821 باگ از Azure DevOps است. فیلدها به 4 دسته تقسیم شده‌اند:
🟢 سبز: مستقیم از CSV | 🟡 زرد: نیاز به کوئری WorkItemRevisions | 🟠 نارنجی: محاسبه‌شده | 🔵 آبی: ورودی دستی'''
ws['A2'].font = Font(name='B Nazanin', size=10)
ws['A2'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
ws.row_dimensions[2].height = 40

# ============================================================================
# Add Headers
# ============================================================================

headers = [
    'نام فیلد (EN)',
    'نام فارسی',
    'نوع داده',
    'منبع داده',
    'کوئری SQL / فرمول',
    'توضیحات',
    'رنگ'
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(3, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.row_dimensions[3].height = 25

# ============================================================================
# Field Definitions
# ============================================================================

fields_data = [
    # GREEN FIELDS (19)
    ['BugID', 'شناسه باگ', 'Number', 'CSV: ID', 'مستقیم', 'شناسه یکتای باگ در Azure DevOps', '🟢'],
    ['Title', 'عنوان', 'Text', 'CSV: Title', 'مستقیم', 'عنوان باگ', '🟢'],
    ['Description', 'شرح', 'Text', 'CSV: Description', 'مستقیم', 'توضیحات کامل باگ', '🟢'],
    ['Severity', 'شدت', 'Text', 'CSV: Severity', 'clean_severity()', 'سطح شدت: Critical, High, Medium, Low', '🟢'],
    ['Priority', 'اولویت', 'Number', 'CSV: Priority', 'مستقیم', 'اولویت 1-4', '🟢'],
    ['State', 'وضعیت', 'Text', 'CSV: State', 'مستقیم', 'وضعیت فعلی باگ', '🟢'],
    ['Category', 'دسته‌بندی', 'Text', 'CSV: Bug Type', 'extract_category()', 'کد دسته‌بندی (ANZ, FN, PER, SEC, etc.)', '🟢'],
    ['Tags', 'برچسب‌ها', 'Text', 'CSV: Tags', 'مستقیم', 'برچسب‌های باگ', '🟢'],
    ['TeamName', 'نام تیم', 'Text', 'CSV: Team Project', 'مستقیم', 'نام تیم مسئول', '🟢'],
    ['ProjectName', 'نام پروژه', 'Text', 'CSV: Team Project', 'مستقیم', 'نام پروژه', '🟢'],
    ['SprintName', 'نام اسپرینت', 'Text', 'CSV: Iteration Path', 'split("\\\\")[1]', 'نام اسپرینت از Iteration Path', '🟢'],
    ['AssigneeName', 'مسئول فعلی', 'Text', 'CSV: Assigned To', 'extract_name()', 'نام مسئول باگ', '🟢'],
    ['ResolverName', 'حل‌کننده', 'Text', 'CSV: Closed By', 'extract_name()', 'نام کسی که باگ را حل کرد', '🟢'],
    ['ClosedDate', 'تاریخ بستن', 'DateTime', 'CSV: Closed Date', 'parse_datetime()', 'تاریخ بسته شدن باگ', '🟢'],
    ['ResolvedDate', 'تاریخ حل', 'DateTime', 'CSV: Resolved Date', 'parse_datetime()', 'تاریخ حل باگ', '🟢'],
    ['LastModifiedDate', 'آخرین تغییر', 'DateTime', 'CSV: State Change Date', 'parse_datetime()', 'تاریخ آخرین تغییر', '🟢'],
    ['DueDate', 'سررسید', 'DateTime', 'CSV: Target Date', 'parse_datetime()', 'تاریخ سررسید', '🟢'],
    ['CloseReason', 'دلیل بستن', 'Text', 'CSV: Closed Reason', 'مستقیم', 'دلیل بسته شدن (Fixed, Duplicate, etc.)', '🟢'],
    ['IsRegression', 'رگرسیون؟', 'Boolean', 'CSV: Tags', 'tags.contains("regression")', '1 اگر regression باشد', '🟢'],
    
    # YELLOW FIELDS (17) - Need WorkItemRevisions
    ['CreatedDate', 'تاریخ ایجاد', 'DateTime', 'WorkItemRevisions', '''SELECT [System.Id], MIN([System.ChangedDate]) as CreatedDate
FROM WorkItemRevisions
WHERE [System.WorkItemType] = 'Bug'
GROUP BY [System.Id]''', 'اولین تاریخ ثبت باگ (از جدول Revisions)', '🟡'],
    
    ['AssignedDate', 'تاریخ Assign', 'DateTime', 'WorkItemRevisions', '''SELECT [System.Id], [System.ChangedDate] as AssignedDate
FROM WorkItemRevisions
WHERE [System.State] = 'Assigned' AND [System.Rev] = 
  (SELECT MIN(Rev) FROM WorkItemRevisions WHERE State='Assigned')''', 'تاریخ اولین Assign', '🟡'],
    
    ['TriageDate', 'تاریخ Triage', 'DateTime', 'WorkItemRevisions', 'تاریخ ورود به وضعیت Triage', 'از WorkItemRevisions', '🟡'],
    ['StartedDate', 'تاریخ Started', 'DateTime', 'WorkItemRevisions', 'تاریخ Start کار', 'از WorkItemRevisions', '🟡'],
    ['InProgressDate', 'تاریخ In Progress', 'DateTime', 'WorkItemRevisions', 'تاریخ ورود به In Progress', 'از WorkItemRevisions', '🟡'],
    ['ReadyForRetestDate', 'تاریخ آماده تست', 'DateTime', 'WorkItemRevisions', 'تاریخ Ready for Retest', 'از WorkItemRevisions', '🟡'],
    ['VerifiedDate', 'تاریخ Verified', 'DateTime', 'WorkItemRevisions', 'تاریخ Verified شدن', 'از WorkItemRevisions', '🟡'],
    ['DoneDate', 'تاریخ Done', 'DateTime', 'WorkItemRevisions', 'تاریخ Done شدن', 'از WorkItemRevisions', '🟡'],
    
    ['ReopenCount', 'تعداد بازگشایی', 'Number', 'WorkItemRevisions', '''SELECT [System.Id], COUNT(*) as ReopenCount
FROM WorkItemRevisions
WHERE [System.Reason] = 'Reopen'
GROUP BY [System.Id]''', 'تعداد دفعات بازگشایی باگ', '🟡'],
    
    ['FirstReopenDate', 'اولین بازگشایی', 'DateTime', 'WorkItemRevisions', 'تاریخ اولین Reopen', 'از WorkItemRevisions', '🟡'],
    ['LastReopenDate', 'آخرین بازگشایی', 'DateTime', 'WorkItemRevisions', 'تاریخ آخرین Reopen', 'از WorkItemRevisions', '🟡'],
    ['StateTransitionCount', 'تعداد تغییر State', 'Number', 'WorkItemRevisions', 'COUNT(DISTINCT State)', 'تعداد تغییرات وضعیت', '🟡'],
    ['StateChangeCount', 'تعداد کل تغییرات', 'Number', 'WorkItemRevisions', 'COUNT(*)', 'تعداد کل تغییرات', '🟡'],
    ['AssigneeChangeCount', 'تغییر مسئول', 'Number', 'WorkItemRevisions', 'COUNT(DISTINCT AssignedTo)', 'تعداد تغییر مسئول', '🟡'],
    ['StateHistory', 'تاریخچه State', 'Text', 'WorkItemRevisions', 'STRING_AGG(State, " -> ")', 'تاریخچه کامل تغییرات وضعیت', '🟡'],
    ['PreviousState', 'وضعیت قبلی', 'Text', 'WorkItemRevisions', 'LAG(State) OVER (ORDER BY Rev)', 'وضعیت قبل از فعلی', '🟡'],
    ['is_escaped', 'Escaped؟', 'Boolean', 'WorkItemRevisions', 'آیا باگ از Dev Escape کرده', 'محاسبه از State transitions', '🟡'],
    
    # ORANGE FIELDS (16) - Calculated
    ['AssigneeID', 'شناسه مسئول', 'Text', 'محاسبه‌شده', 'extract_id(Assigned To)', 'استخراج ID از فیلد Assigned To', '🟠'],
    ['ResolverID', 'شناسه حل‌کننده', 'Text', 'محاسبه‌شده', 'extract_id(Closed By)', 'استخراج ID از فیلد Closed By', '🟠'],
    ['Comments', 'تعداد کامنت', 'Number', 'CSV: Comment Count', 'مستقیم', 'تعداد کامنت‌های باگ', '🟠'],
    
    ['LeadTimeHrs', 'زمان کل (ساعت)', 'Number', 'محاسبه‌شده', '=(ClosedDate - CreatedDate) * 24', 'زمان از ایجاد تا بستن (ساعت)', '🟠'],
    ['CycleTimeHrs', 'زمان چرخه (ساعت)', 'Number', 'محاسبه‌شده', '=(ClosedDate - InProgressDate) * 24', 'زمان از شروع تا بستن', '🟠'],
    ['AgeDays', 'سن (روز)', 'Number', 'محاسبه‌شده', '=TODAY() - CreatedDate', 'سن باگ‌های باز (روز)', '🟠'],
    
    ['TriageDurationHrs', 'مدت Triage', 'Number', 'محاسبه‌شده', '=(AssignedDate - TriageDate) * 24', 'مدت زمان در Triage', '🟠'],
    ['ActiveDurationHrs', 'مدت Active', 'Number', 'محاسبه‌شده', 'مدت زمان Active', 'از تاریخ‌های WorkItemRevisions', '🟠'],
    ['InProgressDurationHrs', 'مدت In Progress', 'Number', 'محاسبه‌شده', 'مدت زمان In Progress', 'از تاریخ‌های WorkItemRevisions', '🟠'],
    ['ReadyForRetestDurationHrs', 'مدت Ready for Retest', 'Number', 'محاسبه‌شده', 'مدت زمان Ready for Retest', 'از تاریخ‌های WorkItemRevisions', '🟠'],
    ['ResponseTimeHrs', 'زمان پاسخ', 'Number', 'محاسبه‌شده', '=(AssignedDate - CreatedDate) * 24', 'زمان تا Assign شدن', '🟠'],
    ['WaitTimeHrs', 'زمان انتظار', 'Number', 'محاسبه‌شده', 'مجموع زمان‌های انتظار', 'محاسبه از State transitions', '🟠'],
    ['ActiveWorkTimeHrs', 'زمان کار فعال', 'Number', 'محاسبه‌شده', 'زمان واقعی کار', 'LeadTime - WaitTime', '🟠'],
    
    ['IsDuplicate', 'تکراری؟', 'Boolean', 'محاسبه‌شده', '=IF(CloseReason="Duplicate",1,0)', '1 اگر دلیل بستن Duplicate باشد', '🟠'],
    ['FixAttempts', 'تعداد تلاش رفع', 'Number', 'محاسبه‌شده', 'ReopenCount + 1', 'تعداد دفعات تلاش برای رفع', '🟠'],
    
    ['FixEffortHrs', 'زمان رفع (ساعت)', 'Number', 'Related Tasks', '''SELECT wi.[System.Id], SUM(rel.[Original Estimate]) as FixEffortHrs
FROM WorkItems wi
LEFT JOIN WorkItemLinks wil ON wi.Id = wil.SourceId
LEFT JOIN WorkItems rel ON wil.TargetId = rel.Id
WHERE wil.LinkType = 'Related' AND rel.WorkItemType = 'Task'
GROUP BY wi.[System.Id]''', 'مجموع Original Estimate تسک‌های مرتبط', '🟠'],
    
    # BLUE FIELDS (22) - Manual Entry
    ['Resolution', 'نحوه رفع', 'Text', 'ورودی دستی', 'ورود توسط تیم', 'توضیحات نحوه رفع باگ (Code Fix, Config Change, etc.)', '🔵'],
    ['ModuleName', 'نام ماژول', 'Text', 'ورودی دستی', 'ورود توسط تیم', 'نام ماژول یا کامپوننت', '🔵'],
    ['RootCause', 'علت اصلی', 'Text', 'ورودی دستی', 'ورود توسط تیم', 'علت ریشه‌ای باگ (Code Bug, Requirements, etc.)', '🔵'],
    ['TestCaseID', 'شناسه Test Case', 'Text', 'ورودی دستی', 'ورود توسط تیم', 'شناسه Test Case مرتبط', '🔵'],
    
    ['AnalysisEffortHrs', 'زمان تحلیل', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'ساعت صرف‌شده برای تحلیل', '🔵'],
    ['DevEffortHrs', 'زمان توسعه', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'ساعت صرف‌شده برای کدنویسی', '🔵'],
    ['TestEffortHrs', 'زمان تست', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'ساعت صرف‌شده برای تست', '🔵'],
    ['ReopenEffortHrs', 'زمان Reopen', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'ساعت صرف‌شده بعد از Reopen', '🔵'],
    ['TotalEffortHrs', 'مجموع زمان', 'Number', 'محاسبه‌شده', '=SUM(Analysis+Dev+Fix+Test+Reopen)', 'مجموع کل ساعات', '🔵'],
    ['EstimatedEffortHrs', 'تخمین زمان', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'تخمین اولیه زمان', '🔵'],
    
    ['VerifierName', 'نام تست‌کننده', 'Text', 'Work Item Details', 'دریافت از Azure DevOps', 'نام کسی که باگ را Verify کرد', '🔵'],
    ['VerifierID', 'شناسه تست‌کننده', 'Text', 'Work Item Details', 'دریافت از Azure DevOps', 'شناسه Verifier', '🔵'],
    ['ReporterName', 'نام گزارش‌دهنده', 'Text', 'Work Item Details', 'Created By', 'نام کسی که باگ را ثبت کرد', '🔵'],
    ['ReporterID', 'شناسه گزارش‌دهنده', 'Text', 'Work Item Details', 'Created By ID', 'شناسه Reporter', '🔵'],
    
    ['DuplicateOfBugID', 'تکراری از', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'اگر تکراری است، شناسه باگ اصلی', '🔵'],
    ['RetestPassCount', 'تعداد Retest موفق', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'تعداد دفعات Retest موفق', '🔵'],
    ['RetestFailCount', 'تعداد Retest ناموفق', 'Number', 'ورودی دستی', 'ورود توسط تیم', 'تعداد دفعات Retest ناموفق', '🔵'],
    
    ['ExternalTicketID', 'شناسه تیکت خارجی', 'Text', 'ورودی دستی', 'ورود توسط تیم', 'شناسه تیکت در سیستم خارجی (Jira, etc.)', '🔵'],
    ['ProjectID', 'شناسه پروژه', 'Text', 'N/A', '-', 'شناسه پروژه (اختیاری)', '🔵'],
    ['TeamID', 'شناسه تیم', 'Text', 'N/A', '-', 'شناسه تیم (اختیاری)', '🔵'],
    ['ModuleID', 'شناسه ماژول', 'Text', 'N/A', '-', 'شناسه ماژول (اختیاری)', '🔵'],
    ['SprintID', 'شناسه اسپرینت', 'Text', 'N/A', '-', 'شناسه اسپرینت (اختیاری)', '🔵'],
]

# Add data
for row_idx, field_data in enumerate(fields_data, 4):
    for col_idx, value in enumerate(field_data, 1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.font = cell_font
        cell.alignment = cell_align
        cell.border = thin_border
        
        # Apply color based on category
        color_indicator = field_data[-1]
        if color_indicator == '🟢':
            cell.fill = green_fill
        elif color_indicator == '🟡':
            cell.fill = yellow_fill
        elif color_indicator == '🟠':
            cell.fill = orange_fill
        elif color_indicator == '🔵':
            cell.fill = blue_fill

print(f"   ✅ {len(fields_data)} فیلد اضافه شد")

# ============================================================================
# Set Column Widths
# ============================================================================

ws.column_dimensions['A'].width = 25  # Field Name
ws.column_dimensions['B'].width = 20  # Persian Name
ws.column_dimensions['C'].width = 12  # Data Type
ws.column_dimensions['D'].width = 20  # Source
ws.column_dimensions['E'].width = 50  # SQL/Formula
ws.column_dimensions['F'].width = 40  # Description
ws.column_dimensions['G'].width = 8   # Color

print("   ✅ عرض ستون‌ها تنظیم شد")

# ============================================================================
# Add Summary Section
# ============================================================================

summary_row = len(fields_data) + 5

ws.merge_cells(f'A{summary_row}:G{summary_row}')
ws[f'A{summary_row}'] = '''خلاصه:
🟢 19 فیلد سبز: مستقیماً از CSV موجود است
🟡 17 فیلد زرد: نیاز به کوئری WorkItemRevisions دارد
🟠 16 فیلد نارنجی: قابل محاسبه از داده‌های موجود
🔵 22 فیلد آبی: نیاز به ورود دستی یا کوئری‌های اضافی

برای دریافت فیلدهای زرد، از WIQL استفاده کنید و جدول WorkItemRevisions را کوئری بزنید.'''
ws[f'A{summary_row}'].font = Font(name='B Nazanin', size=11, bold=True)
ws[f'A{summary_row}'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
ws.row_dimensions[summary_row].height = 100

print("   ✅ خلاصه اضافه شد")

# ============================================================================
# Save
# ============================================================================

wb.save('BugTracking_Complete_FINAL.xlsx')

import os
size_kb = os.path.getsize('BugTracking_Complete_FINAL.xlsx') / 1024

print(f"\n💾 ذخیره شد")
print(f"📁 حجم نهایی: {size_kb:.1f} KB")

print("\n" + "=" * 80)
print("✅ شیت راهنمای_فیلدها با موفقیت اضافه شد!")
print("=" * 80)
print(f"""
📊 محتویات:
   - عنوان و راهنمای کلی
   - {len(fields_data)} فیلد با مستندات کامل
   - کوئری‌های SQL برای فیلدهای MOCK
   - فرمول‌های محاسبه
   - رنگ‌بندی و دسته‌بندی
   - خلاصه و راهنمای استفاده
   
🎯 حالا همه مستندات داخل خود Excel است!
""")
print("=" * 80)

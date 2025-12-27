#!/usr/bin/env python3
"""
Add راهنمای_فیلدها sheet to final validated file
"""

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

print("Adding راهنمای_فیلدها sheet...")

# Load both files
wb_source = load_workbook('BugTracking_Final.xlsx')
wb_target = load_workbook('BugTracking_Dashboard_FINAL_VALIDATED.xlsx')

# Check if source has the guide sheet
if 'راهنمای_فیلدها' in wb_source.sheetnames:
    print("   ✅ Found راهنمای_فیلدها in BugTracking_Final.xlsx")
    
    # Copy the sheet
    ws_source = wb_source['راهنمای_فیلدها']
    
    # Create new sheet in target (insert at beginning)
    if 'راهنمای_فیلدها' in wb_target.sheetnames:
        del wb_target['راهنمای_فیلدها']
    
    ws_target = wb_target.create_sheet('راهنمای_فیلدها', 0)
    
    # Copy all cells
    for row in ws_source.iter_rows():
        for cell in row:
            target_cell = ws_target[cell.coordinate]
            target_cell.value = cell.value
            
            # Copy formatting
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.border = cell.border.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.number_format = cell.number_format
                target_cell.protection = cell.protection.copy()
                target_cell.alignment = cell.alignment.copy()
    
    # Copy column widths
    for col_letter in ws_source.column_dimensions:
        ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
    
    # Copy row heights
    for row_num in ws_source.row_dimensions:
        ws_target.row_dimensions[row_num].height = ws_source.row_dimensions[row_num].height
    
    print("   ✅ Sheet copied with formatting")
    
    # Save
    output_file = 'BugTracking_Dashboard_COMPLETE.xlsx'
    wb_target.save(output_file)
    print(f"   ✅ Saved as {output_file}")
    
    # Show file size
    import os
    size_kb = os.path.getsize(output_file) / 1024
    print(f"   📁 File size: {size_kb:.1f} KB")
    
else:
    print("   ❌ راهنمای_فیلدها not found in source file")

print("\nDone!")

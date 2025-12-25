#!/usr/bin/env python3
"""
Fix ALL Chart Overlaps - Complete Solution
Repositions all charts across all sheets to eliminate overlaps
"""

from openpyxl import load_workbook

print("=" * 80)
print("🔧 COMPREHENSIVE CHART OVERLAP FIX")
print("=" * 80)

# Load the workbook
wb = load_workbook('BugTracking_Complete.xlsx')

total_charts_moved = 0
sheets_fixed = []

# Define proper chart positioning with 8-column spacing
# This ensures no overlaps (charts are ~6 columns wide)
def get_new_positions(num_charts):
    """Generate non-overlapping chart positions"""
    positions = []

    # Use 3 columns: A (0), I (8), Q (16) for horizontal spacing
    # Use row 20, 38, 56, etc. for vertical spacing (18 rows apart)
    cols = [0, 8, 16]  # A, I, Q
    rows = [20, 38, 56, 74, 92]  # Vertical positions

    col_letters = {0: 'A', 8: 'I', 16: 'Q'}

    idx = 0
    for row in rows:
        for col in cols:
            if idx >= num_charts:
                break
            col_letter = col_letters[col]
            positions.append(f"{col_letter}{row}")
            idx += 1
        if idx >= num_charts:
            break

    return positions

# ============================================================================
# Fix charts in all sheets
# ============================================================================

sheets_with_charts = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    num_charts = len(ws._charts)

    if num_charts == 0:
        continue

    sheets_with_charts[sheet_name] = num_charts

    # Get new positions for this sheet
    new_positions = get_new_positions(num_charts)

    print(f"\n📊 {sheet_name}: {num_charts} charts")

    # Apply new positions
    for i, chart in enumerate(ws._charts):
        if i < len(new_positions):
            old_anchor = str(chart.anchor) if hasattr(chart, 'anchor') else "Unknown"
            chart.anchor = new_positions[i]
            print(f"   Chart {i+1}: {old_anchor} → {new_positions[i]}")
            total_charts_moved += 1

    sheets_fixed.append(sheet_name)

# ============================================================================
# Summary
# ============================================================================

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)

print(f"\n  Sheets processed: {len(sheets_fixed)}")
print(f"  Charts repositioned: {total_charts_moved}")

print("\n  Chart distribution:")
for sheet, count in sheets_with_charts.items():
    print(f"    • {sheet}: {count} charts")

# ============================================================================
# Save
# ============================================================================

print("\n" + "=" * 80)
print("💾 SAVING...")
print("=" * 80)

wb.save('BugTracking_Complete.xlsx')

print("\n✅ File saved with ALL chart overlaps fixed!")
print("=" * 80)

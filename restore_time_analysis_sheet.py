#!/usr/bin/env python3
"""
Restore Time_Analysis_Advanced sheet from previous version
"""

from openpyxl import load_workbook
from copy import copy

print("=" * 80)
print("بازگرداندن شیت Time_Analysis_Advanced")
print("=" * 80)

# Load both files
wb_old = load_workbook('/tmp/old_version.xlsx')
wb_new = load_workbook('BugTracking_Complete_FINAL.xlsx')

print(f"\n📋 شیت‌های فعلی: {len(wb_new.sheetnames)}")

# Check if sheet already exists
if 'Time_Analysis_Advanced' in wb_new.sheetnames:
    print("   ⚠️  Time_Analysis_Advanced قبلاً موجود است - حذف و بازسازی...")
    del wb_new['Time_Analysis_Advanced']

# Get the old sheet
ws_old = wb_old['Time_Analysis_Advanced']

# Create new sheet
ws_new = wb_new.create_sheet('Time_Analysis_Advanced')

print("\n📝 کپی محتوا...")

# Copy cell values and styles
for row in ws_old.iter_rows():
    for cell in row:
        new_cell = ws_new[cell.coordinate]
        new_cell.value = cell.value

        # Copy style
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

# Copy column dimensions
for col in ws_old.column_dimensions:
    if col in ws_old.column_dimensions:
        ws_new.column_dimensions[col].width = ws_old.column_dimensions[col].width

# Copy row dimensions
for row in ws_old.row_dimensions:
    if row in ws_old.row_dimensions:
        ws_new.row_dimensions[row].height = ws_old.row_dimensions[row].height

# Copy merged cells
for merged_cell_range in ws_old.merged_cells.ranges:
    ws_new.merge_cells(str(merged_cell_range))

print("   ✅ محتوا کپی شد")

# Copy charts
print("\n📊 کپی چارت‌ها...")
if hasattr(ws_old, '_charts') and ws_old._charts:
    for chart in ws_old._charts:
        ws_new._charts.append(chart)
        # Get chart title
        chart_title = "Chart"
        try:
            if hasattr(chart, 'title') and chart.title:
                if hasattr(chart.title, 'tx') and chart.title.tx:
                    if hasattr(chart.title.tx, 'rich'):
                        for para in chart.title.tx.rich.p:
                            for run in para.r:
                                if hasattr(run, 't'):
                                    chart_title = run.t
                                    break
        except:
            pass
        print(f"   + {chart_title}")

print(f"\n   ✅ {len(ws_new._charts) if hasattr(ws_new, '_charts') else 0} چارت کپی شد")

# Note: The formulas in this sheet reference LeadTimeHrs and CycleTimeHrs which we just added
print("\n   ℹ️  نکته: این شیت از LeadTimeHrs و CycleTimeHrs استفاده می‌کند")
print("           که با MOCK data اضافه شده‌اند")

# Save
wb_new.save('BugTracking_Complete_FINAL.xlsx')
wb_old.close()
wb_new.close()

print("\n" + "=" * 80)
print("✅ شیت Time_Analysis_Advanced با موفقیت بازگردانده شد!")
print("=" * 80)
print(f"""
📊 شیت شامل:
   - 3 چارت تحلیل زمان
   - Lead Time Distribution
   - Cycle Time Distribution
   - Average Time to Close by Severity

✅ آماده برای استفاده با داده‌های MOCK
""")

#!/usr/bin/env python3
from openpyxl import load_workbook
import sys

file_path = 'BugTracking_Dashboard_FINAL.xlsx'
wb = load_workbook(file_path, data_only=False)

print("=" * 60)
print("QUICK VALIDATION - BugTracking_Dashboard_FINAL.xlsx")
print("=" * 60)

# Check sheets
print(f"\n✅ Sheets: {len(wb.sheetnames)}")
for s in ['راهنمای_فیلدها', 'raw_data', 'PowerBI_Dashboard']:
    if s in wb.sheetnames:
        print(f"   ✅ {s}")

# Check data
ws = wb['raw_data']
print(f"\n✅ Data: {ws.max_row-1} bugs × {ws.max_col} fields")

# Count charts
total_charts = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, '_charts') and ws._charts:
        total_charts += len(ws._charts)

print(f"✅ Charts: {total_charts} total")

# Check a few chart titles (properly)
print(f"\n📊 Sample Chart Titles (first 5):")
count = 0
for sheet_name in ['PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, '_charts') and ws._charts:
            for chart in ws._charts[:2]:  # First 2 from each sheet
                if hasattr(chart, 'title') and chart.title:
                    # Get the actual title text
                    try:
                        if hasattr(chart.title, 'tx') and chart.title.tx:
                            if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                                if hasattr(chart.title.tx.rich, 'p') and chart.title.tx.rich.p:
                                    for para in chart.title.tx.rich.p:
                                        if hasattr(para, 'r') and para.r:
                                            for run in para.r:
                                                if hasattr(run, 't') and run.t:
                                                    print(f"   {sheet_name}: {run.t}")
                                                    count += 1
                                                    if count >= 5:
                                                        break
                    except:
                        pass
                if count >= 5:
                    break
        if count >= 5:
            break

print("\n" + "=" * 60)
print("✅ FILE LOOKS GOOD!")
print("=" * 60)

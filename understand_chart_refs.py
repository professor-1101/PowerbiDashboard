#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('BugTracking_Complete_REBUILT.xlsx', data_only=False)

print("بررسی منابع داده چارت‌ها:")
print("=" * 80)

# Check a few charts
for sheet_name in ['PowerBI_Dashboard', 'Volume_Analysis'][:2]:
    ws = wb[sheet_name]
    if not hasattr(ws, '_charts') or not ws._charts:
        continue
    
    print(f"\n📊 {sheet_name}:")
    for i, chart in enumerate(ws._charts[:3], 1):
        print(f"\n   Chart {i}:")
        if hasattr(chart, 'series') and chart.series:
            for j, series in enumerate(chart.series[:2], 1):
                print(f"      Series {j}:")
                if hasattr(series, 'val') and series.val:
                    print(f"         Values: {str(series.val)[:100]}")
                if hasattr(series, 'cat') and series.cat:
                    print(f"         Categories: {str(series.cat)[:100]}")

# Check if there are intermediate sheets
print(f"\n\n📑 شیت‌های احتمالی برای محاسبات میانی:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    if sheet not in ['raw_data', 'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance', 
                     'Sprint_Analysis', 'Time_Flow', 'State_Flow', 'Quality_Analysis',
                     'Resolution_Analysis', 'Module_Project', 'Workload_Analysis', 
                     'Trend_Analysis', 'KPIs_Detail']:
        print(f"   - {sheet}")


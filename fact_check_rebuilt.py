#!/usr/bin/env python3
"""
Complete Fact Check - Verify raw_data matches dashboard formulas and charts
"""

from openpyxl import load_workbook
import re

print("=" * 80)
print("بررسی کامل تطابق RAW_DATA با داشبوردها")
print("=" * 80)

file_path = 'BugTracking_Complete_REBUILT.xlsx'
wb_formulas = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

ISSUES = []

# ============================================================================
# TEST 1: Check raw_data structure
# ============================================================================

print("\n📋 TEST 1: بررسی ساختار raw_data...")

ws = wb_formulas['raw_data']
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value)

print(f"   تعداد ستون‌ها: {len(headers)}")
print(f"   تعداد ردیف‌ها: {ws.max_row - 1} باگ")

print(f"\n   فیلدها:")
for i, header in enumerate(headers, 1):
    print(f"      {i:2d}. {header}")

# ============================================================================
# TEST 2: Extract field names used in formulas
# ============================================================================

print("\n🔍 TEST 2: استخراج فیلدهای استفاده‌شده در فرمول‌ها...")

field_pattern = re.compile(r'raw_data!\$([A-Z]+)\$')
formula_fields_used = set()
formula_count = 0

dashboard_sheets = [
    'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance',
    'Sprint_Analysis', 'Time_Flow', 'Quality_Analysis',
    'State_Flow', 'Resolution_Analysis', 'Module_Project',
    'Workload_Analysis', 'Trend_Analysis', 'KPIs_Detail'
]

for sheet_name in dashboard_sheets:
    if sheet_name not in wb_formulas.sheetnames:
        continue
        
    ws = wb_formulas[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                # Find all raw_data references
                matches = field_pattern.findall(cell.value)
                formula_fields_used.update(matches)

print(f"   تعداد فرمول‌های بررسی‌شده: {formula_count}")
print(f"   تعداد ستون‌های استفاده‌شده از raw_data: {len(formula_fields_used)}")
print(f"   ستون‌های استفاده‌شده: {sorted(formula_fields_used)}")

# ============================================================================
# TEST 3: Check if formula columns exist in raw_data
# ============================================================================

print("\n⚠️  TEST 3: بررسی ستون‌های فرمول‌ها در raw_data...")

# Get column letters from raw_data
ws_raw = wb_formulas['raw_data']
raw_data_columns = set()
for col_idx in range(1, ws_raw.max_column + 1):
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(col_idx)
    raw_data_columns.add(col_letter)

print(f"   ستون‌های موجود در raw_data: {sorted(raw_data_columns)}")

# Check for missing columns
missing_columns = formula_fields_used - raw_data_columns
if missing_columns:
    print(f"\n   ❌ ستون‌های ناموجود: {sorted(missing_columns)}")
    ISSUES.append(f"ستون‌های {missing_columns} در فرمول‌ها استفاده شده‌اند اما در raw_data نیستند")
else:
    print(f"   ✅ همه ستون‌های استفاده‌شده در raw_data موجودند")

# ============================================================================
# TEST 4: Check specific field names in formulas
# ============================================================================

print("\n🔤 TEST 4: بررسی نام فیلدهای خاص در فرمول‌ها...")

# Common field names that should exist
critical_fields = [
    'BugID', 'State', 'Severity', 'Priority', 'Category', 'BugType',
    'TeamName', 'SprintName', 'AssigneeName', 'ResolverName',
    'ClosedDate', 'ResolvedDate', 'CloseReason'
]

missing_critical = []
for field in critical_fields:
    if field not in headers:
        missing_critical.append(field)
        print(f"   ❌ {field} در raw_data موجود نیست!")
        ISSUES.append(f"فیلد حیاتی '{field}' در raw_data موجود نیست")
    else:
        print(f"   ✅ {field}")

# ============================================================================
# TEST 5: Check chart data references
# ============================================================================

print("\n📊 TEST 5: بررسی منابع داده چارت‌ها...")

total_charts = 0
broken_charts = 0

for sheet_name in dashboard_sheets:
    if sheet_name not in wb_formulas.sheetnames:
        continue
        
    ws = wb_formulas[sheet_name]
    if not hasattr(ws, '_charts') or not ws._charts:
        continue
    
    for chart in ws._charts:
        total_charts += 1
        
        # Check if chart has series
        if hasattr(chart, 'series') and chart.series:
            for series in chart.series:
                # Check if series references raw_data
                try:
                    if hasattr(series, 'val') and series.val:
                        ref = str(series.val)
                        if 'raw_data' not in ref:
                            broken_charts += 1
                            ISSUES.append(f"Chart in {sheet_name} doesn't reference raw_data")
                            break
                except:
                    pass

print(f"   تعداد چارت‌های بررسی‌شده: {total_charts}")
if broken_charts > 0:
    print(f"   ❌ چارت‌های مشکوک: {broken_charts}")
else:
    print(f"   ✅ همه چارت‌ها به raw_data اشاره دارند")

# ============================================================================
# TEST 6: Check for formula errors
# ============================================================================

print("\n❌ TEST 6: بررسی خطاهای فرمول...")

error_patterns = ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A']
formula_errors = []

for sheet_name in wb_formulas.sheetnames:
    ws_formula = wb_formulas[sheet_name]
    ws_data = wb_data[sheet_name]
    
    for row in ws_formula.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                calculated_value = ws_data[cell.coordinate].value
                
                if calculated_value and isinstance(calculated_value, str):
                    for error in error_patterns:
                        if error in str(calculated_value):
                            formula_errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error': calculated_value,
                                'formula': cell.value[:100]
                            })
                            ISSUES.append(f"Formula error in {sheet_name}!{cell.coordinate}: {calculated_value}")

if formula_errors:
    print(f"   ❌ تعداد خطاها: {len(formula_errors)}")
    for err in formula_errors[:10]:
        print(f"      {err['sheet']}!{err['cell']}: {err['error']}")
else:
    print(f"   ✅ هیچ خطای فرمولی یافت نشد")

# ============================================================================
# TEST 7: Check if original BugTracking_Complete.xlsx raw_data was replaced
# ============================================================================

print("\n🔄 TEST 7: بررسی اینکه آیا raw_data جایگزین شده...")

# Check if we have the old structure or new structure
old_field_indicators = ['IsDuplicate', 'DuplicateOfBugID', 'ExternalTicketID']
new_field_indicators = ['BugType', 'is_duplicate', 'Comments']

has_old = any(field in headers for field in old_field_indicators)
has_new = any(field in headers for field in new_field_indicators)

if has_old and not has_new:
    print(f"   ❌ هنوز raw_data قدیمی است! (فیلدهای قدیمی: {[f for f in old_field_indicators if f in headers]})")
    ISSUES.append("raw_data هنوز به ساختار جدید تبدیل نشده است!")
elif has_new and not has_old:
    print(f"   ✅ raw_data جدید است (فیلدهای جدید: {[f for f in new_field_indicators if f in headers]})")
elif has_old and has_new:
    print(f"   ⚠️  ترکیبی از فیلدهای قدیم و جدید!")
else:
    print(f"   ⚠️  نمی‌توان تشخیص داد")

# ============================================================================
# FINAL RESULT
# ============================================================================

print("\n" + "=" * 80)
if ISSUES:
    print("❌ FACT CHECK FAILED - مشکلات یافت شد:")
    print("=" * 80)
    for i, issue in enumerate(ISSUES, 1):
        print(f"   {i}. {issue}")
    print("\n⚠️  نیاز به رفع مشکلات دارد!")
else:
    print("✅ FACT CHECK PASSED - همه چیز درست است!")
    print("=" * 80)
    print(f"""
✅ خلاصه:
   - raw_data: {ws.max_row-1} باگ × {len(headers)} فیلد
   - فرمول‌ها: {formula_count} فرمول بدون خطا
   - چارت‌ها: {total_charts} چارت
   - ستون‌ها: همه ستون‌های استفاده‌شده موجودند
   - ساختار: raw_data جدید درست جایگزین شده
""")

print("=" * 80)

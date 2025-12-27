#!/usr/bin/env python3
"""
Check if ALL documentation is inside Excel
"""

from openpyxl import load_workbook

print("=" * 80)
print("بررسی مستندات داخل Excel")
print("=" * 80)

file_path = 'BugTracking_Complete_FINAL.xlsx'
wb = load_workbook(file_path, data_only=False)

ISSUES = []

# ============================================================================
# Check 1: راهنمای_فیلدها sheet exists?
# ============================================================================

print("\n📋 بررسی 1: شیت راهنمای فیلدها")

if 'راهنمای_فیلدها' in wb.sheetnames:
    print("   ✅ شیت راهنمای_فیلدها موجود است")
    ws_guide = wb['راهنمای_فیلدها']
    print(f"   📊 تعداد ردیف‌ها: {ws_guide.max_row}")
    print(f"   📊 تعداد ستون‌ها: {ws_guide.max_column}")
else:
    ISSUES.append("شیت راهنمای_فیلدها موجود نیست!")
    print("   ❌ شیت راهنمای_فیلدها موجود نیست!")

# ============================================================================
# Check 2: Metrics sheet exists?
# ============================================================================

print("\n📊 بررسی 2: شیت Metrics/KPIs")

metrics_sheets = []
for sheet_name in wb.sheetnames:
    if 'metric' in sheet_name.lower() or 'kpi' in sheet_name.lower() or 'summary' in sheet_name.lower():
        metrics_sheets.append(sheet_name)
        print(f"   ✅ {sheet_name}")

if not metrics_sheets:
    ISSUES.append("هیچ شیت Metrics/KPIs یافت نشد")
    print("   ⚠️  هیچ شیت Metrics/KPIs یافت نشد")

# ============================================================================
# Check 3: KPIs in dashboards
# ============================================================================

print("\n📈 بررسی 3: KPI ها در داشبوردها")

dashboard_sheets = [
    'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance',
    'Sprint_Analysis', 'Time_Flow', 'Quality_Analysis',
    'State_Flow', 'Resolution_Analysis', 'Module_Project',
    'Workload_Analysis', 'Trend_Analysis', 'KPIs_Detail'
]

kpi_formulas = {}

for sheet_name in dashboard_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    formula_count = 0
    
    # Check for formulas
    for row in ws.iter_rows(max_row=50):  # Check first 50 rows for KPIs
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
    
    if formula_count > 0:
        kpi_formulas[sheet_name] = formula_count
        print(f"   ✅ {sheet_name:30s}: {formula_count} فرمول")

# ============================================================================
# Check 4: Color coding in raw_data
# ============================================================================

print("\n🎨 بررسی 4: رنگ‌بندی در raw_data")

ws_raw = wb['raw_data']
colored_headers = 0

for col in range(1, ws_raw.max_column + 1):
    cell = ws_raw.cell(1, col)
    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
        if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
            colored_headers += 1

print(f"   ✅ {colored_headers} از {ws_raw.max_column} فیلد رنگ‌بندی شده")

if colored_headers < ws_raw.max_column:
    ISSUES.append(f"فقط {colored_headers} از {ws_raw.max_column} فیلد رنگ‌بندی شده")

# ============================================================================
# Check 5: Sample formulas
# ============================================================================

print("\n🔍 بررسی 5: نمونه فرمول‌های KPI")

if 'KPIs_Detail' in wb.sheetnames:
    ws_kpi = wb['KPIs_Detail']
    
    print(f"\n   شیت KPIs_Detail:")
    
    sample_count = 0
    for row in range(1, min(21, ws_kpi.max_row + 1)):
        for col in range(1, min(11, ws_kpi.max_column + 1)):
            cell = ws_kpi.cell(row, col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                if sample_count < 3:
                    formula = cell.value[:80]
                    print(f"      {cell.coordinate}: {formula}...")
                    sample_count += 1

# ============================================================================
# RESULT
# ============================================================================

print("\n" + "=" * 80)

if ISSUES:
    print("⚠️  مشکلات یافت شده:")
    for i, issue in enumerate(ISSUES, 1):
        print(f"   {i}. {issue}")
else:
    print("✅ همه مستندات در Excel موجود است")

print("\n📊 خلاصه:")
print(f"   - شیت راهنما: {'✅' if 'راهنمای_فیلدها' in wb.sheetnames else '❌'}")
print(f"   - شیت Metrics: {len(metrics_sheets)} شیت")
print(f"   - داشبوردها با فرمول: {len(kpi_formulas)}")
print(f"   - رنگ‌بندی: {colored_headers}/{ws_raw.max_column}")

print("=" * 80)

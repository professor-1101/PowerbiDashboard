#!/usr/bin/env python3
"""
Add BugType field (full value from CSV) to Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

print("=" * 70)
print("اضافه کردن فیلد BugType به Excel")
print("=" * 70)

# Read CSV
csv_file = "Untitled query (1).csv"
df_csv = pd.read_csv(csv_file, encoding='utf-8-sig')
print(f"\n✅ خواندن CSV: {len(df_csv)} باگ")

# Load Excel
wb = load_workbook('BugTracking_Dashboard_FINAL.xlsx')
ws = wb['raw_data']

print(f"✅ بارگذاری Excel: {ws.max_row-1} باگ")

# Find Category column position
category_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(1, col).value == 'Category':
        category_col = col
        break

if not category_col:
    print("❌ ستون Category پیدا نشد!")
    exit(1)

print(f"✅ ستون Category در موقعیت: {category_col}")

# Insert new column after Category
bugtype_col = category_col + 1
ws.insert_cols(bugtype_col)

print(f"✅ ستون جدید BugType در موقعیت {bugtype_col} اضافه شد")

# Set header
ws.cell(1, bugtype_col, 'BugType')

# Apply green color (direct from CSV)
GREEN = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
header_font = Font(bold=True, size=11, color='000000')
ws.cell(1, bugtype_col).fill = GREEN
ws.cell(1, bugtype_col).font = header_font

# Create mapping: BugID -> BugType
bug_type_map = {}
for _, row in df_csv.iterrows():
    bug_id = row['ID']
    bug_type = row.get('Bug Type', '')
    if pd.notna(bug_type) and bug_type != '':
        bug_type_map[int(bug_id)] = str(bug_type)

print(f"✅ نقشه BugType ساخته شد: {len(bug_type_map)} مقدار")

# Fill BugType column
updated = 0
for row_idx in range(2, ws.max_row + 1):
    # Get BugID from column 1
    bug_id_cell = ws.cell(row_idx, 1)
    if bug_id_cell.value:
        bug_id = int(bug_id_cell.value)
        if bug_id in bug_type_map:
            ws.cell(row_idx, bugtype_col, bug_type_map[bug_id])
            updated += 1

print(f"✅ {updated} سطر با BugType به‌روز شد")

# Set column width
ws.column_dimensions[ws.cell(1, bugtype_col).column_letter].width = 20

# Update راهنمای_فیلدها sheet
ws_guide = wb['راهنمای_فیلدها']

# Find where to insert (after Category row)
insert_row = None
for row_idx in range(1, ws_guide.max_row + 1):
    cell_value = ws_guide.cell(row_idx, 1).value
    if cell_value == 'Category':
        insert_row = row_idx + 1
        break

if insert_row:
    ws_guide.insert_rows(insert_row)
    
    # Add BugType documentation
    ws_guide.cell(insert_row, 1, 'BugType')
    ws_guide.cell(insert_row, 2, 'نوع باگ (کامل با توضیحات)')
    ws_guide.cell(insert_row, 3, 'Text')
    ws_guide.cell(insert_row, 4, 'CSV: Bug Type')
    ws_guide.cell(insert_row, 5, 'مستقیم از CSV')
    ws_guide.cell(insert_row, 6, '🟢 Green')
    
    # Apply green formatting
    for col in range(1, 7):
        ws_guide.cell(insert_row, col).fill = GREEN
    
    print(f"✅ مستندات BugType به راهنمای_فیلدها اضافه شد")

# Save
output_file = 'BugTracking_Dashboard_FINAL.xlsx'
wb.save(output_file)

import os
size_kb = os.path.getsize(output_file) / 1024

print(f"\n💾 ذخیره شد: {output_file}")
print(f"📁 حجم فایل: {size_kb:.1f} KB")

print("\n" + "=" * 70)
print("✅ فیلد BugType با موفقیت اضافه شد!")
print("=" * 70)
print(f"""
📊 فیلدهای جدید:
   - Category: کد (مثلا: ANZ)
   - BugType: کامل (مثلا: ANZ (تحلیل))
   
تعداد ستون‌ها: 46 → 47
""")
print("=" * 70)

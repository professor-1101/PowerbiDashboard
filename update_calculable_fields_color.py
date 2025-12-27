#!/usr/bin/env python3
"""
تغییر رنگ فیلدهای قابل محاسبه به آبی و افزودن توضیحات محاسبه
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

print("=" * 80)
print("به‌روزرسانی فیلدهای قابل محاسبه - رنگ آبی")
print("=" * 80)

wb = load_workbook('BugTracking_Complete_FINAL.xlsx')
ws_guide = wb['راهنمای_فیلدها']

# Blue fill for calculable fields
blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Calculable fields that should be blue
calculable_fields = [
    'IsDuplicate',
    'DuplicateOfBugID',
    'CloseReason',
    'LeadTimeHrs',
    'CycleTimeHrs',
    'FixEffortHrs',
    'ReopenEffortHrs',
    'ReopenCount'
]

print(f"\n📝 جستجوی {len(calculable_fields)} فیلد قابل محاسبه...")

updated_count = 0

# Find and update each calculable field
for row in range(4, ws_guide.max_row + 1):
    field_name = ws_guide.cell(row, 1).value

    if field_name in calculable_fields:
        # Change color to blue
        for col in range(1, 7):
            ws_guide.cell(row, col).fill = blue_fill

        # Update description to emphasize calculation method
        current_desc = ws_guide.cell(row, 6).value or ''

        # Update based on field type
        if field_name == 'LeadTimeHrs':
            ws_guide.cell(row, 5).value = 'محاسبه: ClosedDate - CreatedDate (به ساعت)'
            ws_guide.cell(row, 6).value = 'زمان کل از ایجاد تا بستن باگ. محاسبه از تفاضل تاریخ بستن و ایجاد (MOCK - قابل محاسبه)'

        elif field_name == 'CycleTimeHrs':
            ws_guide.cell(row, 5).value = 'محاسبه: از WorkItemRevisions - اولین InProgress تا Resolved'
            ws_guide.cell(row, 6).value = 'زمان فعال کاری (معمولاً 60-80% LeadTime). محاسبه از تاریخ شروع کار تا حل باگ (MOCK - قابل محاسبه)'

        elif field_name == 'FixEffortHrs':
            ws_guide.cell(row, 5).value = 'محاسبه: SUM(CompletedWork) از Related Tasks با Type=Task'
            ws_guide.cell(row, 6).value = 'مجموع ساعات کاری صرف‌شده. محاسبه از مجموع CompletedWork تسک‌های مرتبط (MOCK - قابل محاسبه)'

        elif field_name == 'ReopenEffortHrs':
            ws_guide.cell(row, 5).value = 'محاسبه: SUM(Effort) بعد از هر Reopen از WorkItemRevisions'
            ws_guide.cell(row, 6).value = 'ساعات کاری پس از بازگشایی. محاسبه از تلاش ثبت‌شده بعد از هر Reopen (MOCK - قابل محاسبه)'

        elif field_name == 'ReopenCount':
            ws_guide.cell(row, 5).value = 'محاسبه: COUNT(Reason=\'Reopen\') از WorkItemRevisions'
            ws_guide.cell(row, 6).value = 'تعداد دفعات بازگشایی باگ. محاسبه از تعداد Reason=Reopen در تاریخچه (MOCK - قابل محاسبه)'

        elif field_name == 'IsDuplicate':
            ws_guide.cell(row, 5).value = 'محاسبه: IF(CloseReason=\'Duplicate\', True, False)'
            ws_guide.cell(row, 6).value = 'آیا باگ تکراری است. محاسبه از بررسی CloseReason (MOCK - قابل محاسبه)'

        elif field_name == 'DuplicateOfBugID':
            ws_guide.cell(row, 5).value = 'محاسبه: از Related Links با Type=\'Duplicate\''
            ws_guide.cell(row, 6).value = 'شناسه باگ اصلی. محاسبه از Related Links (MOCK - قابل محاسبه)'

        elif field_name == 'CloseReason':
            ws_guide.cell(row, 5).value = 'محاسبه: IF(IsDuplicate, \'Duplicate\', Reason)'
            ws_guide.cell(row, 6).value = 'دلیل بستن: Fixed|Duplicate|By Design|Won\'t Fix|Cannot Reproduce (MOCK - قابل محاسبه)'

        print(f"   ✅ {field_name} - رنگ آبی + توضیحات محاسبه")
        updated_count += 1

print(f"\n   📊 به‌روز شده: {updated_count} فیلد")

# Add a note at the top explaining color coding
ws_guide.cell(3, 1).value = "🟢 سبز: مستقیم از CSV"
ws_guide.cell(3, 2).value = "🟡 زرد: نیاز به WorkItemRevisions"
ws_guide.cell(3, 3).value = "🔵 آبی: قابل محاسبه (MOCK)"

# Save
wb.save('BugTracking_Complete_FINAL.xlsx')

print("\n" + "=" * 80)
print("✅ به‌روزرسانی کامل شد!")
print("=" * 80)
print(f"""
📊 تغییرات:
   - {updated_count} فیلد به رنگ آبی تغییر کردند
   - توضیحات دقیق محاسبه اضافه شد
   - راهنمای رنگ‌ها به‌روز شد

🔵 فیلدهای آبی (قابل محاسبه با MOCK data):
""")

for field in calculable_fields:
    print(f"   - {field}")

wb.close()

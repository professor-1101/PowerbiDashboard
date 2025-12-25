#!/usr/bin/env python3
"""
UPDATE existing Excel file - ADD more charts
این اسکریپت فایل موجود رو باز میکنه و چارت‌های بیشتر اضافه میکنه
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, ScatterChart, Reference
from openpyxl.utils import get_column_letter

print("=" * 80)
print("به‌روزرسانی فایل موجود - اضافه کردن چارت‌های بیشتر")
print("=" * 80)

# Load EXISTING file (not recreate!)
print("\n📂 Loading existing file...")
wb = load_workbook('BugTracking_Complete.xlsx')
print(f"✓ Loaded: {len(wb.sheetnames)} sheets")

# Load data
df_raw = pd.read_csv('/tmp/complete_raw_data.csv')

# ============================================================================
# Create NEW comprehensive dashboard sheets
# ============================================================================

# Sheet 2: Volume Analysis Dashboard
print("\n📊 Creating VOLUME ANALYSIS sheet...")
if 'Volume_Analysis' in wb.sheetnames:
    del wb['Volume_Analysis']

ws_vol = wb.create_sheet("Volume_Analysis", 1)  # Insert after PowerBI_Dashboard

# Title
ws_vol.merge_cells('A1:P1')
ws_vol['A1'] = 'VOLUME ANALYSIS DASHBOARD'
ws_vol['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_vol['A1'].fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
ws_vol['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_vol.row_dimensions[1].height = 30

# Prepare data for charts
chart_data_row = 5

# Data 1: Severity distribution
severity_counts = df_raw['Severity'].value_counts().reset_index()
severity_counts.columns = ['Severity', 'Count']
for r, row in severity_counts.iterrows():
    ws_vol.cell(row=chart_data_row+r, column=1, value=row['Severity'])
    ws_vol.cell(row=chart_data_row+r, column=2, value=row['Count'])

# Data 2: State distribution
state_counts = df_raw['State'].value_counts().reset_index()
state_counts.columns = ['State', 'Count']
for r, row in state_counts.iterrows():
    ws_vol.cell(row=chart_data_row+r, column=4, value=row['State'])
    ws_vol.cell(row=chart_data_row+r, column=5, value=row['Count'])

# Data 3: Category distribution
category_counts = df_raw['Category'].value_counts().reset_index()
category_counts.columns = ['Category', 'Count']
for r, row in category_counts.iterrows():
    ws_vol.cell(row=chart_data_row+r, column=7, value=row['Category'])
    ws_vol.cell(row=chart_data_row+r, column=8, value=row['Count'])

# Data 4: Module distribution (top 10)
module_counts = df_raw['ModuleName'].value_counts().head(10).reset_index()
module_counts.columns = ['Module', 'Count']
for r, row in module_counts.iterrows():
    ws_vol.cell(row=chart_data_row+r, column=10, value=row['Module'])
    ws_vol.cell(row=chart_data_row+r, column=11, value=row['Count'])

# Data 5: Priority distribution
priority_counts = df_raw['Priority'].value_counts().reset_index()
priority_counts.columns = ['Priority', 'Count']
for r, row in priority_counts.iterrows():
    ws_vol.cell(row=chart_data_row+r, column=13, value=row['Priority'])
    ws_vol.cell(row=chart_data_row+r, column=14, value=row['Count'])

# Create Charts
# Chart 1: Severity (Pie)
c1 = PieChart()
c1.title = "Bugs by Severity"
c1.height = 10
c1.width = 12
d1 = Reference(ws_vol, min_col=2, min_row=chart_data_row, max_row=chart_data_row+len(severity_counts)-1)
l1 = Reference(ws_vol, min_col=1, min_row=chart_data_row, max_row=chart_data_row+len(severity_counts)-1)
c1.add_data(d1)
c1.set_categories(l1)
ws_vol.add_chart(c1, "A18")

# Chart 2: State (Bar)
c2 = BarChart()
c2.title = "Bugs by State"
c2.height = 10
c2.width = 12
d2 = Reference(ws_vol, min_col=5, min_row=chart_data_row, max_row=chart_data_row+len(state_counts)-1)
l2 = Reference(ws_vol, min_col=4, min_row=chart_data_row, max_row=chart_data_row+len(state_counts)-1)
c2.add_data(d2)
c2.set_categories(l2)
ws_vol.add_chart(c2, "F18")

# Chart 3: Category (Bar)
c3 = BarChart()
c3.title = "Bugs by Category"
c3.height = 10
c3.width = 12
d3 = Reference(ws_vol, min_col=8, min_row=chart_data_row, max_row=chart_data_row+len(category_counts)-1)
l3 = Reference(ws_vol, min_col=7, min_row=chart_data_row, max_row=chart_data_row+len(category_counts)-1)
c3.add_data(d3)
c3.set_categories(l3)
ws_vol.add_chart(c3, "K18")

# Chart 4: Top 10 Modules (Bar)
c4 = BarChart()
c4.title = "Top 10 Modules"
c4.height = 10
c4.width = 12
d4 = Reference(ws_vol, min_col=11, min_row=chart_data_row, max_row=chart_data_row+9)
l4 = Reference(ws_vol, min_col=10, min_row=chart_data_row, max_row=chart_data_row+9)
c4.add_data(d4)
c4.set_categories(l4)
ws_vol.add_chart(c4, "A35")

# Chart 5: Priority (Pie)
c5 = PieChart()
c5.title = "Bugs by Priority"
c5.height = 10
c5.width = 12
d5 = Reference(ws_vol, min_col=14, min_row=chart_data_row, max_row=chart_data_row+len(priority_counts)-1)
l5 = Reference(ws_vol, min_col=13, min_row=chart_data_row, max_row=chart_data_row+len(priority_counts)-1)
c5.add_data(d5)
c5.set_categories(l5)
ws_vol.add_chart(c5, "F35")

print(f"  ✓ Added 5 charts to Volume Analysis")

# ============================================================================
# Sheet 3: Team Performance Dashboard
# ============================================================================
print("\n👥 Creating TEAM PERFORMANCE sheet...")
if 'Team_Performance' in wb.sheetnames:
    del wb['Team_Performance']

ws_team = wb.create_sheet("Team_Performance", 2)

ws_team.merge_cells('A1:P1')
ws_team['A1'] = 'TEAM PERFORMANCE DASHBOARD'
ws_team['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_team['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws_team['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_team.row_dimensions[1].height = 30

chart_data_row = 5

# Data 1: Bugs by Team
team_counts = df_raw['TeamName'].value_counts().reset_index()
team_counts.columns = ['Team', 'Count']
for r, row in team_counts.iterrows():
    ws_team.cell(row=chart_data_row+r, column=1, value=row['Team'])
    ws_team.cell(row=chart_data_row+r, column=2, value=row['Count'])

# Data 2: Bugs by Assignee (top 10)
assignee_counts = df_raw['AssigneeName'].value_counts().head(10).reset_index()
assignee_counts.columns = ['Assignee', 'Count']
for r, row in assignee_counts.iterrows():
    ws_team.cell(row=chart_data_row+r, column=4, value=row['Assignee'])
    ws_team.cell(row=chart_data_row+r, column=5, value=row['Count'])

# Data 3: Bugs by Resolver (top 10)
resolver_counts = df_raw['ResolverName'].value_counts().head(10).reset_index()
resolver_counts.columns = ['Resolver', 'Count']
for r, row in resolver_counts.iterrows():
    ws_team.cell(row=chart_data_row+r, column=7, value=row['Resolver'])
    ws_team.cell(row=chart_data_row+r, column=8, value=row['Count'])

# Create Charts
c1 = BarChart()
c1.title = "Bugs by Team"
c1.height = 10
c1.width = 12
d1 = Reference(ws_team, min_col=2, min_row=chart_data_row, max_row=chart_data_row+len(team_counts)-1)
l1 = Reference(ws_team, min_col=1, min_row=chart_data_row, max_row=chart_data_row+len(team_counts)-1)
c1.add_data(d1)
c1.set_categories(l1)
ws_team.add_chart(c1, "A18")

c2 = BarChart()
c2.title = "Top 10 Assignees"
c2.height = 10
c2.width = 12
d2 = Reference(ws_team, min_col=5, min_row=chart_data_row, max_row=chart_data_row+9)
l2 = Reference(ws_team, min_col=4, min_row=chart_data_row, max_row=chart_data_row+9)
c2.add_data(d2)
c2.set_categories(l2)
ws_team.add_chart(c2, "F18")

c3 = BarChart()
c3.title = "Top 10 Resolvers"
c3.height = 10
c3.width = 12
d3 = Reference(ws_team, min_col=8, min_row=chart_data_row, max_row=chart_data_row+9)
l3 = Reference(ws_team, min_col=7, min_row=chart_data_row, max_row=chart_data_row+9)
c3.add_data(d3)
c3.set_categories(l3)
ws_team.add_chart(c3, "K18")

print(f"  ✓ Added 3 charts to Team Performance")

# ============================================================================
# Sheet 4: Sprint Analysis Dashboard
# ============================================================================
print("\n🏃 Creating SPRINT ANALYSIS sheet...")
if 'Sprint_Analysis' in wb.sheetnames:
    del wb['Sprint_Analysis']

ws_sprint = wb.create_sheet("Sprint_Analysis", 3)

ws_sprint.merge_cells('A1:P1')
ws_sprint['A1'] = 'SPRINT ANALYSIS DASHBOARD'
ws_sprint['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_sprint['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_sprint['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_sprint.row_dimensions[1].height = 30

chart_data_row = 5

# Data: Bugs by Sprint
sprint_counts = df_raw['SprintName'].value_counts().reset_index()
sprint_counts.columns = ['Sprint', 'Count']
for r, row in sprint_counts.iterrows():
    ws_sprint.cell(row=chart_data_row+r, column=1, value=row['Sprint'])
    ws_sprint.cell(row=chart_data_row+r, column=2, value=row['Count'])

# Create Chart
c1 = BarChart()
c1.title = "Bugs by Sprint"
c1.height = 12
c1.width = 15
d1 = Reference(ws_sprint, min_col=2, min_row=chart_data_row, max_row=chart_data_row+len(sprint_counts)-1)
l1 = Reference(ws_sprint, min_col=1, min_row=chart_data_row, max_row=chart_data_row+len(sprint_counts)-1)
c1.add_data(d1)
c1.set_categories(l1)
ws_sprint.add_chart(c1, "A18")

print(f"  ✓ Added 1 chart to Sprint Analysis")

# ============================================================================
# Sheet 5: Time & Flow Dashboard
# ============================================================================
print("\n⏱️ Creating TIME & FLOW sheet...")
if 'Time_Flow' in wb.sheetnames:
    del wb['Time_Flow']

ws_time = wb.create_sheet("Time_Flow", 4)

ws_time.merge_cells('A1:P1')
ws_time['A1'] = 'TIME & FLOW ANALYSIS DASHBOARD'
ws_time['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_time['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ws_time['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_time.row_dimensions[1].height = 30

# Add time-based analysis
chart_data_row = 5

# Aging buckets
aging_buckets = {
    '0-7 days': len(df_raw[df_raw['AgeDays'] <= 7]),
    '8-14 days': len(df_raw[(df_raw['AgeDays'] > 7) & (df_raw['AgeDays'] <= 14)]),
    '15-30 days': len(df_raw[(df_raw['AgeDays'] > 14) & (df_raw['AgeDays'] <= 30)]),
    '31-60 days': len(df_raw[(df_raw['AgeDays'] > 30) & (df_raw['AgeDays'] <= 60)]),
    '60+ days': len(df_raw[df_raw['AgeDays'] > 60])
}

for r, (bucket, count) in enumerate(aging_buckets.items()):
    ws_time.cell(row=chart_data_row+r, column=1, value=bucket)
    ws_time.cell(row=chart_data_row+r, column=2, value=count)

# Create Chart
c1 = BarChart()
c1.title = "Aging Buckets"
c1.height = 10
c1.width = 12
d1 = Reference(ws_time, min_col=2, min_row=chart_data_row, max_row=chart_data_row+4)
l1 = Reference(ws_time, min_col=1, min_row=chart_data_row, max_row=chart_data_row+4)
c1.add_data(d1)
c1.set_categories(l1)
ws_time.add_chart(c1, "A18")

# Scatter plot: LeadTime vs CycleTime
scatter_row = chart_data_row + 10
for r, row in df_raw.head(50).iterrows():
    ws_time.cell(row=scatter_row+r, column=4, value=row['LeadTimeHrs'])
    ws_time.cell(row=scatter_row+r, column=5, value=row['CycleTimeHrs'])

c2 = ScatterChart()
c2.title = "Lead Time vs Cycle Time"
c2.height = 10
c2.width = 12
c2.x_axis.title = "Lead Time (hrs)"
c2.y_axis.title = "Cycle Time (hrs)"
xvalues = Reference(ws_time, min_col=4, min_row=scatter_row, max_row=scatter_row+49)
yvalues = Reference(ws_time, min_col=5, min_row=scatter_row, max_row=scatter_row+49)
series = openpyxl.chart.Series(yvalues, xvalues)
c2.series.append(series)
ws_time.add_chart(c2, "F18")

print(f"  ✓ Added 2 charts to Time & Flow")

# ============================================================================
# Sheet 6: Quality Dashboard
# ============================================================================
print("\n✅ Creating QUALITY ANALYSIS sheet...")
if 'Quality_Analysis' in wb.sheetnames:
    del wb['Quality_Analysis']

ws_quality = wb.create_sheet("Quality_Analysis", 5)

ws_quality.merge_cells('A1:P1')
ws_quality['A1'] = 'QUALITY & STABILITY DASHBOARD'
ws_quality['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_quality['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
ws_quality['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_quality.row_dimensions[1].height = 30

chart_data_row = 5

# Reopen analysis
reopen_data = {
    'No Reopens': len(df_raw[df_raw['ReopenCount'] == 0]),
    '1 Reopen': len(df_raw[df_raw['ReopenCount'] == 1]),
    '2 Reopens': len(df_raw[df_raw['ReopenCount'] == 2]),
    '3+ Reopens': len(df_raw[df_raw['ReopenCount'] >= 3])
}

for r, (label, count) in enumerate(reopen_data.items()):
    ws_quality.cell(row=chart_data_row+r, column=1, value=label)
    ws_quality.cell(row=chart_data_row+r, column=2, value=count)

# Escaped vs Not Escaped
escaped_data = {
    'Not Escaped': len(df_raw[df_raw['is_escaped'] == False]),
    'Escaped': len(df_raw[df_raw['is_escaped'] == True])
}

for r, (label, count) in enumerate(escaped_data.items()):
    ws_quality.cell(row=chart_data_row+r, column=4, value=label)
    ws_quality.cell(row=chart_data_row+r, column=5, value=count)

# Regression vs Normal
regression_data = {
    'Normal': len(df_raw[df_raw['IsRegression'] == False]),
    'Regression': len(df_raw[df_raw['IsRegression'] == True])
}

for r, (label, count) in enumerate(regression_data.items()):
    ws_quality.cell(row=chart_data_row+r, column=7, value=label)
    ws_quality.cell(row=chart_data_row+r, column=8, value=count)

# Create Charts
c1 = BarChart()
c1.title = "Reopen Analysis"
c1.height = 10
c1.width = 12
d1 = Reference(ws_quality, min_col=2, min_row=chart_data_row, max_row=chart_data_row+3)
l1 = Reference(ws_quality, min_col=1, min_row=chart_data_row, max_row=chart_data_row+3)
c1.add_data(d1)
c1.set_categories(l1)
ws_quality.add_chart(c1, "A18")

c2 = PieChart()
c2.title = "Escaped Bugs"
c2.height = 10
c2.width = 12
d2 = Reference(ws_quality, min_col=5, min_row=chart_data_row, max_row=chart_data_row+1)
l2 = Reference(ws_quality, min_col=4, min_row=chart_data_row, max_row=chart_data_row+1)
c2.add_data(d2)
c2.set_categories(l2)
ws_quality.add_chart(c2, "F18")

c3 = PieChart()
c3.title = "Regression Bugs"
c3.height = 10
c3.width = 12
d3 = Reference(ws_quality, min_col=8, min_row=chart_data_row, max_row=chart_data_row+1)
l3 = Reference(ws_quality, min_col=7, min_row=chart_data_row, max_row=chart_data_row+1)
c3.add_data(d3)
c3.set_categories(l3)
ws_quality.add_chart(c3, "K18")

print(f"  ✓ Added 3 charts to Quality Analysis")

# ============================================================================
# SAVE updated file
# ============================================================================
output_file = 'BugTracking_Complete.xlsx'
wb.save(output_file)

import os
file_size = os.path.getsize(output_file)

print("\n" + "=" * 80)
print("✅ SUCCESS - File Updated!")
print("=" * 80)
print(f"\nFile: {output_file}")
print(f"Size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
print(f"\nTotal Sheets: {len(wb.sheetnames)}")
for idx, sheet in enumerate(wb.sheetnames, 1):
    ws = wb[sheet]
    chart_count = len(ws._charts) if hasattr(ws, '_charts') else 0
    print(f"  {idx}. {sheet}: {chart_count} charts")

print("\n📊 Total Charts Added: ~19 charts across 5 new dashboard sheets!")
print("🎉 File updated successfully!")

#!/usr/bin/env python3
"""
اصلاح MOCK data برای LeadTimeHrs, CycleTimeHrs, DuplicateOfBugID
"""

from openpyxl import load_workbook
import random

print("=" * 80)
print("اصلاح MOCK Data فیلدهای زمان")
print("=" * 80)

wb = load_workbook('BugTracking_Complete_FINAL.xlsx')
ws = wb['raw_data']

# Get headers
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

# Get column indices
lead_time_col = headers.index('LeadTimeHrs') + 1 if 'LeadTimeHrs' in headers else None
cycle_time_col = headers.index('CycleTimeHrs') + 1 if 'CycleTimeHrs' in headers else None
duplicate_of_col = headers.index('DuplicateOfBugID') + 1 if 'DuplicateOfBugID' in headers else None
is_duplicate_col = headers.index('IsDuplicate') + 1 if 'IsDuplicate' in headers else None
state_col = headers.index('State') + 1 if 'State' in headers else None
bugid_col = headers.index('BugID') + 1 if 'BugID' in headers else 1

print(f"\n📊 فیلدها:")
print(f"   LeadTimeHrs: Column {lead_time_col}")
print(f"   CycleTimeHrs: Column {cycle_time_col}")
print(f"   DuplicateOfBugID: Column {duplicate_of_col}")

# Collect all bug IDs for duplicate reference
all_bug_ids = []
for row in range(2, ws.max_row + 1):
    bug_id = ws.cell(row, bugid_col).value
    if bug_id:
        all_bug_ids.append(bug_id)

print(f"\n📝 تولید MOCK data برای {ws.max_row - 1} باگ...")

stats = {
    'lead_time_added': 0,
    'cycle_time_added': 0,
    'duplicate_fixed': 0
}

for row in range(2, ws.max_row + 1):
    state = ws.cell(row, state_col).value if state_col else None
    is_duplicate = ws.cell(row, is_duplicate_col).value if is_duplicate_col else None
    current_bug_id = ws.cell(row, bugid_col).value

    # Fix LeadTimeHrs - realistic MOCK data
    if lead_time_col:
        if state in ['Closed', 'Resolved']:
            # Realistic lead time distribution:
            # Small bugs: 4-48 hours (50%)
            # Medium bugs: 48-168 hours (1-7 days) (30%)
            # Large bugs: 168-720 hours (7-30 days) (20%)
            rand = random.random()
            if rand < 0.5:
                lead_time = random.uniform(4, 48)
            elif rand < 0.8:
                lead_time = random.uniform(48, 168)
            else:
                lead_time = random.uniform(168, 720)

            ws.cell(row, lead_time_col, round(lead_time, 2))
            stats['lead_time_added'] += 1
        else:
            # Open bugs - calculate from current time (partial lead time)
            partial_lead = random.uniform(1, 72)
            ws.cell(row, lead_time_col, round(partial_lead, 2))
            stats['lead_time_added'] += 1

    # Fix CycleTimeHrs - typically 60-80% of LeadTime
    if cycle_time_col and lead_time_col:
        lead_val = ws.cell(row, lead_time_col).value
        if lead_val and lead_val != 'N/A':
            cycle_time = lead_val * random.uniform(0.6, 0.8)
            ws.cell(row, cycle_time_col, round(cycle_time, 2))
            stats['cycle_time_added'] += 1
        else:
            ws.cell(row, cycle_time_col, 'N/A')

    # Fix DuplicateOfBugID - if IsDuplicate is Yes, assign a valid bug ID
    if duplicate_of_col and is_duplicate_col:
        if is_duplicate == 'Yes':
            # Pick a random bug that's not the current one
            available_bugs = [bid for bid in all_bug_ids if bid != current_bug_id]
            if available_bugs:
                original_bug = random.choice(available_bugs)
                ws.cell(row, duplicate_of_col, original_bug)
                stats['duplicate_fixed'] += 1
        elif ws.cell(row, duplicate_of_col).value == 'N/A':
            # Already correct
            pass

print(f"\n   ✅ MOCK data اضافه شد:")
print(f"      - LeadTimeHrs: {stats['lead_time_added']} باگ")
print(f"      - CycleTimeHrs: {stats['cycle_time_added']} باگ")
print(f"      - DuplicateOfBugID: {stats['duplicate_fixed']} باگ")

# Save
wb.save('BugTracking_Complete_FINAL.xlsx')

print("\n" + "=" * 80)
print("✅ اصلاح MOCK data کامل شد!")
print("=" * 80)
print(f"""
📊 نتیجه:
   - همه {ws.max_row - 1} باگ دارای LeadTimeHrs واقع‌گرایانه
   - همه باگ‌ها دارای CycleTimeHrs (60-80% از LeadTime)
   - باگ‌های Duplicate اصلاح شدند

✅ فایل آماده است!
""")

wb.close()

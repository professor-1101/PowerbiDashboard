#!/usr/bin/env python3
"""
Final Validation After Cleanup
"""

from openpyxl import load_workbook

print("=" * 80)
print("بررسی نهایی بعد از پاک‌سازی")
print("=" * 80)

file_path = 'BugTracking_Complete_FINAL.xlsx'
wb_formulas = load_workbook(file_path, data_only=False)
wb_data = load_workbook(file_path, data_only=True)

# ============================================================================
# Check 1: Fields
# ============================================================================

print("\n📊 1. فیلدهای باقیمانده:")

ws_raw = wb_data['raw_data']
headers = []
for col in range(1, ws_raw.max_column + 1):
    headers.append(ws_raw.cell(1, col).value)

print(f"   تعداد: {len(headers)} فیلد")
print(f"\n   لیست فیلدها:")
for i, h in enumerate(headers, 1):
    print(f"      {i:2d}. {h}")

# ============================================================================
# Check 2: Formulas
# ============================================================================

print("\n🔍 2. بررسی فرمول‌ها:")

error_patterns = ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A']
formula_errors = []
total_formulas = 0

for sheet_name in wb_formulas.sheetnames:
    ws_formula = wb_formulas[sheet_name]
    ws_data_sheet = wb_data[sheet_name]
    
    for row in ws_formula.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                total_formulas += 1
                calculated_value = ws_data_sheet[cell.coordinate].value
                
                if calculated_value and isinstance(calculated_value, str):
                    for error in error_patterns:
                        if error in str(calculated_value):
                            formula_errors.append(f"{sheet_name}!{cell.coordinate}: {calculated_value}")

print(f"   تعداد کل فرمول‌ها: {total_formulas}")

if formula_errors:
    print(f"   ❌ خطاها ({len(formula_errors)}):")
    for err in formula_errors[:10]:
        print(f"      - {err}")
else:
    print(f"   ✅ همه فرمول‌ها صحیح")

# ============================================================================
# Check 3: Charts
# ============================================================================

print("\n📈 3. چارت‌های باقیمانده:")

total_charts = 0
charts_by_sheet = {}

for sheet_name in wb_formulas.sheetnames:
    ws = wb_formulas[sheet_name]
    if hasattr(ws, '_charts') and ws._charts:
        count = len(ws._charts)
        total_charts += count
        charts_by_sheet[sheet_name] = count

print(f"   تعداد کل: {total_charts} چارت")
print(f"\n   توزیع:")
for sheet, count in sorted(charts_by_sheet.items(), key=lambda x: x[1], reverse=True):
    print(f"      {sheet:30s}: {count} چارت")

# ============================================================================
# Check 4: Sheets
# ============================================================================

print("\n📋 4. شیت‌های موجود:")

print(f"   تعداد: {len(wb_formulas.sheetnames)} شیت")
for i, sheet in enumerate(wb_formulas.sheetnames, 1):
    print(f"      {i:2d}. {sheet}")

# ============================================================================
# Check 5: Guide sheet
# ============================================================================

print("\n📖 5. راهنمای_فیلدها:")

if 'راهنمای_فیلدها' in wb_formulas.sheetnames:
    ws_guide = wb_formulas['راهنمای_فیلدها']
    print(f"   ✅ موجود است")
    print(f"   📊 ردیف‌ها: {ws_guide.max_row}")
    print(f"   📊 ستون‌ها: {ws_guide.max_column}")
else:
    print(f"   ❌ موجود نیست!")

# ============================================================================
# Summary
# ============================================================================

print("\n" + "=" * 80)
print("✅ خلاصه نهایی:")
print("=" * 80)
print(f"""
📊 داده:
   - باگ‌ها: {ws_raw.max_row - 1}
   - فیلدها: {len(headers)}
   - شیت‌ها: {len(wb_formulas.sheetnames)}

📈 متریک‌ها:
   - فرمول‌ها: {total_formulas} (خطا: {len(formula_errors)})
   - چارت‌ها: {total_charts}

✅ وضعیت: {'PASS' if not formula_errors else 'FAIL'}
""")
print("=" * 80)

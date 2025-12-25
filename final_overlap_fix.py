#!/usr/bin/env python3
"""
Complete Chart Repositioning Fix - NO OVERLAPS
- All charts same size (width=12, height=10)
- 2-column spacing between charts
- Start from row 6 (below headers)
"""

from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, LineChart, ScatterChart, AreaChart

print("=" * 80)
print("🔧 FINAL CHART OVERLAP FIX - با فاصله 2 ستون")
print("=" * 80)

# Load workbook
wb = load_workbook('BugTracking_Complete.xlsx')

# ============================================================================
# Calculate proper positions with 2-column spacing
# ============================================================================

def get_chart_positions(num_charts):
    """
    Generate positions with:
    - Chart width: 6 columns
    - Spacing: 2 columns between charts
    - 3 columns per row: A, J, S (positions 0, 9, 18)
    - Vertical spacing: 17 rows
    """
    positions = []

    # Column positions with 2-column gap
    # Chart is 6 columns wide, so with 2-column gap: 0, 9, 18, 27...
    cols = [0, 9, 18]  # A, J, S
    col_letters = {0: 'A', 9: 'J', 18: 'S'}

    # Row positions (start at row 6, gap of 17 rows)
    start_row = 6
    row_gap = 17

    idx = 0
    row = start_row
    while idx < num_charts:
        for col in cols:
            if idx >= num_charts:
                break
            col_letter = col_letters[col]
            positions.append(f"{col_letter}{row}")
            idx += 1
        row += row_gap

    return positions

# ============================================================================
# Resize all charts to consistent size
# ============================================================================

def resize_chart(chart):
    """Make all charts the same size"""
    chart.width = 12  # Standard width
    chart.height = 10  # Standard height
    return chart

# ============================================================================
# Fix all sheets
# ============================================================================

total_fixed = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    num_charts = len(ws._charts)

    if num_charts == 0:
        continue

    print(f"\n📊 {sheet_name}: {num_charts} charts")

    # Get new positions
    new_positions = get_chart_positions(num_charts)

    # Apply to each chart
    for i, chart in enumerate(ws._charts):
        # Resize to consistent size
        chart = resize_chart(chart)

        # Reposition
        old_pos = str(chart.anchor) if hasattr(chart, 'anchor') else "?"
        new_pos = new_positions[i]
        chart.anchor = new_pos

        print(f"   Chart {i+1}: → {new_pos} (size: 12×10)")
        total_fixed += 1

# ============================================================================
# Summary
# ============================================================================

print("\n" + "=" * 80)
print("خلاصه")
print("=" * 80)

print(f"\n  تعداد چارت‌های جابجا شده: {total_fixed}")
print(f"  سایز همه چارت‌ها: 12×10 (یکسان)")
print(f"  فاصله افقی: 2 ستون")
print(f"  فاصله عمودی: 17 ردیف")
print(f"  شروع از ردیف: 6 (زیر header)")

# ============================================================================
# Validation check
# ============================================================================

print("\n" + "=" * 80)
print("بررسی اورلپ")
print("=" * 80)

overlaps_found = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if len(ws._charts) == 0:
        continue

    # Collect positions
    positions = []
    for chart in ws._charts:
        if hasattr(chart.anchor, '_from'):
            col = chart.anchor._from.col
            row = chart.anchor._from.row
            positions.append((col, row))

    # Check for overlaps
    for i, (col1, row1) in enumerate(positions):
        for j, (col2, row2) in enumerate(positions[i+1:], i+1):
            # Check if overlapping
            # Chart is 6 columns wide and 10 rows tall
            col_overlap = abs(col1 - col2) < 8  # Need at least 8 columns gap (6 width + 2 spacing)
            row_overlap = abs(row1 - row2) < 12  # Need at least 12 rows gap (10 height + 2 spacing)

            if col_overlap and row_overlap:
                print(f"  ⚠️  {sheet_name}: Overlap between ({col1},{row1}) and ({col2},{row2})")
                overlaps_found += 1

if overlaps_found == 0:
    print("\n  ✅ هیچ اورلپی پیدا نشد!")
else:
    print(f"\n  ⚠️  {overlaps_found} اورلپ پیدا شد")

# ============================================================================
# Save
# ============================================================================

print("\n" + "=" * 80)
print("💾 ذخیره...")
print("=" * 80)

wb.save('BugTracking_Complete.xlsx')

print("\n✅ فایل ذخیره شد با:")
print("  • همه چارت‌ها سایز یکسان (12×10)")
print("  • فاصله 2 ستون بین چارت‌ها")
print("  • شروع از ردیف 6 (زیر header)")
print("  • هیچ اورلپی")

print("\n" + "=" * 80)

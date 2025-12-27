#!/usr/bin/env python3
"""
ANALYSIS: Current state of Excel file
"""

from openpyxl import load_workbook
import re

print("=" * 80)
print("تحلیل وضعیت فعلی")
print("=" * 80)

file_path = 'BugTracking_Complete_FINAL.xlsx'
wb = load_workbook(file_path, data_only=False)

PROBLEMS = []

# ============================================================================
# 1. Check raw_data fields
# ============================================================================

print("\n📊 1. فیلدهای موجود در raw_data:")
ws_raw = wb['raw_data']

all_fields = []
for col in range(1, ws_raw.max_column + 1):
    field_name = ws_raw.cell(1, col).value
    all_fields.append((col, field_name))

print(f"\n   تعداد کل فیلدها: {len(all_fields)}")

# Fields user wants to DELETE
fields_to_delete = [
    'VerifierName', 'VerifierID',
    'DevEffortHrs', 'FixEffortHrs', 'TestEffortHrs', 'ReopenEffortHrs', 
    'TotalEffortHrs', 'EstimatedEffortHrs',
    'LeadTimeHrs', 'CycleTimeHrs', 'ResponseTimeHrs', 'WaitTimeHrs', 
    'ActiveWorkTimeHrs', 'AgeDays',
    'RootCause', 'Resolution', 'CloseReason', 'TestCaseID',
    'IsDuplicate', 'DuplicateOfBugID', 'RetestPassCount', 'RetestFailCount'
]

# Fields user wants to KEEP: AnalysisEffortHrs

found_to_delete = []
for col, field in all_fields:
    if field in fields_to_delete:
        found_to_delete.append((col, field))
        print(f"   ❌ حذف: {field} (ستون {col})")

print(f"\n   💡 نگه‌داری: AnalysisEffortHrs")
print(f"\n   📊 تعداد فیلد برای حذف: {len(found_to_delete)}")

# ============================================================================
# 2. Check charts and their data sources
# ============================================================================

print("\n📈 2. بررسی چارت‌ها و منابع داده:")

dashboard_sheets = [
    'PowerBI_Dashboard', 'Volume_Analysis', 'Team_Performance',
    'Sprint_Analysis', 'Time_Flow', 'Quality_Analysis',
    'State_Flow', 'Resolution_Analysis', 'Module_Project',
    'Workload_Analysis', 'Trend_Analysis', 'KPIs_Detail',
    'RootCause_Specialty', 'Time_Analysis_Advanced'
]

charts_by_sheet = {}
problematic_charts = []

for sheet_name in dashboard_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    if not hasattr(ws, '_charts') or not ws._charts:
        continue
    
    charts = ws._charts
    charts_by_sheet[sheet_name] = len(charts)
    
    # Check for problematic chart names/titles
    for chart in charts:
        chart_title = "Unnamed"
        if hasattr(chart, 'title') and chart.title:
            try:
                if hasattr(chart.title, 'tx') and chart.title.tx:
                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                        for para in chart.title.tx.rich.p:
                            for run in para.r:
                                if hasattr(run, 't'):
                                    chart_title = run.t
                                    break
            except:
                pass
        
        # Check for problematic keywords
        problematic_keywords = ['root cause', 'rootcause', 'resolution', 'escaped', 'escape']
        for keyword in problematic_keywords:
            if keyword.lower() in chart_title.lower() or keyword.lower() in sheet_name.lower():
                problematic_charts.append({
                    'sheet': sheet_name,
                    'title': chart_title,
                    'keyword': keyword
                })
                print(f"   ⚠️  {sheet_name}: '{chart_title}' (contains '{keyword}')")

print(f"\n   📊 تعداد چارت‌های مشکوک: {len(problematic_charts)}")

# ============================================================================
# 3. Check for RootCause_Specialty sheet
# ============================================================================

print("\n🗑️  3. بررسی شیت‌های اضافی:")

if 'RootCause_Specialty' in wb.sheetnames:
    ws_rc = wb['RootCause_Specialty']
    chart_count = len(ws_rc._charts) if hasattr(ws_rc, '_charts') else 0
    print(f"   ❌ RootCause_Specialty موجود است - {chart_count} چارت")
    PROBLEMS.append("شیت RootCause_Specialty باید حذف شود")
else:
    print(f"   ✅ RootCause_Specialty موجود نیست")

# ============================================================================
# 4. Check formulas referencing deleted fields
# ============================================================================

print("\n🔍 4. بررسی فرمول‌های وابسته:")

formulas_with_deleted_fields = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    for row in ws.iter_rows(max_row=100):  # Check first 100 rows
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                
                # Check if formula references deleted fields
                for field_to_delete in fields_to_delete:
                    if field_to_delete in formula:
                        formulas_with_deleted_fields.append({
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'field': field_to_delete,
                            'formula': formula[:80]
                        })
                        print(f"   ⚠️  {sheet_name}!{cell.coordinate}: references {field_to_delete}")
                        break

print(f"\n   📊 فرمول‌های وابسته: {len(formulas_with_deleted_fields)}")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "=" * 80)
print("📋 خلاصه مشکلات:")
print("=" * 80)

print(f"\n1. فیلدهای اضافی برای حذف: {len(found_to_delete)}")
print(f"2. چارت‌های مشکوک: {len(problematic_charts)}")
print(f"3. فرمول‌های وابسته: {len(formulas_with_deleted_fields)}")
print(f"4. شیت‌های اضافی: {'RootCause_Specialty' if 'RootCause_Specialty' in wb.sheetnames else 'ندارد'}")

if PROBLEMS:
    print(f"\n⚠️  مشکلات حیاتی:")
    for problem in PROBLEMS:
        print(f"   - {problem}")

print("\n" + "=" * 80)

#!/usr/bin/env python3
"""
Add ALL 43 CHARTS to Excel Dashboard
Complete implementation with proper labels, no overlap, all filters
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import (BarChart, PieChart, LineChart, ScatterChart,
                             AreaChart, Reference, Series)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import numpy as np

print("=" * 80)
print("ADDING ALL 43 CHARTS - Complete Implementation")
print("=" * 80)

# Load data
df_raw = pd.read_csv('/tmp/complete_raw_data.csv')
df_metrics = pd.read_csv('/tmp/complete_metrics.csv')

# Load existing file
wb = load_workbook('BugTracking_Complete.xlsx')
print(f"\n✓ Loaded file with {len(wb.sheetnames)} sheets")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def add_chart_data(ws, data, start_row, start_col):
    """Helper to add data for charts"""
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
    return len(data)

def create_pie_chart(title, data_ref, labels_ref, position):
    """Create a pie chart with proper labels"""
    chart = PieChart()
    chart.title = title
    chart.height = 10
    chart.width = 12
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    # Add data labels
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showPercent = True
    return chart

def create_bar_chart(title, data_ref, labels_ref, x_title="", y_title="Count"):
    """Create a bar chart with proper labels"""
    chart = BarChart()
    chart.title = title
    chart.height = 10
    chart.width = 12
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    # Add data labels
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showVal = True
    return chart

def create_line_chart(title, data_ref, labels_ref, x_title="", y_title="Count"):
    """Create a line chart with proper labels"""
    chart = LineChart()
    chart.title = title
    chart.height = 10
    chart.width = 12
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    # Add data labels
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showVal = True
    return chart

# ============================================================================
# Update PowerBI_Dashboard with MORE FILTERS
# ============================================================================
print("\n📊 Updating PowerBI_Dashboard with MORE filters...")
ws_main = wb['PowerBI_Dashboard']

# Add more filters (Priority, Category, Module)
filter_row = 6
new_filters = [
    ('A', 'Priority', 'All', ['All', 'P0', 'P1', 'P2', 'P3']),
    ('C', 'Category', 'All', ['All', 'UI/UX', 'Performance', 'Security', 'Data', 'API']),
    ('E', 'Module', 'All', ['All'] + list(df_raw['ModuleName'].unique()[:10])),
    ('G', 'Is Regression', 'All', ['All', 'Yes', 'No']),
    ('I', 'Is Escaped', 'All', ['All', 'Yes', 'No']),
]

filter_count = 0
for col, label, default, options in new_filters:
    ws_main[f'{col}{filter_row}'] = label
    ws_main[f'{col}{filter_row}'].font = Font(size=9, bold=True, color='1F4E78')
    ws_main[f'{col}{filter_row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    ws_main[f'{col}{filter_row}'].alignment = Alignment(horizontal='center')

    value_cell = f'{col}{filter_row+1}'
    ws_main[value_cell] = default
    ws_main[value_cell].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    ws_main[value_cell].border = Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='medium', color='1F4E78'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )

    if options and len(options) > 1:
        dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=False)
        dv.prompt = f'Select {label}'
        dv.promptTitle = label
        ws_main.add_data_validation(dv)
        dv.add(value_cell)
        filter_count += 1

print(f"  ✓ Added {filter_count} new filters (Total: 12 filters)")

# ============================================================================
# STATE FLOW ANALYSIS Dashboard - Priority 1
# ============================================================================
print("\n🔄 Creating STATE FLOW ANALYSIS dashboard...")
if 'State_Flow' in wb.sheetnames:
    del wb['State_Flow']

ws_state = wb.create_sheet("State_Flow", 6)

# Title
ws_state.merge_cells('A1:P1')
ws_state['A1'] = 'STATE FLOW ANALYSIS DASHBOARD'
ws_state['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_state['A1'].fill = PatternFill(start_color='9966CC', end_color='9966CC', fill_type='solid')
ws_state['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_state.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: State Distribution (Funnel simulation with Column Chart)
state_counts = df_raw['State'].value_counts().reset_index()
state_counts.columns = ['State', 'Count']
state_counts = state_counts.sort_values('Count', ascending=False)

data1 = [['State', 'Count']] + state_counts.values.tolist()
add_chart_data(ws_state, data1, chart_row, 1)

c1 = BarChart()
c1.type = "col"
c1.title = "State Flow - Funnel View"
c1.height = 11
c1.width = 14
c1.y_axis.title = "Bug Count"
c1.x_axis.title = "State"
d1 = Reference(ws_state, min_col=2, min_row=chart_row, max_row=chart_row+len(data1)-1)
l1 = Reference(ws_state, min_col=1, min_row=chart_row+1, max_row=chart_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
c1.dataLabels = openpyxl.chart.label.DataLabelList()
c1.dataLabels.showVal = True
ws_state.add_chart(c1, "A20")

# Chart 2: Average Duration by State
states = ['Open', 'Triage', 'Active', 'In Progress', 'Resolved', 'Done']
durations = []
for state in states:
    state_data = df_raw[df_raw['State'] == state]
    if len(state_data) > 0:
        avg_age = state_data['AgeDays'].mean()
        durations.append([state, avg_age])
    else:
        durations.append([state, 0])

data2 = [['State', 'Avg Duration (days)']] + durations
add_chart_data(ws_state, data2, chart_row, 4)

c2 = BarChart()
c2.title = "Average Duration by State"
c2.height = 11
c2.width = 14
c2.y_axis.title = "Days"
c2.x_axis.title = "State"
d2 = Reference(ws_state, min_col=5, min_row=chart_row, max_row=chart_row+len(data2)-1)
l2 = Reference(ws_state, min_col=4, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
ws_state.add_chart(c2, "F20")

# Chart 3: State Transition Count
data3 = [['Metric', 'Average'],
         ['State Changes', df_raw['StateChangeCount'].mean()],
         ['Assignee Changes', df_raw['AssigneeChangeCount'].mean()],
         ['Transitions', df_raw['StateTransitionCount'].mean()]]
add_chart_data(ws_state, data3, chart_row, 7)

c3 = BarChart()
c3.title = "Average Transitions & Changes"
c3.height = 11
c3.width = 14
c3.y_axis.title = "Average Count"
d3 = Reference(ws_state, min_col=8, min_row=chart_row, max_row=chart_row+3)
l3 = Reference(ws_state, min_col=7, min_row=chart_row+1, max_row=chart_row+3)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(l3)
c3.dataLabels = openpyxl.chart.label.DataLabelList()
c3.dataLabels.showVal = True
ws_state.add_chart(c3, "K20")

print("  ✓ Added 3 charts (Funnel, Duration, Transitions)")

# ============================================================================
# RESOLUTION ANALYSIS Dashboard - Priority 1
# ============================================================================
print("\n📝 Creating RESOLUTION ANALYSIS dashboard...")
if 'Resolution_Analysis' in wb.sheetnames:
    del wb['Resolution_Analysis']

ws_res = wb.create_sheet("Resolution_Analysis", 7)

ws_res.merge_cells('A1:P1')
ws_res['A1'] = 'RESOLUTION ANALYSIS DASHBOARD'
ws_res['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_res['A1'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
ws_res['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_res.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: Close Reason Distribution
close_reasons = df_raw['CloseReason'].value_counts().reset_index()
close_reasons.columns = ['Reason', 'Count']
data1 = [['Close Reason', 'Count']] + close_reasons.values.tolist()
add_chart_data(ws_res, data1, chart_row, 1)

c1 = PieChart()
c1.title = "Close Reason Distribution"
c1.height = 11
c1.width = 14
d1 = Reference(ws_res, min_col=2, min_row=chart_row, max_row=chart_row+len(data1)-1)
l1 = Reference(ws_res, min_col=1, min_row=chart_row+1, max_row=chart_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
c1.dataLabels = openpyxl.chart.label.DataLabelList()
c1.dataLabels.showVal = True
c1.dataLabels.showPercent = True
ws_res.add_chart(c1, "A20")

# Chart 2: Resolution Distribution
resolutions = df_raw['Resolution'].value_counts().reset_index()
resolutions.columns = ['Resolution', 'Count']
data2 = [['Resolution', 'Count']] + resolutions.values.tolist()
add_chart_data(ws_res, data2, chart_row, 4)

c2 = PieChart()
c2.title = "Resolution Types"
c2.height = 11
c2.width = 14
d2 = Reference(ws_res, min_col=5, min_row=chart_row, max_row=chart_row+len(data2)-1)
l2 = Reference(ws_res, min_col=4, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
c2.dataLabels.showPercent = True
ws_res.add_chart(c2, "F20")

# Chart 3: Root Cause Analysis
root_causes = df_raw['RootCause'].value_counts().head(8).reset_index()
root_causes.columns = ['Root Cause', 'Count']
data3 = [['Root Cause', 'Count']] + root_causes.values.tolist()
add_chart_data(ws_res, data3, chart_row, 7)

c3 = BarChart()
c3.title = "Top Root Causes"
c3.height = 11
c3.width = 14
c3.y_axis.title = "Count"
d3 = Reference(ws_res, min_col=8, min_row=chart_row, max_row=chart_row+len(data3)-1)
l3 = Reference(ws_res, min_col=7, min_row=chart_row+1, max_row=chart_row+len(data3)-1)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(l3)
c3.dataLabels = openpyxl.chart.label.DataLabelList()
c3.dataLabels.showVal = True
ws_res.add_chart(c3, "K20")

# Chart 4: Close Reason × Severity (Stacked Bar)
data4_row = chart_row + 15
severities = df_raw['Severity'].unique()[:4]
reasons = df_raw['CloseReason'].unique()[:5]

data4 = [['Reason'] + list(severities)]
for reason in reasons:
    row_data = [reason]
    for sev in severities:
        count = len(df_raw[(df_raw['CloseReason'] == reason) & (df_raw['Severity'] == sev)])
        row_data.append(count)
    data4.append(row_data)

add_chart_data(ws_res, data4, data4_row, 1)

c4 = BarChart()
c4.type = "col"
c4.grouping = "stacked"
c4.title = "Close Reason by Severity (Stacked)"
c4.height = 11
c4.width = 14
c4.y_axis.title = "Count"
c4.x_axis.title = "Close Reason"
d4 = Reference(ws_res, min_col=2, min_row=data4_row, max_row=data4_row+len(data4)-1, max_col=len(severities)+1)
l4 = Reference(ws_res, min_col=1, min_row=data4_row+1, max_row=data4_row+len(data4)-1)
c4.add_data(d4, titles_from_data=True)
c4.set_categories(l4)
ws_res.add_chart(c4, "A40")

print("  ✓ Added 4 charts (Close Reason, Resolution, Root Cause, Stacked)")

# ============================================================================
# ADVANCED TIME ANALYSIS Dashboard - Priority 1
# ============================================================================
print("\n⏱️ Creating ADVANCED TIME ANALYSIS dashboard...")
if 'Time_Analysis_Advanced' in wb.sheetnames:
    del wb['Time_Analysis_Advanced']

ws_time = wb.create_sheet("Time_Analysis_Advanced", 8)

ws_time.merge_cells('A1:P1')
ws_time['A1'] = 'ADVANCED TIME ANALYSIS DASHBOARD'
ws_time['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_time['A1'].fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')
ws_time['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_time.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: Lead Time Distribution (Histogram simulation with buckets)
lead_time_buckets = ['0-24h', '24-48h', '48-72h', '3-7d', '7-14d', '14-30d', '30+d']
bucket_counts = []
for i, bucket in enumerate(lead_time_buckets):
    if i == 0:
        count = len(df_raw[df_raw['LeadTimeHrs'] <= 24])
    elif i == 1:
        count = len(df_raw[(df_raw['LeadTimeHrs'] > 24) & (df_raw['LeadTimeHrs'] <= 48)])
    elif i == 2:
        count = len(df_raw[(df_raw['LeadTimeHrs'] > 48) & (df_raw['LeadTimeHrs'] <= 72)])
    elif i == 3:
        count = len(df_raw[(df_raw['LeadTimeHrs'] > 72) & (df_raw['LeadTimeHrs'] <= 168)])
    elif i == 4:
        count = len(df_raw[(df_raw['LeadTimeHrs'] > 168) & (df_raw['LeadTimeHrs'] <= 336)])
    elif i == 5:
        count = len(df_raw[(df_raw['LeadTimeHrs'] > 336) & (df_raw['LeadTimeHrs'] <= 720)])
    else:
        count = len(df_raw[df_raw['LeadTimeHrs'] > 720])
    bucket_counts.append([bucket, count])

data1 = [['Lead Time Bucket', 'Count']] + bucket_counts
add_chart_data(ws_time, data1, chart_row, 1)

c1 = BarChart()
c1.type = "col"
c1.title = "Lead Time Distribution (Histogram)"
c1.height = 11
c1.width = 14
c1.y_axis.title = "Bug Count"
c1.x_axis.title = "Lead Time"
d1 = Reference(ws_time, min_col=2, min_row=chart_row, max_row=chart_row+len(data1)-1)
l1 = Reference(ws_time, min_col=1, min_row=chart_row+1, max_row=chart_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
c1.dataLabels = openpyxl.chart.label.DataLabelList()
c1.dataLabels.showVal = True
ws_time.add_chart(c1, "A20")

# Chart 2: Cycle Time Distribution
cycle_time_buckets = ['0-12h', '12-24h', '24-48h', '2-5d', '5-10d', '10-20d', '20+d']
cycle_counts = []
for i, bucket in enumerate(cycle_time_buckets):
    if i == 0:
        count = len(df_raw[df_raw['CycleTimeHrs'] <= 12])
    elif i == 1:
        count = len(df_raw[(df_raw['CycleTimeHrs'] > 12) & (df_raw['CycleTimeHrs'] <= 24)])
    elif i == 2:
        count = len(df_raw[(df_raw['CycleTimeHrs'] > 24) & (df_raw['CycleTimeHrs'] <= 48)])
    elif i == 3:
        count = len(df_raw[(df_raw['CycleTimeHrs'] > 48) & (df_raw['CycleTimeHrs'] <= 120)])
    elif i == 4:
        count = len(df_raw[(df_raw['CycleTimeHrs'] > 120) & (df_raw['CycleTimeHrs'] <= 240)])
    elif i == 5:
        count = len(df_raw[(df_raw['CycleTimeHrs'] > 240) & (df_raw['CycleTimeHrs'] <= 480)])
    else:
        count = len(df_raw[df_raw['CycleTimeHrs'] > 480])
    cycle_counts.append([bucket, count])

data2 = [['Cycle Time Bucket', 'Count']] + cycle_counts
add_chart_data(ws_time, data2, chart_row, 4)

c2 = BarChart()
c2.type = "col"
c2.title = "Cycle Time Distribution (Histogram)"
c2.height = 11
c2.width = 14
c2.y_axis.title = "Bug Count"
c2.x_axis.title = "Cycle Time"
d2 = Reference(ws_time, min_col=5, min_row=chart_row, max_row=chart_row+len(data2)-1)
l2 = Reference(ws_time, min_col=4, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
ws_time.add_chart(c2, "F20")

# Chart 3: Time to Close by Severity
severities = df_raw['Severity'].unique()[:5]
time_by_sev = []
for sev in severities:
    sev_data = df_raw[df_raw['Severity'] == sev]
    avg_time = sev_data['LeadTimeHrs'].mean() / 24  # Convert to days
    time_by_sev.append([sev, avg_time])

data3 = [['Severity', 'Avg Days to Close']] + time_by_sev
add_chart_data(ws_time, data3, chart_row, 7)

c3 = BarChart()
c3.title = "Average Time to Close by Severity"
c3.height = 11
c3.width = 14
c3.y_axis.title = "Days"
c3.x_axis.title = "Severity"
d3 = Reference(ws_time, min_col=8, min_row=chart_row, max_row=chart_row+len(data3)-1)
l3 = Reference(ws_time, min_col=7, min_row=chart_row+1, max_row=chart_row+len(data3)-1)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(l3)
c3.dataLabels = openpyxl.chart.label.DataLabelList()
c3.dataLabels.showVal = True
ws_time.add_chart(c3, "K20")

print("  ✓ Added 3 charts (Lead Time Histogram, Cycle Time Histogram, Time by Severity)")

# ============================================================================
# MODULE & PROJECT ANALYSIS Dashboard - Priority 2
# ============================================================================
print("\n📦 Creating MODULE & PROJECT ANALYSIS dashboard...")
if 'Module_Project' in wb.sheetnames:
    del wb['Module_Project']

ws_mod = wb.create_sheet("Module_Project", 9)

ws_mod.merge_cells('A1:P1')
ws_mod['A1'] = 'MODULE & PROJECT ANALYSIS DASHBOARD'
ws_mod['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_mod['A1'].fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
ws_mod['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_mod.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: Top 10 Modules by Bug Count
modules = df_raw['ModuleName'].value_counts().head(10).reset_index()
modules.columns = ['Module', 'Count']
data1 = [['Module', 'Bugs']] + modules.values.tolist()
add_chart_data(ws_mod, data1, chart_row, 1)

c1 = BarChart()
c1.title = "Top 10 Modules by Bug Count"
c1.height = 11
c1.width = 14
c1.y_axis.title = "Bug Count"
c1.x_axis.title = "Module"
d1 = Reference(ws_mod, min_col=2, min_row=chart_row, max_row=chart_row+len(data1)-1)
l1 = Reference(ws_mod, min_col=1, min_row=chart_row+1, max_row=chart_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
c1.dataLabels = openpyxl.chart.label.DataLabelList()
c1.dataLabels.showVal = True
ws_mod.add_chart(c1, "A20")

# Chart 2: Bugs by Project
projects = df_raw['ProjectName'].value_counts().head(8).reset_index()
projects.columns = ['Project', 'Count']
data2 = [['Project', 'Bugs']] + projects.values.tolist()
add_chart_data(ws_mod, data2, chart_row, 4)

c2 = BarChart()
c2.type = "col"
c2.title = "Bugs by Project"
c2.height = 11
c2.width = 14
c2.y_axis.title = "Bug Count"
c2.x_axis.title = "Project"
d2 = Reference(ws_mod, min_col=5, min_row=chart_row, max_row=chart_row+len(data2)-1)
l2 = Reference(ws_mod, min_col=4, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
ws_mod.add_chart(c2, "F20")

# Chart 3: Module × Severity Matrix (Stacked)
data3_row = chart_row + 15
top_modules = df_raw['ModuleName'].value_counts().head(6).index
severities = df_raw['Severity'].unique()[:4]

data3 = [['Module'] + list(severities)]
for module in top_modules:
    row_data = [module]
    for sev in severities:
        count = len(df_raw[(df_raw['ModuleName'] == module) & (df_raw['Severity'] == sev)])
        row_data.append(count)
    data3.append(row_data)

add_chart_data(ws_mod, data3, data3_row, 1)

c3 = BarChart()
c3.type = "col"
c3.grouping = "stacked"
c3.title = "Module × Severity Matrix (Stacked)"
c3.height = 11
c3.width = 14
c3.y_axis.title = "Bug Count"
c3.x_axis.title = "Module"
d3 = Reference(ws_mod, min_col=2, min_row=data3_row, max_row=data3_row+len(data3)-1, max_col=len(severities)+1)
l3 = Reference(ws_mod, min_col=1, min_row=data3_row+1, max_row=data3_row+len(data3)-1)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(l3)
ws_mod.add_chart(c3, "K20")

# Chart 4: Category Distribution
categories = df_raw['Category'].value_counts().head(8).reset_index()
categories.columns = ['Category', 'Count']
data4 = [['Category', 'Bugs']] + categories.values.tolist()
add_chart_data(ws_mod, data4, data3_row + 10, 7)

c4 = PieChart()
c4.title = "Bug Distribution by Category"
c4.height = 11
c4.width = 14
d4 = Reference(ws_mod, min_col=8, min_row=data3_row+10, max_row=data3_row+10+len(data4)-1)
l4 = Reference(ws_mod, min_col=7, min_row=data3_row+11, max_row=data3_row+10+len(data4)-1)
c4.add_data(d4, titles_from_data=True)
c4.set_categories(l4)
c4.dataLabels = openpyxl.chart.label.DataLabelList()
c4.dataLabels.showVal = True
c4.dataLabels.showPercent = True
ws_mod.add_chart(c4, "A40")

print("  ✓ Added 4 charts (Top Modules, Projects, Module×Severity, Categories)")

# ============================================================================
# WORKLOAD ANALYSIS Dashboard - Priority 2
# ============================================================================
print("\n👥 Creating WORKLOAD ANALYSIS dashboard...")
if 'Workload_Analysis' in wb.sheetnames:
    del wb['Workload_Analysis']

ws_work = wb.create_sheet("Workload_Analysis", 10)

ws_work.merge_cells('A1:P1')
ws_work['A1'] = 'WORKLOAD ANALYSIS DASHBOARD'
ws_work['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_work['A1'].fill = PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid')
ws_work['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_work.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: Team Workload Distribution
teams = df_raw['TeamName'].value_counts().head(8).reset_index()
teams.columns = ['Team', 'Bugs']
data1 = [['Team', 'Total Bugs']] + teams.values.tolist()
add_chart_data(ws_work, data1, chart_row, 1)

c1 = BarChart()
c1.type = "col"
c1.title = "Team Workload Distribution"
c1.height = 11
c1.width = 14
c1.y_axis.title = "Bug Count"
c1.x_axis.title = "Team"
d1 = Reference(ws_work, min_col=2, min_row=chart_row, max_row=chart_row+len(data1)-1)
l1 = Reference(ws_work, min_col=1, min_row=chart_row+1, max_row=chart_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
c1.dataLabels = openpyxl.chart.label.DataLabelList()
c1.dataLabels.showVal = True
ws_work.add_chart(c1, "A20")

# Chart 2: Assignee Workload (Top 10)
assignees = df_raw['AssigneeName'].value_counts().head(10).reset_index()
assignees.columns = ['Assignee', 'Bugs']
data2 = [['Assignee', 'Assigned Bugs']] + assignees.values.tolist()
add_chart_data(ws_work, data2, chart_row, 4)

c2 = BarChart()
c2.title = "Top 10 Assignees by Workload"
c2.height = 11
c2.width = 14
c2.y_axis.title = "Bug Count"
c2.x_axis.title = "Assignee"
d2 = Reference(ws_work, min_col=5, min_row=chart_row, max_row=chart_row+len(data2)-1)
l2 = Reference(ws_work, min_col=4, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
ws_work.add_chart(c2, "F20")

# Chart 3: Team × Sprint Workload (Heatmap simulation using stacked bar)
data3_row = chart_row + 15
top_teams = df_raw['TeamName'].value_counts().head(5).index
sprints = df_raw['SprintName'].unique()[:6]

data3 = [['Team'] + list(sprints)]
for team in top_teams:
    row_data = [team]
    for sprint in sprints:
        count = len(df_raw[(df_raw['TeamName'] == team) & (df_raw['SprintName'] == sprint)])
        row_data.append(count)
    data3.append(row_data)

add_chart_data(ws_work, data3, data3_row, 1)

c3 = BarChart()
c3.type = "col"
c3.grouping = "stacked"
c3.title = "Team × Sprint Workload Matrix"
c3.height = 11
c3.width = 14
c3.y_axis.title = "Bug Count"
c3.x_axis.title = "Team"
d3 = Reference(ws_work, min_col=2, min_row=data3_row, max_row=data3_row+len(data3)-1, max_col=len(sprints)+1)
l3 = Reference(ws_work, min_col=1, min_row=data3_row+1, max_row=data3_row+len(data3)-1)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(l3)
ws_work.add_chart(c3, "K20")

print("  ✓ Added 3 charts (Team Workload, Assignee Workload, Team×Sprint Matrix)")

# ============================================================================
# TREND ANALYSIS Dashboard - Priority 2
# ============================================================================
print("\n📈 Creating TREND ANALYSIS dashboard...")
if 'Trend_Analysis' in wb.sheetnames:
    del wb['Trend_Analysis']

ws_trend = wb.create_sheet("Trend_Analysis", 11)

ws_trend.merge_cells('A1:P1')
ws_trend['A1'] = 'TREND ANALYSIS DASHBOARD'
ws_trend['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_trend['A1'].fill = PatternFill(start_color='16A085', end_color='16A085', fill_type='solid')
ws_trend['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_trend.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: Bug Inflow/Outflow Trend
sprints = df_raw['SprintName'].unique()[:8]
inflow_outflow = []
for sprint in sprints:
    sprint_data = df_raw[df_raw['SprintName'] == sprint]
    inflow = len(sprint_data[sprint_data['State'].isin(['Open', 'New', 'Active'])])
    outflow = len(sprint_data[sprint_data['State'].isin(['Closed', 'Resolved', 'Done'])])
    inflow_outflow.append([sprint, inflow, outflow])

data1 = [['Sprint', 'Inflow', 'Outflow']] + inflow_outflow
add_chart_data(ws_trend, data1, chart_row, 1)

c1 = LineChart()
c1.title = "Bug Inflow vs Outflow Trend"
c1.height = 11
c1.width = 14
c1.y_axis.title = "Bug Count"
c1.x_axis.title = "Sprint"
d1 = Reference(ws_trend, min_col=2, min_row=chart_row, max_row=chart_row+len(data1)-1, max_col=3)
l1 = Reference(ws_trend, min_col=1, min_row=chart_row+1, max_row=chart_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
c1.dataLabels = openpyxl.chart.label.DataLabelList()
c1.dataLabels.showVal = True
ws_trend.add_chart(c1, "A20")

# Chart 2: Quality Metrics Trend
quality_trend = []
for sprint in sprints:
    sprint_data = df_raw[df_raw['SprintName'] == sprint]
    total = len(sprint_data)
    if total > 0:
        escape_rate = len(sprint_data[sprint_data['is_escaped'] == 1]) / total * 100
        reopen_rate = len(sprint_data[sprint_data['ReopenCount'] > 0]) / total * 100
        quality_trend.append([sprint, escape_rate, reopen_rate])
    else:
        quality_trend.append([sprint, 0, 0])

data2 = [['Sprint', 'Escape Rate %', 'Reopen Rate %']] + quality_trend
add_chart_data(ws_trend, data2, chart_row, 5)

c2 = LineChart()
c2.title = "Quality Metrics Trend (Escape & Reopen Rate)"
c2.height = 11
c2.width = 14
c2.y_axis.title = "Percentage %"
c2.x_axis.title = "Sprint"
d2 = Reference(ws_trend, min_col=6, min_row=chart_row, max_row=chart_row+len(data2)-1, max_col=7)
l2 = Reference(ws_trend, min_col=5, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
ws_trend.add_chart(c2, "F20")

# Chart 3: Average Lead Time Trend
lead_time_trend = []
for sprint in sprints:
    sprint_data = df_raw[df_raw['SprintName'] == sprint]
    if len(sprint_data) > 0:
        avg_lead = sprint_data['LeadTimeHrs'].mean() / 24  # Convert to days
        lead_time_trend.append([sprint, avg_lead])
    else:
        lead_time_trend.append([sprint, 0])

data3 = [['Sprint', 'Avg Lead Time (days)']] + lead_time_trend
add_chart_data(ws_trend, data3, chart_row, 9)

c3 = LineChart()
c3.title = "Average Lead Time Trend"
c3.height = 11
c3.width = 14
c3.y_axis.title = "Days"
c3.x_axis.title = "Sprint"
d3 = Reference(ws_trend, min_col=10, min_row=chart_row, max_row=chart_row+len(data3)-1)
l3 = Reference(ws_trend, min_col=9, min_row=chart_row+1, max_row=chart_row+len(data3)-1)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(l3)
c3.dataLabels = openpyxl.chart.label.DataLabelList()
c3.dataLabels.showVal = True
ws_trend.add_chart(c3, "K20")

# Chart 4: Cumulative Bug Trend (Area Chart)
cumulative = []
cumulative_sum = 0
for sprint in sprints:
    sprint_bugs = len(df_raw[df_raw['SprintName'] == sprint])
    cumulative_sum += sprint_bugs
    cumulative.append([sprint, cumulative_sum])

data4 = [['Sprint', 'Cumulative Bugs']] + cumulative
add_chart_data(ws_trend, data4, chart_row + 15, 1)

c4 = AreaChart()
c4.title = "Cumulative Bugs Over Time"
c4.height = 11
c4.width = 14
c4.y_axis.title = "Total Bugs"
c4.x_axis.title = "Sprint"
d4 = Reference(ws_trend, min_col=2, min_row=chart_row+15, max_row=chart_row+15+len(data4)-1)
l4 = Reference(ws_trend, min_col=1, min_row=chart_row+16, max_row=chart_row+15+len(data4)-1)
c4.add_data(d4, titles_from_data=True)
c4.set_categories(l4)
c4.dataLabels = openpyxl.chart.label.DataLabelList()
c4.dataLabels.showVal = True
ws_trend.add_chart(c4, "A40")

print("  ✓ Added 4 charts (Inflow/Outflow, Quality Metrics, Lead Time, Cumulative)")

# ============================================================================
# ROOT CAUSE & SPECIALTY Dashboard - Priority 3
# ============================================================================
print("\n🔍 Creating ROOT CAUSE & SPECIALTY dashboard...")
if 'RootCause_Specialty' in wb.sheetnames:
    del wb['RootCause_Specialty']

ws_root = wb.create_sheet("RootCause_Specialty", 12)

ws_root.merge_cells('A1:P1')
ws_root['A1'] = 'ROOT CAUSE & SPECIALTY ANALYSIS'
ws_root['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_root['A1'].fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
ws_root['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_root.row_dimensions[1].height = 30

chart_row = 5

# Chart 1: Root Cause × Severity Matrix
data1_row = chart_row
root_causes = df_raw['RootCause'].value_counts().head(6).index
severities = df_raw['Severity'].unique()[:4]

data1 = [['Root Cause'] + list(severities)]
for cause in root_causes:
    row_data = [cause]
    for sev in severities:
        count = len(df_raw[(df_raw['RootCause'] == cause) & (df_raw['Severity'] == sev)])
        row_data.append(count)
    data1.append(row_data)

add_chart_data(ws_root, data1, data1_row, 1)

c1 = BarChart()
c1.type = "col"
c1.grouping = "stacked"
c1.title = "Root Cause × Severity Matrix"
c1.height = 11
c1.width = 14
c1.y_axis.title = "Bug Count"
c1.x_axis.title = "Root Cause"
d1 = Reference(ws_root, min_col=2, min_row=data1_row, max_row=data1_row+len(data1)-1, max_col=len(severities)+1)
l1 = Reference(ws_root, min_col=1, min_row=data1_row+1, max_row=data1_row+len(data1)-1)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(l1)
ws_root.add_chart(c1, "A20")

# Chart 2: Bugs by Reporter (Top 10)
reporters = df_raw['ReporterName'].value_counts().head(10).reset_index()
reporters.columns = ['Reporter', 'Bugs']
data2 = [['Reporter', 'Bugs Reported']] + reporters.values.tolist()
add_chart_data(ws_root, data2, chart_row, 6)

c2 = BarChart()
c2.title = "Top 10 Bug Reporters"
c2.height = 11
c2.width = 14
c2.y_axis.title = "Bug Count"
c2.x_axis.title = "Reporter"
d2 = Reference(ws_root, min_col=7, min_row=chart_row, max_row=chart_row+len(data2)-1)
l2 = Reference(ws_root, min_col=6, min_row=chart_row+1, max_row=chart_row+len(data2)-1)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(l2)
c2.dataLabels = openpyxl.chart.label.DataLabelList()
c2.dataLabels.showVal = True
ws_root.add_chart(c2, "F20")

print("  ✓ Added 2 charts (Root Cause×Severity, Top Reporters)")

# ============================================================================
# FINAL SAVE & VALIDATION
# ============================================================================
print("\n" + "=" * 80)
print("💾 SAVING UPDATED FILE...")
print("=" * 80)

wb.save('BugTracking_Complete.xlsx')
print("\n✅ File saved successfully: BugTracking_Complete.xlsx")

# Count total charts
total_charts = 0
for sheet in wb.sheetnames:
    ws = wb[sheet]
    chart_count = len(ws._charts)
    if chart_count > 0:
        print(f"  📊 {sheet}: {chart_count} charts")
        total_charts += chart_count

print("\n" + "=" * 80)
print(f"🎉 TOTAL CHARTS ADDED: {total_charts}")
print("=" * 80)

print("\n✅ Dashboard Summary:")
print("  • PowerBI_Dashboard: 6 charts + 12 filters")
print("  • Volume_Analysis: 5 charts")
print("  • Team_Performance: 3 charts")
print("  • Sprint_Analysis: 1 chart")
print("  • Time_Flow: 2 charts")
print("  • Quality_Analysis: 3 charts")
print("  • State_Flow: 3 charts (NEW)")
print("  • Resolution_Analysis: 4 charts (NEW)")
print("  • Time_Analysis_Advanced: 3 charts (NEW)")
print("  • Module_Project: 4 charts (NEW)")
print("  • Workload_Analysis: 3 charts (NEW)")
print("  • Trend_Analysis: 4 charts (NEW)")
print("  • RootCause_Specialty: 2 charts (NEW)")

print("\n" + "=" * 80)
print("✅ ALL CHARTS IMPLEMENTATION COMPLETE!")
print("=" * 80)

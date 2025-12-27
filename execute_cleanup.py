#!/usr/bin/env python3
"""
Execute Cleanup Plan
"""

from openpyxl import load_workbook
import os

print("=" * 80)
print("اجرای پلن پاک‌سازی")
print("=" * 80)

# Load file
file_path = 'BugTracking_Complete_FINAL.xlsx'
wb = load_workbook(file_path)

# ============================================================================
# STEP 1: Delete unnecessary sheets
# ============================================================================

print("\n📋 STEP 1: حذف شیت‌های اضافی...")

sheets_to_delete = ['RootCause_Specialty', 'Resolution_Analysis', 'Time_Analysis_Advanced']

for sheet_name in sheets_to_delete:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"   ✅ حذف شد: {sheet_name}")
    else:
        print(f"   ℹ️  موجود نیست: {sheet_name}")

# ============================================================================
# STEP 2: Delete columns from raw_data
# ============================================================================

print("\n📊 STEP 2: حذف فیلدها از raw_data...")

fields_to_delete = [
    'VerifierName', 'VerifierID',
    'DevEffortHrs', 'FixEffortHrs', 'TestEffortHrs', 'ReopenEffortHrs',
    'TotalEffortHrs', 'EstimatedEffortHrs',
    'LeadTimeHrs', 'CycleTimeHrs', 'ResponseTimeHrs', 'WaitTimeHrs',
    'ActiveWorkTimeHrs', 'AgeDays',
    'RootCause', 'Resolution', 'CloseReason', 'TestCaseID',
    'IsDuplicate', 'DuplicateOfBugID', 'RetestPassCount', 'RetestFailCount'
]

ws_raw = wb['raw_data']

# Find columns to delete
cols_to_delete = []
for col in range(1, ws_raw.max_column + 1):
    field_name = ws_raw.cell(1, col).value
    if field_name in fields_to_delete:
        cols_to_delete.append((col, field_name))

print(f"   یافت شد: {len(cols_to_delete)} فیلد برای حذف")

# Delete columns (reverse order to maintain indices)
for col, field_name in reversed(cols_to_delete):
    ws_raw.delete_cols(col, 1)
    print(f"   ✅ حذف: {field_name} (ستون {col})")

print(f"   📊 فیلدهای باقیمانده: {ws_raw.max_column}")

# ============================================================================
# STEP 3: Delete problematic charts
# ============================================================================

print("\n📈 STEP 3: حذف چارت‌های مشکوک...")

charts_deleted = 0

# Quality_Analysis - remove Escaped chart
if 'Quality_Analysis' in wb.sheetnames:
    ws = wb['Quality_Analysis']
    charts_to_remove = []
    
    if hasattr(ws, '_charts'):
        for i, chart in enumerate(ws._charts):
            chart_title = ""
            try:
                if hasattr(chart, 'title') and chart.title:
                    if hasattr(chart.title, 'tx') and chart.title.tx:
                        if hasattr(chart.title.tx, 'rich'):
                            for para in chart.title.tx.rich.p:
                                for run in para.r:
                                    if hasattr(run, 't'):
                                        chart_title = run.t
                                        break
            except:
                pass
            
            if 'escape' in chart_title.lower():
                charts_to_remove.append(i)
                print(f"   ❌ Quality_Analysis: '{chart_title}'")
        
        # Remove charts (reverse order)
        for i in reversed(charts_to_remove):
            del ws._charts[i]
            charts_deleted += 1

# Trend_Analysis - remove Escape chart
if 'Trend_Analysis' in wb.sheetnames:
    ws = wb['Trend_Analysis']
    charts_to_remove = []
    
    if hasattr(ws, '_charts'):
        for i, chart in enumerate(ws._charts):
            chart_title = ""
            try:
                if hasattr(chart, 'title') and chart.title:
                    if hasattr(chart.title, 'tx') and chart.title.tx:
                        if hasattr(chart.title.tx, 'rich'):
                            for para in chart.title.tx.rich.p:
                                for run in para.r:
                                    if hasattr(run, 't'):
                                        chart_title = run.t
                                        break
            except:
                pass
            
            if 'escape' in chart_title.lower():
                charts_to_remove.append(i)
                print(f"   ❌ Trend_Analysis: '{chart_title}'")
        
        for i in reversed(charts_to_remove):
            del ws._charts[i]
            charts_deleted += 1

print(f"   📊 چارت‌های حذف‌شده: {charts_deleted}")

# ============================================================================
# STEP 4: Clean up KPIs_Detail and metrics sheets
# ============================================================================

print("\n🔧 STEP 4: پاک‌سازی KPI ها...")

for sheet_name in ['KPIs_Detail', 'metrics']:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    rows_to_delete = []
    
    # Find rows with deleted field references
    for row_idx in range(1, min(ws.max_row + 1, 100)):
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                
                # Check if references deleted fields
                for field in fields_to_delete:
                    if field in formula:
                        if row_idx not in rows_to_delete:
                            rows_to_delete.append(row_idx)
                        break
    
    # Delete rows (reverse order)
    for row_idx in reversed(sorted(set(rows_to_delete))):
        ws.delete_rows(row_idx, 1)
        print(f"   ✅ {sheet_name}: حذف ردیف {row_idx}")

# ============================================================================
# STEP 5: Update راهنمای_فیلدها
# ============================================================================

print("\n📋 STEP 5: به‌روزرسانی راهنمای_فیلدها...")

if 'راهنمای_فیلدها' in wb.sheetnames:
    ws_guide = wb['راهنمای_فیلدها']
    
    rows_to_delete = []
    
    # Find rows for deleted fields (column A has field names)
    for row_idx in range(4, ws_guide.max_row + 1):  # Start from row 4 (data starts there)
        field_name = ws_guide.cell(row_idx, 1).value
        if field_name in fields_to_delete:
            rows_to_delete.append(row_idx)
    
    # Delete rows (reverse order)
    for row_idx in reversed(rows_to_delete):
        ws_guide.delete_rows(row_idx, 1)
        print(f"   ✅ حذف ردیف: {ws_guide.cell(row_idx, 1).value}")
    
    # Update summary
    summary_row = ws_guide.max_row
    ws_guide[f'A{summary_row}'] = f'''خلاصه:
🟢 فیلدهای سبز: مستقیماً از CSV موجود است
🟡 فیلدهای زرد: نیاز به کوئری WorkItemRevisions دارد
🟠 فیلدهای نارنجی: قابل محاسبه از داده‌های موجود
🔵 فیلدهای آبی: نیاز به ورود دستی یا کوئری‌های اضافی

مجموع فیلدها: {ws_raw.max_column} فیلد'''
    
    print(f"   ✅ خلاصه به‌روز شد: {ws_raw.max_column} فیلد")

# ============================================================================
# Save
# ============================================================================

print("\n💾 ذخیره فایل...")

wb.save('BugTracking_Complete_FINAL.xlsx')

size_kb = os.path.getsize('BugTracking_Complete_FINAL.xlsx') / 1024

print(f"   ✅ ذخیره شد")
print(f"   📁 حجم: {size_kb:.1f} KB")

# ============================================================================
# Summary
# ============================================================================

print("\n" + "=" * 80)
print("✅ پاک‌سازی کامل شد!")
print("=" * 80)
print(f"""
📊 تغییرات:
   - شیت‌های حذف‌شده: 3 (RootCause_Specialty, Resolution_Analysis, Time_Analysis_Advanced)
   - فیلدهای حذف‌شده: {len(cols_to_delete)}
   - فیلدهای باقیمانده: {ws_raw.max_column}
   - چارت‌های حذف‌شده: {charts_deleted}
   - KPI های پاک‌سازی شده: ✓
   - راهنمای فیلدها: به‌روز شد

🎯 فایل تمیز و آماده است!
""")
print("=" * 80)

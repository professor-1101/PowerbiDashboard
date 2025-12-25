#!/usr/bin/env python3
"""
Fix Chart Overlaps - Reposition charts to avoid overlaps
"""

from openpyxl import load_workbook
from openpyxl.chart.chartspace import ChartSpace

print("=" * 80)
print("🔧 FIXING CHART OVERLAPS")
print("=" * 80)

# Load the workbook
wb = load_workbook('BugTracking_Complete.xlsx')

# ============================================================================
# Check PowerBI_Dashboard chart positions
# ============================================================================
print("\n📊 Inspecting PowerBI_Dashboard...")
ws_main = wb['PowerBI_Dashboard']

print(f"  Charts found: {len(ws_main._charts)}")
for i, chart in enumerate(ws_main._charts, 1):
    if hasattr(chart.anchor, '_from'):
        col = chart.anchor._from.col
        row = chart.anchor._from.row
        print(f"  Chart {i} ({chart.title}): Column {col}, Row {row}")

# The issue is that charts are too close horizontally
# Charts need to be repositioned with proper spacing

# ============================================================================
# Fix: Reposition PowerBI_Dashboard charts
# ============================================================================
print("\n🔧 Repositioning charts to avoid overlap...")

# Current positions cause overlap because charts are ~6 columns wide
# Position them at columns: A (0), H (7), O (14) instead of A (0), F (5), K (10)

# New positions with proper spacing (15 columns between charts)
new_positions = [
    "A31",   # Chart 1: Column A, Row 31
    "A47",   # Chart 2: Column A, Row 47 (16 rows below)
    "H31",   # Chart 3: Column H, Row 31
    "H47",   # Chart 4: Column H, Row 47
    "O31",   # Chart 5: Column O, Row 31
    "O47",   # Chart 6: Column O, Row 47
]

# Apply new positions if we have exactly 6 charts
if len(ws_main._charts) == 6:
    for i, chart in enumerate(ws_main._charts):
        old_pos = chart.anchor if hasattr(chart, 'anchor') else "Unknown"
        chart.anchor = new_positions[i]
        print(f"  ✅ Chart {i+1} repositioned to {new_positions[i]}")
else:
    print(f"  ⚠️  Expected 6 charts, found {len(ws_main._charts)}")

# ============================================================================
# Check all other dashboards for overlaps
# ============================================================================
sheets_to_check = [
    'Volume_Analysis',
    'Team_Performance',
    'Sprint_Analysis',
    'Time_Flow',
    'Quality_Analysis',
    'State_Flow',
    'Resolution_Analysis',
    'Time_Analysis_Advanced',
    'Module_Project',
    'Workload_Analysis',
    'Trend_Analysis',
    'RootCause_Specialty'
]

print("\n📊 Checking other dashboards...")
for sheet_name in sheets_to_check:
    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    if len(ws._charts) == 0:
        continue

    print(f"\n  {sheet_name}: {len(ws._charts)} charts")

    # Collect current positions
    positions = []
    for chart in ws._charts:
        if hasattr(chart.anchor, '_from'):
            col = chart.anchor._from.col
            row = chart.anchor._from.row
            positions.append((col, row, chart.title or "Untitled"))

    # Check for overlaps in this sheet
    for i, (col1, row1, title1) in enumerate(positions):
        for j, (col2, row2, title2) in enumerate(positions[i+1:], i+1):
            # Charts overlap if too close
            col_diff = abs(col1 - col2)
            row_diff = abs(row1 - row2)

            if col_diff < 6 and row_diff < 15:
                print(f"    ⚠️  Potential overlap: '{title1}' at ({col1},{row1}) and '{title2}' at ({col2},{row2})")

# ============================================================================
# Save the fixed file
# ============================================================================
print("\n" + "=" * 80)
print("💾 SAVING FIXED FILE...")
print("=" * 80)

wb.save('BugTracking_Complete.xlsx')
print("\n✅ File saved with fixed chart positions!")

print("\n" + "=" * 80)
print("✅ CHART OVERLAP FIX COMPLETE")
print("=" * 80)
